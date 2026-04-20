# ============================================================================
# BULK IMPORT SERVICE
# ============================================================================
"""
Service for handling Excel bulk imports with validation and preview.
"""

import openpyxl
from io import BytesIO
from django.db import transaction
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import InMemoryUploadedFile
from typing import Optional, Dict, List, Any, Tuple
import re
from datetime import datetime

from workforce.core_models import Employee
from workforce.models import (
    BulkImportSession,
    BulkImportRecord,
    Department,
    JobGrade,
    JobTitle,
    Campus,
    Position,
)


class BulkImportService:
    """
    Service for bulk importing data from Excel files.
    Supports staged validation before actual import.
    """
    
    # Field mappings for employee import
    EMPLOYEE_FIELD_MAPPING = {
        'employee_no': {'required': True, 'type': 'string'},
        'first_name': {'required': True, 'type': 'string'},
        'middle_name': {'required': False, 'type': 'string'},
        'last_name': {'required': True, 'type': 'string'},
        'date_of_birth': {'required': True, 'type': 'date'},
        'gender': {'required': True, 'type': 'choice', 'choices': ['male', 'female', 'other']},
        'national_id': {'required': True, 'type': 'string'},
        'personal_email': {'required': True, 'type': 'email'},
        'official_email': {'required': True, 'type': 'email'},
        'phone_primary': {'required': True, 'type': 'string'},
        'phone_secondary': {'required': False, 'type': 'string'},
        'employee_category': {'required': True, 'type': 'choice', 
                             'choices': ['teaching', 'non_teaching', 'contract', 'casual', 'visiting']},
        'payroll_type': {'required': True, 'type': 'choice',
                        'choices': ['monthly', 'hourly', 'contract', 'daily']},
        'hire_date': {'required': True, 'type': 'date'},
        'department_code': {'required': True, 'type': 'lookup', 'model': 'Department'},
        'job_grade_code': {'required': False, 'type': 'lookup', 'model': 'JobGrade'},
    }
    
    DEPARTMENT_FIELD_MAPPING = {
        'code': {'required': True, 'type': 'string'},
        'name': {'required': True, 'type': 'string'},
        'department_type': {'required': True, 'type': 'choice',
                           'choices': ['academic', 'administrative', 'support']},
        'parent_department_code': {'required': False, 'type': 'lookup', 'model': 'Department'},
        'campus_code': {'required': True, 'type': 'lookup', 'model': 'Campus'},
    }
    
    POSITION_FIELD_MAPPING = {
        'position_code': {'required': True, 'type': 'string'},
        'title': {'required': True, 'type': 'string'},
        'job_title_code': {'required': True, 'type': 'lookup', 'model': 'JobTitle'},
        'department_code': {'required': True, 'type': 'lookup', 'model': 'Department'},
        'campus_code': {'required': True, 'type': 'lookup', 'model': 'Campus'},
        'position_type': {'required': True, 'type': 'choice',
                         'choices': ['permanent', 'contract', 'temporary', 'grant_funded']},
        'budgeted_salary': {'required': False, 'type': 'decimal'},
        'is_critical': {'required': False, 'type': 'boolean'},
    }
    
    def __init__(self, user):
        """Initialize with the current user."""
        self.user = user
    
    def _get_field_mapping(self, import_type: str) -> Dict:
        """Get field mapping for the import type."""
        mappings = {
            'employees': self.EMPLOYEE_FIELD_MAPPING,
            'departments': self.DEPARTMENT_FIELD_MAPPING,
            'positions': self.POSITION_FIELD_MAPPING,
        }
        return mappings.get(import_type, {})
    
    @transaction.atomic
    def create_import_session(
        self,
        uploaded_file: InMemoryUploadedFile,
        import_type: str,
        skip_duplicates: bool = True,
        update_existing: bool = False,
    ) -> BulkImportSession:
        """
        Create a new import session from an uploaded file.
        
        Args:
            uploaded_file: The uploaded Excel file
            import_type: Type of import (employees, departments, etc.)
            skip_duplicates: Skip rows with duplicate identifiers
            update_existing: Update existing records instead of skipping
            
        Returns:
            BulkImportSession: The created session
        """
        session = BulkImportSession.objects.create(
            import_type=import_type,
            status=BulkImportSession.Status.UPLOADED,
            original_filename=uploaded_file.name,
            uploaded_file=uploaded_file,
            file_size_bytes=uploaded_file.size,
            uploaded_by=self.user,
            skip_duplicates=skip_duplicates,
            update_existing=update_existing,
        )
        
        # Parse and store raw data
        self._parse_excel(session)
        
        return session
    
    def _parse_excel(self, session: BulkImportSession) -> None:
        """
        Parse Excel file and create BulkImportRecord entries.
        """
        try:
            workbook = openpyxl.load_workbook(session.uploaded_file)
            sheet = workbook.active
            
            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    # Normalize header names
                    header = str(cell.value).strip().lower().replace(' ', '_')
                    headers.append(header)
                else:
                    break
            
            # Process data rows
            row_count = 0
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                # Check if row is empty
                if all(cell.value is None for cell in row[:len(headers)]):
                    continue
                
                row_data = {}
                for col_idx, cell in enumerate(row[:len(headers)]):
                    header = headers[col_idx]
                    value = cell.value
                    
                    # Convert dates to string for JSON storage
                    if isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d')
                    
                    row_data[header] = value
                
                BulkImportRecord.objects.create(
                    import_session=session,
                    row_number=row_idx,
                    raw_data=row_data,
                    status=BulkImportRecord.Status.PENDING,
                )
                row_count += 1
            
            session.total_rows = row_count
            session.save()
            
        except Exception as e:
            session.status = BulkImportSession.Status.FAILED
            session.error_summary = f"Failed to parse Excel file: {str(e)}"
            session.save()
            raise ValidationError(f"Failed to parse Excel file: {str(e)}")
    
    @transaction.atomic
    def validate_import(self, session: BulkImportSession) -> BulkImportSession:
        """
        Validate all records in an import session.
        
        Args:
            session: The import session to validate
            
        Returns:
            BulkImportSession: The updated session with validation results
        """
        session.status = BulkImportSession.Status.VALIDATING
        session.validation_started_at = timezone.now()
        session.save()
        
        field_mapping = self._get_field_mapping(session.import_type)
        valid_count = 0
        error_count = 0
        warning_count = 0
        
        for record in session.records.all():
            errors, warnings, processed_data = self._validate_record(
                record.raw_data,
                field_mapping,
                session.import_type,
                session.skip_duplicates,
            )
            
            record.validation_errors = errors
            record.validation_warnings = warnings
            record.processed_data = processed_data
            
            if errors:
                record.status = BulkImportRecord.Status.INVALID
                error_count += 1
            elif warnings:
                record.status = BulkImportRecord.Status.WARNING
                warning_count += 1
                valid_count += 1
            else:
                record.status = BulkImportRecord.Status.VALID
                valid_count += 1
            
            record.save()
        
        session.valid_rows = valid_count
        session.error_rows = error_count
        session.warning_rows = warning_count
        session.validation_completed_at = timezone.now()
        
        if error_count > 0:
            session.status = BulkImportSession.Status.VALIDATION_FAILED
        else:
            session.status = BulkImportSession.Status.VALIDATED
        
        session.save()
        return session
    
    def _validate_record(
        self,
        raw_data: Dict,
        field_mapping: Dict,
        import_type: str,
        skip_duplicates: bool,
    ) -> Tuple[List[str], List[str], Dict]:
        """
        Validate a single record.
        
        Returns:
            Tuple of (errors, warnings, processed_data)
        """
        errors = []
        warnings = []
        processed_data = {}
        
        for field_name, rules in field_mapping.items():
            value = raw_data.get(field_name)
            
            # Check required fields
            if rules['required'] and (value is None or str(value).strip() == ''):
                errors.append(f"Field '{field_name}' is required")
                continue
            
            if value is None or str(value).strip() == '':
                continue
            
            # Type validation
            field_type = rules['type']
            
            if field_type == 'string':
                processed_data[field_name] = str(value).strip()
                
            elif field_type == 'email':
                email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                if not re.match(email_pattern, str(value)):
                    errors.append(f"Invalid email format for '{field_name}': {value}")
                else:
                    processed_data[field_name] = str(value).strip().lower()
                    
            elif field_type == 'date':
                try:
                    if isinstance(value, datetime):
                        processed_data[field_name] = value.strftime('%Y-%m-%d')
                    elif isinstance(value, str):
                        # Try common date formats
                        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y']:
                            try:
                                parsed = datetime.strptime(value, fmt)
                                processed_data[field_name] = parsed.strftime('%Y-%m-%d')
                                break
                            except ValueError:
                                continue
                        else:
                            errors.append(f"Invalid date format for '{field_name}': {value}")
                except Exception:
                    errors.append(f"Invalid date for '{field_name}': {value}")
                    
            elif field_type == 'choice':
                choices = rules.get('choices', [])
                str_value = str(value).strip().lower()
                if str_value not in choices:
                    errors.append(f"Invalid choice for '{field_name}': {value}. Valid options: {choices}")
                else:
                    processed_data[field_name] = str_value
                    
            elif field_type == 'decimal':
                try:
                    processed_data[field_name] = float(value)
                except (ValueError, TypeError):
                    errors.append(f"Invalid number for '{field_name}': {value}")
                    
            elif field_type == 'boolean':
                str_value = str(value).strip().lower()
                if str_value in ['true', 'yes', '1', 'y']:
                    processed_data[field_name] = True
                elif str_value in ['false', 'no', '0', 'n']:
                    processed_data[field_name] = False
                else:
                    errors.append(f"Invalid boolean for '{field_name}': {value}")
                    
            elif field_type == 'lookup':
                model_name = rules.get('model')
                lookup_result = self._validate_lookup(model_name, value)
                if lookup_result['found']:
                    processed_data[f"{field_name}_id"] = lookup_result['id']
                    processed_data[field_name] = value
                else:
                    if rules['required']:
                        errors.append(f"{model_name} not found for '{field_name}': {value}")
                    else:
                        warnings.append(f"{model_name} not found for '{field_name}': {value}")
        
        # Check for duplicates
        if skip_duplicates and import_type == 'employees':
            emp_no = processed_data.get('employee_no')
            if emp_no and Employee.objects.filter(employee_no=emp_no).exists():
                errors.append(f"Employee with employee_no '{emp_no}' already exists")
        
        return errors, warnings, processed_data
    
    def _validate_lookup(self, model_name: str, value: Any) -> Dict:
        """
        Validate a lookup field value.
        """
        result = {'found': False, 'id': None}
        
        if not value:
            return result
        
        value = str(value).strip()
        
        try:
            if model_name == 'Department':
                dept = Department.objects.filter(
                    code__iexact=value
                ).first() or Department.objects.filter(
                    name__iexact=value
                ).first()
                if dept:
                    result = {'found': True, 'id': dept.id}
                    
            elif model_name == 'JobGrade':
                grade = JobGrade.objects.filter(
                    code__iexact=value
                ).first() or JobGrade.objects.filter(
                    name__iexact=value
                ).first()
                if grade:
                    result = {'found': True, 'id': grade.id}
                    
            elif model_name == 'JobTitle':
                title = JobTitle.objects.filter(
                    code__iexact=value
                ).first() or JobTitle.objects.filter(
                    title__iexact=value
                ).first()
                if title:
                    result = {'found': True, 'id': title.id}
                    
            elif model_name == 'Campus':
                campus = Campus.objects.filter(
                    code__iexact=value
                ).first() or Campus.objects.filter(
                    name__iexact=value
                ).first()
                if campus:
                    result = {'found': True, 'id': campus.id}
                    
        except Exception:
            pass
        
        return result
    
    def get_preview(self, session: BulkImportSession, limit: int = 100) -> Dict:
        """
        Get a preview of validated data.
        
        Args:
            session: The import session
            limit: Maximum records to return
            
        Returns:
            Dict with preview data
        """
        records = session.records.all()[:limit]
        
        return {
            'session_id': session.id,
            'import_type': session.import_type,
            'status': session.status,
            'total_rows': session.total_rows,
            'valid_rows': session.valid_rows,
            'error_rows': session.error_rows,
            'warning_rows': session.warning_rows,
            'records': [
                {
                    'row_number': r.row_number,
                    'status': r.status,
                    'raw_data': r.raw_data,
                    'processed_data': r.processed_data,
                    'errors': r.validation_errors,
                    'warnings': r.validation_warnings,
                }
                for r in records
            ]
        }
    
    @transaction.atomic
    def execute_import(self, session: BulkImportSession) -> BulkImportSession:
        """
        Execute the actual import after validation.
        
        Args:
            session: The validated import session
            
        Returns:
            BulkImportSession: The updated session with import results
        """
        if session.status not in [
            BulkImportSession.Status.VALIDATED,
            BulkImportSession.Status.PREVIEW
        ]:
            raise ValidationError("Session must be validated before import")
        
        session.status = BulkImportSession.Status.IMPORTING
        session.import_started_at = timezone.now()
        session.save()
        
        imported_count = 0
        skipped_count = 0
        
        try:
            for record in session.records.filter(
                status__in=[
                    BulkImportRecord.Status.VALID,
                    BulkImportRecord.Status.WARNING
                ]
            ):
                try:
                    created_obj = self._import_record(session.import_type, record)
                    record.status = BulkImportRecord.Status.IMPORTED
                    
                    if session.import_type == 'employees':
                        record.created_employee_id = created_obj.id
                    record.created_object_id = created_obj.id
                    record.created_object_type = session.import_type
                    
                    imported_count += 1
                except Exception as e:
                    record.status = BulkImportRecord.Status.FAILED
                    record.error_message = str(e)
                    skipped_count += 1
                
                record.save()
            
            session.imported_rows = imported_count
            session.skipped_rows = skipped_count
            session.import_completed_at = timezone.now()
            
            if skipped_count > 0:
                session.status = BulkImportSession.Status.PARTIALLY_COMPLETED
            else:
                session.status = BulkImportSession.Status.COMPLETED
                
        except Exception as e:
            session.status = BulkImportSession.Status.FAILED
            session.error_summary = str(e)
        
        session.save()
        return session
    
    def _import_record(self, import_type: str, record: BulkImportRecord) -> Any:
        """
        Import a single record.
        """
        data = record.processed_data
        
        if import_type == 'employees':
            return self._import_employee(data)
        elif import_type == 'departments':
            return self._import_department(data)
        elif import_type == 'positions':
            return self._import_position(data)
        else:
            raise ValidationError(f"Unknown import type: {import_type}")
    
    def _import_employee(self, data: Dict) -> Employee:
        """Import a single employee record."""
        from datetime import datetime as dt
        
        employee = Employee.objects.create(
            employee_no=data['employee_no'],
            first_name=data['first_name'],
            middle_name=data.get('middle_name', ''),
            last_name=data['last_name'],
            date_of_birth=dt.strptime(data['date_of_birth'], '%Y-%m-%d').date(),
            gender=data['gender'],
            national_id=data['national_id'],
            personal_email=data['personal_email'],
            official_email=data['official_email'],
            phone_primary=data['phone_primary'],
            phone_secondary=data.get('phone_secondary', ''),
            employee_category=data['employee_category'],
            payroll_type=data['payroll_type'],
            hire_date=dt.strptime(data['hire_date'], '%Y-%m-%d').date(),
            employment_status=Employee.EmploymentStatus.PROBATION,
            department_id=data.get('department_code_id'),
            job_grade_id=data.get('job_grade_code_id'),
        )
        return employee
    
    def _import_department(self, data: Dict) -> Department:
        """Import a single department record."""
        department = Department.objects.create(
            code=data['code'],
            name=data['name'],
            department_type=data['department_type'],
            parent_department_id=data.get('parent_department_code_id'),
            campus_id=data.get('campus_code_id'),
        )
        return department
    
    def _import_position(self, data: Dict) -> Position:
        """Import a single position record."""
        from django.utils import timezone
        
        position = Position.objects.create(
            position_code=data['position_code'],
            title=data['title'],
            job_title_id=data.get('job_title_code_id'),
            department_id=data.get('department_code_id'),
            campus_id=data.get('campus_code_id'),
            position_type=data['position_type'],
            budgeted_salary=data.get('budgeted_salary'),
            is_critical=data.get('is_critical', False),
            effective_from=timezone.now().date(),
        )
        return position
    
    def generate_template(self, import_type: str) -> BytesIO:
        """
        Generate an Excel template for the given import type.
        
        Args:
            import_type: Type of import
            
        Returns:
            BytesIO: Excel file as bytes
        """
        field_mapping = self._get_field_mapping(import_type)
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"{import_type.title()} Import"
        
        # Create headers
        headers = list(field_mapping.keys())
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header.replace('_', ' ').title()
            
            # Add comment with field info
            rules = field_mapping[header]
            comment_text = f"{'Required' if rules['required'] else 'Optional'}\n"
            comment_text += f"Type: {rules['type']}"
            if rules['type'] == 'choice':
                comment_text += f"\nChoices: {', '.join(rules['choices'])}"
        
        # Add sample row
        sample_data = self._get_sample_data(import_type)
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=2, column=col_idx).value = sample_data.get(header, '')
        
        # Auto-adjust column widths
        for col in sheet.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = max(max_length + 2, 12)
        
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output
    
    def _get_sample_data(self, import_type: str) -> Dict:
        """Get sample data for template."""
        samples = {
            'employees': {
                'employee_no': 'EMP001',
                'first_name': 'John',
                'middle_name': 'M',
                'last_name': 'Doe',
                'date_of_birth': '1990-01-15',
                'gender': 'male',
                'national_id': '12345678',
                'personal_email': 'john.doe@email.com',
                'official_email': 'john.doe@company.com',
                'phone_primary': '+1234567890',
                'employee_category': 'teaching',
                'payroll_type': 'monthly',
                'hire_date': '2024-01-01',
                'department_code': 'ACAD',
            },
            'departments': {
                'code': 'DEPT001',
                'name': 'Academic Affairs',
                'department_type': 'academic',
                'campus_code': 'MAIN',
            },
            'positions': {
                'position_code': 'POS001',
                'title': 'Senior Lecturer',
                'job_title_code': 'SLEC',
                'department_code': 'ACAD',
                'campus_code': 'MAIN',
                'position_type': 'permanent',
            }
        }
        return samples.get(import_type, {})
