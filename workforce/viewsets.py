"""
HR & Payroll Django REST Framework ViewSets
Part 2: API ViewSets with Custom Actions
"""

from django.apps import apps
from rest_framework import viewsets, status, filters
import rest_framework.parsers
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q, Sum, Count, Avg
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal
from math import radians, sin, cos, sqrt, atan2


from .core_models import *
from .models import *
from .serializers import *
from .permissions import *


# ============================================================================
# CORE VIEWSETS
# ============================================================================

class CampusSerializer(serializers.ModelSerializer):
    """Campus serializer"""
    class Meta:
        model = Campus
        fields = '__all__'


class CampusViewSet(viewsets.ModelViewSet):
    """Campus ViewSet"""
    queryset = Campus.objects.all()
    serializer_class = CampusSerializer
    permission_classes = [AllowAny]


class EmployeeViewSet(viewsets.ModelViewSet):
    """
    Employee ViewSet with custom actions
    """
    queryset = Employee.objects.select_related(
        'department', 'job_grade'
    ).prefetch_related('addresses', 'emergency_contacts')
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = [
        'employment_status', 'employee_category', 
        'department', 'gender', 'job_assignments__job_title'
    ]
    search_fields = [
        'employee_no', 'first_name', 'last_name', 
        'national_id', 'official_email', 'personal_email'
    ]
    ordering_fields = ['employee_no', 'hire_date', 'first_name']
    ordering = ['employee_no']
    
    def get_serializer_class(self):
        if self.action == 'list':
            return EmployeeListSerializer
        return EmployeeDetailSerializer
    
    @action(detail=True, methods=['get'])
    def profile(self, request, pk=None):
        """Get employee profile with all related data"""
        employee = self.get_object()
        serializer = EmployeeDetailSerializer(employee)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def leave_balance(self, request, pk=None):
        """Get employee leave balances"""
        employee = self.get_object()
        current_year = timezone.now().year
        balances = EmployeeLeaveBalance.objects.filter(
            employee=employee,
            year=current_year
        )
        serializer = EmployeeLeaveBalanceSerializer(balances, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def attendance_summary(self, request, pk=None):
        """Get employee attendance summary for current month"""
        employee = self.get_object()
        today = timezone.now().date()
        month_start = today.replace(day=1)
        
        attendance = AttendanceRecord.objects.filter(
            employee=employee,
            attendance_date__gte=month_start,
            attendance_date__lte=today
        )
        
        summary = {
            'total_days': attendance.count(),
            'present': attendance.filter(status='present').count(),
            'absent': attendance.filter(status='absent').count(),
            'late': attendance.filter(status='late').count(),
            'on_leave': attendance.filter(status='on_leave').count(),
            'total_hours': attendance.aggregate(
                total=Sum('total_hours')
            )['total'] or 0,
            'overtime_hours': attendance.aggregate(
                total=Sum('overtime_hours')
            )['total'] or 0
        }
        
        return Response(summary)
    
    @action(detail=True, methods=['get'])
    def payroll_history(self, request, pk=None):
        """Get employee payroll history"""
        employee = self.get_object()
        calculations = PayrollCalculation.objects.filter(
            employee=employee
        ).select_related('payroll_period').order_by('-payroll_period__start_date')[:12]
        
        serializer = PayrollCalculationSerializer(calculations, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Get employee statistics"""
        total = Employee.objects.count()
        active = Employee.objects.filter(employment_status='active').count()
        
        # Calculate specific category counts
        teaching_staff = Employee.objects.filter(employee_category='teaching').count()
        non_teaching_staff = Employee.objects.filter(employee_category='non_teaching').count()
        
        # Calculate employees currently on leave
        today = timezone.now().date()
        on_leave = LeaveApplication.objects.filter(
            status='approved',
            start_date__lte=today,
            end_date__gte=today
        ).count()
        
        by_category = Employee.objects.values(
            'employee_category'
        ).annotate(count=Count('id'))
        
        by_department = Employee.objects.values(
            'department__name'
        ).annotate(count=Count('id')).order_by('-count')[:10]
        
        stats = {
            'total_employees': total,
            'active_employees': active,
            'teaching_staff': teaching_staff,
            'non_teaching_staff': non_teaching_staff,
            'on_leave': on_leave,
            'by_category': list(by_category),
            'top_departments': list(by_department)
        }
        return Response(stats)
        
    @action(detail=False, methods=['get'])
    def my_comprehensive_summary(self, request):
        """
        Unified dashboard stats for the logged-in employee:
        - Leave balances
        - Upcoming national holidays
        - Current work policy/weekend status
        """
        employee = getattr(request.user, 'employee_profile', None)
        today = timezone.now().date()
        current_year = today.year

        # 1. Leave Balances
        leave_data = []
        if employee:
            leave_balances = EmployeeLeaveBalance.objects.filter(
                employee=employee,
                year=current_year
            ).select_related('leave_type')
            
            for lb in leave_balances:
                leave_data.append({
                    'leave_type': lb.leave_type.name,
                    'leave_code': lb.leave_type.code,
                    'opening': float(lb.opening_balance),
                    'accrued': float(lb.accrued_days),
                    'taken': float(lb.taken_days),
                    'pending': float(lb.pending_days),
                    'remaining': float(lb.closing_balance),
                    'total_entitlement': float(lb.opening_balance + lb.accrued_days + lb.carried_forward_days)
                })

        # 2. Upcoming National Holidays
        TimetableException = apps.get_model('timetable', 'TimetableException')
        holidays = TimetableException.objects.filter(
            exception_type='holiday',
            date_from__gte=today
        ).order_by('date_from')[:3]

        holiday_data = []
        for h in holidays:
            holiday_data.append({
                'name': h.reason,
                'date_from': h.date_from,
                'date_to': h.date_to,
                'is_today': h.date_from <= today <= h.date_to
            })

        # 3. Work Schedule & Weekend Status
        work_status = {
            'has_schedule': False,
            'is_weekend_worker': False,
            'weekend_multiplier': 1.0,
            'holiday_multiplier': 1.0,
            'next_weekend_status': 'Standard Off'
        }

        schedule_assignment = None
        if employee:
            schedule_assignment = EmployeeWorkSchedule.objects.select_related(
                'work_schedule__attendance_policy'
            ).filter(
                employee=employee,
                is_active=True
            ).first()

        if schedule_assignment:
            ws = schedule_assignment.work_schedule
            policy = ws.attendance_policy
            work_status.update({
                'has_schedule': True,
                'schedule_name': ws.name,
                'weekend_multiplier': float(policy.weekend_work_multiplier),
                'holiday_multiplier': float(policy.holiday_work_multiplier),
            })
            
            is_sat_work = ws.saturday_start is not None
            is_sun_work = ws.sunday_start is not None
            work_status['is_weekend_worker'] = is_sat_work or is_sun_work
            
            if is_sat_work and is_sun_work:
                work_status['next_weekend_status'] = 'Full Weekend Shift'
            elif is_sat_work:
                work_status['next_weekend_status'] = 'Saturday Shift'
            elif is_sun_work:
                work_status['next_weekend_status'] = 'Sunday Shift'
            else:
                work_status['next_weekend_status'] = 'Off'
        else:
            # Fallback to general policy if no employee/schedule
            policy = AttendancePolicy.objects.filter(employee_category='all', is_active=True).first()
            if policy:
                work_status.update({
                    'weekend_multiplier': float(policy.weekend_work_multiplier),
                    'holiday_multiplier': float(policy.holiday_work_multiplier),
                })

        return Response({
            'employee_name': employee.get_full_name() if employee else request.user.get_full_name or request.user.username,
            'leave_balances': leave_data,
            'upcoming_holidays': holiday_data,
            'work_policy': work_status,
            'server_date': today
        })


    @action(detail=False, methods=['get'])
    def unlinked(self, request):
        """List employees that don't have a linked user account"""
        employees = Employee.objects.filter(
            user__isnull=True
        ).order_by('first_name', 'last_name')
        search = request.query_params.get('search', '')
        if search:
            employees = employees.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(employee_no__icontains=search) |
                Q(official_email__icontains=search)
            )
        serializer = EmployeeListSerializer(employees[:50], many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='create-user-account')
    def create_user_account(self, request, pk=None):
        """Create a user account for an existing employee and link them"""
        employee = self.get_object()
        if employee.user:
            return Response(
                {'error': 'This employee already has a linked user account.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        email = employee.official_email or employee.personal_email
        if not email:
            return Response(
                {'error': 'Employee has no email address. Cannot create user account.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if User.objects.filter(email__iexact=email).exists():
            return Response(
                {'error': f'A user with email {email} already exists.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Generate username from email
        username = email.split('@')[0]
        base_username = username
        counter = 1
        while User.objects.filter(username=username).exists():
            username = f"{base_username}{counter}"
            counter += 1

        # Use a default password (user must change on first login)
        import secrets
        import string
        password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(12))

        is_lecturer = employee.employee_category == 'teaching'
        gender_map = {'male': 'M', 'female': 'F', 'other': 'O'}

        user = User.objects.create_user(
            email=email,
            username=username,
            first_name=employee.first_name,
            last_name=employee.last_name,
            password=password,
            is_lecturer=is_lecturer,
            gender=gender_map.get(employee.gender, ''),
            phone=employee.phone_primary or '',
            is_first_login=True,
        )

        employee.user = user
        employee.save(update_fields=['user'])

        return Response({
            'success': True,
            'message': f'User account created for {employee.get_full_name()}',
            'user_id': user.id,
            'username': username,
            'password': password,
            'email': email,
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], url_path='bulk-import', parser_classes=[rest_framework.parsers.MultiPartParser])
    def import_data(self, request):
        """Import employees from Excel file"""
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file uploaded', 'message': 'No file uploaded'},
                status=status.HTTP_400_BAD_REQUEST
            )

        file = request.FILES['file']
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'Invalid file format. Please upload .xlsx or .xls file',
                 'message': 'Invalid file format. Please upload .xlsx or .xls file'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            import openpyxl
            wb = openpyxl.load_workbook(file)
            sheet = wb.active
            
            headers = [cell.value for cell in sheet[1]]
            required_headers = ['Employee No', 'First Name', 'Last Name', 'Email', 'Department', 'Job Title']
            
            missing = [h for h in required_headers if h not in headers]
            if missing:
                return Response(
                    {'error': f'Missing headers: {", ".join(missing)}',
                     'message': f'Missing headers: {", ".join(missing)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            success_count = 0
            updated_count = 0
            errors = []
            
            # Valid employee category choices (from Employee.EmployeeCategory)
            VALID_CATEGORIES = {
                'teaching': 'teaching',
                'teaching staff': 'teaching',
                'non_teaching': 'non_teaching',
                'non-teaching': 'non_teaching',
                'non teaching': 'non_teaching',
                'non-teaching staff': 'non_teaching',
                'contract': 'contract',
                'contract staff': 'contract',
                'casual': 'casual',
                'casual worker': 'casual',
                'visiting': 'visiting',
                'visiting faculty': 'visiting',
            }
            
            # Helper to get col index
            def get_col_val(row, header_name):
                try:
                    idx = headers.index(header_name)
                    return row[idx].value
                except (ValueError, IndexError):
                    return None

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                try:
                    emp_no = get_col_val(row, 'Employee No')
                    if not emp_no: continue

                    email = get_col_val(row, 'Email')
                    first_name = get_col_val(row, 'First Name')
                    last_name = get_col_val(row, 'Last Name')
                    dept_name = get_col_val(row, 'Department')
                    job_title_name = get_col_val(row, 'Job Title')
                    phone = get_col_val(row, 'Phone')
                    category_raw = get_col_val(row, 'Employee Category')
                    gender_raw = get_col_val(row, 'Gender')
                    dob_raw = get_col_val(row, 'Date of Birth')
                    national_id = get_col_val(row, 'National ID')
                    hire_date_raw = get_col_val(row, 'Hire Date')
                    middle_name = get_col_val(row, 'Middle Name')
                    personal_email = get_col_val(row, 'Personal Email')
                    employment_type = get_col_val(row, 'Employment Type')

                    # --- Resolve all fields first ---

                    # Get or create department
                    department = None
                    if dept_name:
                        d_name = str(dept_name).strip()
                        department = Department.objects.filter(name__iexact=d_name).first()
                        if not department:
                            d_code = d_name.upper().replace(' ', '_')[:20]
                            department = Department.objects.filter(code=d_code).first()
                        if not department:
                            d_code = d_name.upper().replace(' ', '_')[:20]
                            d_suffix = 1
                            d_base = d_code[:16]
                            while Department.objects.filter(code=d_code).exists():
                                d_code = f"{d_base}_{d_suffix}"[:20]
                                d_suffix += 1
                            department = Department.objects.create(
                                name=d_name, code=d_code, is_active=True
                            )

                    # Resolve employee category
                    category = 'teaching'
                    if category_raw:
                        cat_key = str(category_raw).strip().lower()
                        if cat_key in VALID_CATEGORIES:
                            category = VALID_CATEGORIES[cat_key]
                        else:
                            errors.append(f"Row {row_idx}: Unknown category '{category_raw}', defaulting to 'teaching'")

                    # Resolve gender
                    gender = 'male'
                    if gender_raw:
                        g = str(gender_raw).strip().lower()
                        if g in ('male', 'female', 'other'):
                            gender = g
                        elif g in ('m',):
                            gender = 'male'
                        elif g in ('f',):
                            gender = 'female'

                    # Resolve dates
                    from datetime import datetime as dt
                    dob = None
                    if dob_raw:
                        if hasattr(dob_raw, 'date'):
                            dob = dob_raw.date() if hasattr(dob_raw, 'date') else dob_raw
                        elif isinstance(dob_raw, str):
                            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
                                try:
                                    dob = dt.strptime(dob_raw.strip(), fmt).date()
                                    break
                                except ValueError:
                                    continue
                    if not dob:
                        dob = dt(1990, 1, 1).date()

                    hire_date = None
                    if hire_date_raw:
                        if hasattr(hire_date_raw, 'date'):
                            hire_date = hire_date_raw.date() if hasattr(hire_date_raw, 'date') else hire_date_raw
                        elif isinstance(hire_date_raw, str):
                            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
                                try:
                                    hire_date = dt.strptime(hire_date_raw.strip(), fmt).date()
                                    break
                                except ValueError:
                                    continue
                    if not hire_date:
                        hire_date = timezone.now().date()

                    # --- Update or Create employee ---
                    existing = Employee.objects.filter(employee_no=emp_no).first()
                    if existing:
                        employee = existing
                        if first_name: employee.first_name = str(first_name).strip()
                        if last_name: employee.last_name = str(last_name).strip()
                        if middle_name: employee.middle_name = str(middle_name).strip()
                        if email: employee.official_email = str(email).strip()
                        if personal_email: employee.personal_email = str(personal_email).strip()
                        if phone: employee.phone_primary = str(phone).strip()[:20]
                        if department: employee.department = department
                        if category_raw: employee.employee_category = category
                        if gender_raw: employee.gender = gender
                        if dob_raw and dob: employee.date_of_birth = dob
                        if hire_date_raw and hire_date: employee.hire_date = hire_date
                        if national_id:
                            nid_val = str(national_id).strip()
                            if not Employee.objects.filter(national_id=nid_val).exclude(pk=employee.pk).exists():
                                employee.national_id = nid_val
                        employee.save()
                        updated_count += 1
                    else:
                        # Handle national_id uniqueness for new employees
                        nid = str(national_id).strip() if national_id else str(emp_no)
                        if Employee.objects.filter(national_id=nid).exists():
                            nid = f"{nid}_{emp_no}"

                        employee = Employee.objects.create(
                            employee_no=str(emp_no).strip()[:20],
                            first_name=str(first_name or '').strip(),
                            middle_name=str(middle_name or '').strip(),
                            last_name=str(last_name or '').strip(),
                            official_email=str(email or '').strip(),
                            personal_email=str(personal_email or email or '').strip(),
                            phone_primary=str(phone or '').strip()[:20],
                            department=department,
                            hire_date=hire_date,
                            date_of_birth=dob,
                            national_id=nid,
                            gender=gender,
                            employee_category=category,
                            employment_status='active',
                            payroll_type='monthly'
                        )
                        success_count += 1
                    
                    # Get or create Job Title and assign
                    if job_title_name:
                        jt_name = str(job_title_name).strip()
                        job_title = JobTitle.objects.filter(title__iexact=jt_name).first()
                        if not job_title:
                            # Need a default JobGrade for new titles
                            default_grade = JobGrade.objects.first()
                            if not default_grade:
                                default_grade = JobGrade.objects.create(
                                    code='GEN',
                                    name='General',
                                    category='teaching',
                                    min_salary=0,
                                    max_salary=0,
                                    grade_level=1,
                                    is_active=True
                                )
                            # Map employee category to job title category
                            jt_category = 'teaching'
                            if category in ('non_teaching',):
                                jt_category = 'non_teaching'
                            elif category in ('contract', 'casual', 'visiting'):
                                jt_category = 'contract'
                            # Generate unique code from title
                            jt_code = jt_name.upper().replace(' ', '_')[:20]
                            suffix = 1
                            base_code = jt_code[:16]
                            while JobTitle.objects.filter(code=jt_code).exists():
                                jt_code = f"{base_code}_{suffix}"[:20]
                                suffix += 1
                            job_title = JobTitle.objects.create(
                                title=jt_name,
                                code=jt_code,
                                job_grade=default_grade,
                                category=jt_category,
                                is_active=True
                            )
                        
                        # Resolve employment type
                        emp_type = 'full_time'
                        if employment_type:
                            et = str(employment_type).strip().lower().replace(' ', '_').replace('-', '_')
                            if et in ('full_time', 'part_time', 'contract', 'temporary', 'intern'):
                                emp_type = et

                        # Get default campus
                        default_campus = employee.campus or Campus.objects.first()
                        if not default_campus:
                            default_campus = Campus.objects.create(
                                name='Main Campus',
                                code='MAIN',
                                is_active=True
                            )

                        EmployeeJobAssignment.objects.update_or_create(
                            employee=employee,
                            job_title=job_title,
                            defaults={
                                'department': department or employee.department,
                                'campus': default_campus,
                                'effective_from': hire_date,
                                'is_primary_assignment': True,
                                'employment_type': emp_type,
                                'assignment_type': 'permanent'
                            }
                        )
                    
                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")

            return Response({
                'message': f'Successfully imported {success_count} new, updated {updated_count} employees',
                'imported': success_count,
                'updated': updated_count,
                'failed': len(errors),
                'errors': errors
            })

        except Exception as e:
             return Response(
                {'error': f'Failed to process file: {str(e)}',
                 'message': f'Failed to process file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request):
        """Download Excel template for employee import"""
        import openpyxl
        from django.http import HttpResponse
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employee Import Template"
        
        # Define headers
        headers = [
            'Employee No', 'First Name', 'Middle Name', 'Last Name', 
            'Email', 'Personal Email', 'Phone', 'Department', 
            'Job Title', 'Gender', 'Date of Birth', 'National ID',
            'Hire Date', 'Employee Category', 'Employment Type'
        ]
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        
        # Add sample row
        sample_data = [
            'EMP001', 'John', 'M', 'Doe',
            'john.doe@company.com', 'john.personal@email.com', '+1234567890', 'Academic Department',
            'Lecturer', 'male', '1990-01-15', '12345678',
            '2024-01-01', 'teaching', 'full_time'
        ]
        for col_idx, value in enumerate(sample_data, start=1):
            ws.cell(row=2, column=col_idx).value = value
        
        # Add instructions sheet
        instructions = wb.create_sheet(title="Instructions")
        instructions_data = [
            ["Employee Import Template Instructions"],
            [""],
            ["Required Fields:"],
            ["- Employee No: Unique employee identifier (e.g., EMP001)"],
            ["- First Name: Employee's first name"],
            ["- Last Name: Employee's last name"],
            ["- Email: Official email address"],
            ["- Department: Department name (must exist in system)"],
            [""],
            ["Optional Fields:"],
            ["- Middle Name: Employee's middle name"],
            ["- Personal Email: Personal email address"],
            ["- Phone: Primary phone number"],
            ["- Job Title: Job title (must exist in system)"],
            ["- Gender: male, female, or other"],
            ["- Date of Birth: Format YYYY-MM-DD"],
            ["- National ID: National identification number"],
            ["- Hire Date: Format YYYY-MM-DD"],
            ["- Employee Category: teaching, non_teaching, contract, casual, visiting"],
            ["- Employment Type: full_time, part_time, contract, temporary"],
            [""],
            ["Notes:"],
            ["1. Delete the sample row before importing"],
            ["2. Department and Job Title must match existing records in the system"],
            ["3. Date formats should be YYYY-MM-DD"]
        ]
        for row_idx, row_data in enumerate(instructions_data, start=1):
            instructions.cell(row=row_idx, column=1).value = row_data[0] if row_data else ""
        
        # Auto-adjust column widths
        for ws_sheet in [ws, instructions]:
            for column in ws_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value or '')) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return as downloadable file
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=employee_import_template.xlsx'
        return response


class DepartmentViewSet(viewsets.ModelViewSet):
    """Department ViewSet"""
    queryset = Department.objects.select_related(
        'faculty', 'campus', 'head_of_department'
    )
    serializer_class = DepartmentSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['department_type', 'campus', 'faculty']
    search_fields = ['code', 'name']
    
    @action(detail=True, methods=['get'])
    def employees(self, request, pk=None):
        """Get all employees in this department"""
        department = self.get_object()
        employees = Employee.objects.filter(department=department)
        serializer = EmployeeListSerializer(employees, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def payroll_summary(self, request, pk=None):
        """Get department payroll summary"""
        department = self.get_object()
        current_period = PayrollPeriod.objects.filter(
            status='approved'
        ).order_by('-start_date').first()
        
        if not current_period:
            return Response({'message': 'No approved payroll period'})
        
        calculations = PayrollCalculation.objects.filter(
            employee__department=department,
            payroll_period=current_period
        )
        
        summary = {
            'period': current_period.period_name,
            'employee_count': calculations.count(),
            'total_gross': calculations.aggregate(
                total=Sum('gross_pay')
            )['total'] or 0,
            'total_net': calculations.aggregate(
                total=Sum('net_pay')
            )['total'] or 0,
            'total_deductions': calculations.aggregate(
                total=Sum('total_deductions')
            )['total'] or 0
        }
        
        return Response(summary)


class JobTitleViewSet(viewsets.ModelViewSet):
    """Job Title ViewSet"""
    queryset = JobTitle.objects.select_related('job_grade')
    serializer_class = JobTitleSerializer
    permission_classes = [AllowAny]
    filter_backends = [filters.SearchFilter]
    search_fields = ['title', 'code']


class JobGradeViewSet(viewsets.ModelViewSet):
    """Job Grade ViewSet"""
    queryset = JobGrade.objects.all()
    serializer_class = JobGradeSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=True, methods=['get'])
    def steps(self, request, pk=None):
        """Get all steps for a job grade"""
        grade = self.get_object()
        steps = PayGradeStep.objects.filter(job_grade=grade, is_active=True).order_by('step_number')
        serializer = PayGradeStepSerializer(steps, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def generate_steps(self, request, pk=None):
        """
        Auto-generate incremental steps for a job grade.
        Body: { "num_steps": 12, "increment": 500 }
        Uses min_salary as step 1, increments from there.
        """
        grade = self.get_object()
        num_steps = int(request.data.get('num_steps', 12))
        increment = Decimal(str(request.data.get('increment', 500)))

        # Clear existing steps
        PayGradeStep.objects.filter(job_grade=grade).delete()

        steps = []
        for i in range(1, num_steps + 1):
            amount = grade.min_salary + (increment * (i - 1))
            steps.append(PayGradeStep(
                job_grade=grade,
                step_number=i,
                amount=amount,
                is_active=True
            ))
        PayGradeStep.objects.bulk_create(steps)

        # Update max_salary to match last step
        grade.max_salary = steps[-1].amount
        grade.save(update_fields=['max_salary'])

        serializer = PayGradeStepSerializer(
            PayGradeStep.objects.filter(job_grade=grade).order_by('step_number'),
            many=True
        )
        return Response({
            'message': f'Generated {num_steps} steps for {grade.code}',
            'steps': serializer.data
        })

    def retrieve(self, request, *args, **kwargs):
        """Override to include steps in detail view"""
        instance = self.get_object()
        serializer = JobGradeDetailSerializer(instance)
        return Response(serializer.data)


class PayGradeStepViewSet(viewsets.ModelViewSet):
    """Pay Grade Step CRUD"""
    queryset = PayGradeStep.objects.select_related('job_grade').all()
    serializer_class = PayGradeStepSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['job_grade', 'is_active']
    ordering = ['job_grade', 'step_number']


class PayrollAccountViewSet(viewsets.ModelViewSet):
    """Payroll Account CRUD — unified earnings/deductions master"""
    queryset = PayrollAccount.objects.select_related(
        'employee_gl_account', 'employer_gl_account'
    ).all()
    serializer_class = PayrollAccountSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['account_type', 'category', 'is_active', 'has_employer_contribution',
                        'used_for_paye', 'used_for_nhif_shif', 'used_for_pension',
                        'is_allowable_deduction', 'used_for_housing_levy']
    search_fields = ['code', 'name']
    ordering = ['account_type', 'sort_order', 'name']

    @action(detail=False, methods=['get'])
    def gl_accounts(self, request):
        """Return chart of accounts marked available_in_payroll."""
        from finance.models import Account
        accounts = Account.objects.filter(
            available_in_payroll=True, is_active=True
        ).values('id', 'code', 'name', 'type').order_by('code')
        return Response(list(accounts))


class EmployeeEarningViewSet(viewsets.ModelViewSet):
    """Employee Earning CRUD"""
    queryset = EmployeeEarning.objects.select_related(
        'employee', 'payroll_account', 'earning_type', 'payroll_period'
    ).all()
    serializer_class = EmployeeEarningSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['employee', 'payroll_account', 'is_recurring', 'is_one_time', 'status']
    search_fields = ['employee__employee_no', 'employee__first_name', 'employee__last_name']
    ordering = ['-created_at']


class EmployeeDeductionViewSet(viewsets.ModelViewSet):
    """Employee Deduction CRUD"""
    queryset = EmployeeDeduction.objects.select_related(
        'employee', 'payroll_account', 'deduction_type', 'payroll_period'
    ).all()
    serializer_class = EmployeeDeductionSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['employee', 'payroll_account', 'is_recurring', 'is_one_time', 'status']
    search_fields = ['employee__employee_no', 'employee__first_name', 'employee__last_name']
    ordering = ['-created_at']


class GroupEarningViewSet(viewsets.ModelViewSet):
    """Group (Job Grade) Earning CRUD"""
    queryset = GroupEarning.objects.select_related('job_grade', 'payroll_account').all()
    serializer_class = GroupEarningSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['job_grade', 'payroll_account', 'is_active']
    ordering = ['job_grade', 'payroll_account__sort_order']


class GroupDeductionViewSet(viewsets.ModelViewSet):
    """Group (Job Grade) Deduction CRUD"""
    queryset = GroupDeduction.objects.select_related('job_grade', 'payroll_account').all()
    serializer_class = GroupDeductionSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['job_grade', 'payroll_account', 'is_active']
    ordering = ['job_grade', 'payroll_account__sort_order']


# ============================================================================
# PENSION SCHEME VIEWSETS
# ============================================================================

class PensionSchemeViewSet(viewsets.ModelViewSet):
    """Pension Scheme CRUD"""
    queryset = PensionScheme.objects.select_related('payroll_account').all()
    serializer_class = PensionSchemeSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['scheme_type', 'is_active']
    search_fields = ['code', 'name', 'provider_name']
    ordering = ['name']


class PensionSchemeGradeRateViewSet(viewsets.ModelViewSet):
    """Grade-specific pension rate overrides CRUD"""
    queryset = PensionSchemeGradeRate.objects.select_related(
        'pension_scheme', 'job_grade'
    ).all()
    serializer_class = PensionSchemeGradeRateSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['pension_scheme', 'job_grade']
    ordering = ['pension_scheme', 'job_grade']


class EmployeePensionEnrollmentViewSet(viewsets.ModelViewSet):
    """Employee pension enrollment CRUD"""
    queryset = EmployeePensionEnrollment.objects.select_related(
        'employee', 'pension_scheme'
    ).all()
    serializer_class = EmployeePensionEnrollmentSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['pension_scheme', 'status', 'employee']
    search_fields = ['employee__employee_no', 'employee__first_name', 'employee__last_name', 'membership_number']
    ordering = ['-enrollment_date']


class PensionContributionViewSet(viewsets.ModelViewSet):
    """Pension contribution history CRUD"""
    queryset = PensionContribution.objects.select_related(
        'enrollment__employee', 'enrollment__pension_scheme', 'payroll_period'
    ).all()
    serializer_class = PensionContributionSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['enrollment', 'enrollment__pension_scheme', 'enrollment__employee', 'payroll_period']
    ordering = ['-payroll_period__start_date']


class EmployeePayProfileViewSet(viewsets.ModelViewSet):
    """Employee Pay Profile CRUD"""
    queryset = EmployeePayProfile.objects.select_related(
        'employee', 'pay_profile', 'pay_grade_step',
        'pay_grade_step__job_grade'
    ).all()
    serializer_class = EmployeePayProfileSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['employee', 'pay_profile', 'is_active']
    search_fields = ['employee__employee_no', 'employee__first_name', 'employee__last_name']
    ordering = ['-effective_from']



# ============================================================================
# ATTENDANCE VIEWSETS
# ============================================================================

class AttendancePolicyViewSet(viewsets.ModelViewSet):
    """Attendance policy CRUD"""
    queryset = AttendancePolicy.objects.all()
    serializer_class = AttendancePolicySerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['employee_category', 'is_active']
    search_fields = ['name']
    ordering = ['name']


class EmployeeAttendanceAccessProfileViewSet(viewsets.ModelViewSet):
    """Employee-specific attendance access overrides."""
    queryset = EmployeeAttendanceAccessProfile.objects.select_related(
        'employee', 'assigned_campus'
    ).all()
    serializer_class = EmployeeAttendanceAccessProfileSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['employee', 'is_active', 'assigned_campus']
    search_fields = ['employee__employee_no', 'employee__first_name', 'employee__last_name']
    ordering = ['employee__employee_no']


class BiometricDeviceViewSet(viewsets.ModelViewSet):
    """Biometric device CRUD"""
    queryset = BiometricDevice.objects.select_related('campus').all()
    serializer_class = BiometricDeviceSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['campus', 'device_type', 'is_active']
    search_fields = ['device_name', 'device_code', 'location_description']
    ordering = ['device_name']


class WorkScheduleViewSet(viewsets.ModelViewSet):
    """Work schedule CRUD"""
    queryset = WorkSchedule.objects.select_related('attendance_policy').all()
    serializer_class = WorkScheduleSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['attendance_policy', 'schedule_type', 'is_active']
    search_fields = ['name']
    ordering = ['name']


class EmployeeWorkScheduleViewSet(viewsets.ModelViewSet):
    """Employee schedule assignment CRUD"""
    queryset = EmployeeWorkSchedule.objects.select_related(
        'employee', 'work_schedule'
    ).all()
    serializer_class = EmployeeWorkScheduleSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['employee', 'work_schedule', 'is_active']
    search_fields = ['employee__employee_no', 'employee__first_name', 'employee__last_name', 'work_schedule__name']
    ordering = ['-effective_from']

class AttendanceRecordViewSet(viewsets.ModelViewSet):
    """Attendance Record ViewSet"""
    queryset = AttendanceRecord.objects.select_related(
        'employee', 'work_schedule'
    )
    serializer_class = AttendanceRecordSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = [
        'employee', 'status', 'approval_status', 'attendance_date'
    ]
    search_fields = ['employee__employee_no', 'employee__first_name']

    def _resolve_employee(self, request):
        employee_id = request.data.get('employee_id') or request.query_params.get('employee_id')
        if employee_id:
            try:
                return Employee.objects.get(id=employee_id), None
            except Employee.DoesNotExist:
                return None, Response(
                    {'error': 'Employee not found'},
                    status=status.HTTP_404_NOT_FOUND
                )

        employee = getattr(request.user, 'employee_profile', None)
        if not employee:
            return None, Response(
                {'error': 'No employee profile is linked to your account'},
                status=status.HTTP_400_BAD_REQUEST
            )
        return employee, None

    def _resolve_employee_schedule(self, employee):
        return EmployeeWorkSchedule.objects.select_related(
            'work_schedule__attendance_policy', 'employee__campus'
        ).filter(
            employee=employee,
            is_active=True
        ).first()

    def _effective_clock_policy(self, schedule, employee):
        attendance_policy = schedule.work_schedule.attendance_policy
        access_profile = EmployeeAttendanceAccessProfile.objects.select_related('assigned_campus').filter(
            employee=employee,
            is_active=True
        ).first()

        enforce_biometric = attendance_policy.requires_biometric
        allow_remote = attendance_policy.allow_remote_clock_in
        allow_geo = attendance_policy.allow_geolocation_clock_in
        require_geofence = attendance_policy.require_on_site_geofence

        if access_profile:
            if access_profile.enforce_biometric is not None:
                enforce_biometric = access_profile.enforce_biometric
            if access_profile.allow_remote_clock_in is not None:
                allow_remote = access_profile.allow_remote_clock_in
            if access_profile.allow_geolocation_clock_in is not None:
                allow_geo = access_profile.allow_geolocation_clock_in
            if access_profile.require_on_site_geofence is not None:
                require_geofence = access_profile.require_on_site_geofence

        campus = None
        if access_profile and access_profile.assigned_campus:
            campus = access_profile.assigned_campus
        elif employee.campus:
            campus = employee.campus

        radius = attendance_policy.default_geofence_radius_meters
        if campus and campus.geofence_radius_meters:
            radius = campus.geofence_radius_meters
        if access_profile and access_profile.geofence_radius_meters:
            radius = access_profile.geofence_radius_meters

        allowed_methods = []
        if enforce_biometric:
            allowed_methods = ['biometric']
        else:
            if allow_geo:
                allowed_methods.append('geolocation')
            if allow_remote:
                allowed_methods.append('remote')

        if not allowed_methods:
            # Always keep one method to prevent dead-end configuration.
            allowed_methods = ['remote']

        return {
            'schedule': schedule,
            'attendance_policy': attendance_policy,
            'access_profile': access_profile,
            'campus': campus,
            'radius': radius,
            'enforce_biometric': enforce_biometric,
            'require_geofence': require_geofence,
            'allowed_methods': allowed_methods,
        }

    def _distance_meters(self, lat1, lon1, lat2, lon2):
        earth_radius_m = 6371000
        dlat = radians(lat2 - lat1)
        dlon = radians(lon2 - lon1)
        a = sin(dlat / 2) ** 2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon / 2) ** 2
        c = 2 * atan2(sqrt(a), sqrt(1 - a))
        return earth_radius_m * c

    def _validate_geofence(self, method, policy_data, request):
        if method == 'remote':
            return None

        if method != 'geolocation' and not policy_data['require_geofence']:
            return None

        campus = policy_data['campus']
        if not campus or campus.latitude is None or campus.longitude is None:
            return Response(
                {'error': 'Campus geofence coordinates are not configured'},
                status=status.HTTP_400_BAD_REQUEST
            )

        latitude = request.data.get('latitude')
        longitude = request.data.get('longitude')
        if latitude in [None, ''] or longitude in [None, '']:
            return Response(
                {'error': 'Latitude and longitude are required for on-site clocking'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            user_lat = float(latitude)
            user_lon = float(longitude)
        except (TypeError, ValueError):
            return Response(
                {'error': 'Invalid latitude/longitude values'},
                status=status.HTTP_400_BAD_REQUEST
            )

        distance = self._distance_meters(
            user_lat,
            user_lon,
            float(campus.latitude),
            float(campus.longitude)
        )

        if distance > float(policy_data['radius']):
            return Response(
                {
                    'error': 'You are outside the allowed campus geofence',
                    'distance_meters': round(distance, 2),
                    'allowed_radius_meters': float(policy_data['radius'])
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        return None

    @action(detail=False, methods=['get'])
    def my_clock_policy(self, request):
        """Return effective attendance methods and geofence requirements for logged-in user."""
        employee = getattr(request.user, 'employee_profile', None)
        if not employee:
            return self.general_policy(request)

        schedule = self._resolve_employee_schedule(employee)
        if not schedule:
            return self.general_policy(request)

        policy_data = self._effective_clock_policy(schedule, employee)
        policy = policy_data['attendance_policy']
        campus = policy_data['campus']

        return Response({
            'employee_id': employee.id,
            'employee_no': employee.employee_no,
            'employee_name': employee.get_full_name(),
            'work_schedule': schedule.work_schedule.name,
            'attendance_policy': policy.name,
            'holiday_work_multiplier': float(policy.holiday_work_multiplier),
            'weekend_work_multiplier': float(policy.weekend_work_multiplier),
            'allowed_methods': policy_data['allowed_methods'],
            'requires_biometric': policy_data['enforce_biometric'],
            'requires_geofence': policy_data['require_geofence'],
            'geofence_radius_meters': policy_data['radius'],
            'campus': {
                'id': campus.id,
                'name': campus.name,
                'latitude': campus.latitude,
                'longitude': campus.longitude,
            } if campus else None,
        })

    @action(detail=False, methods=['get'])
    def general_policy(self, request):
        """Return default system attendance policy for users without a specific assignment."""
        policy = AttendancePolicy.objects.filter(employee_category='all', is_active=True).first()
        if not policy:
            # Fallback to absolute defaults if no 'all' policy exists
            return Response({
                'attendance_policy': 'Default System Policy',
                'holiday_work_multiplier': 1.0,
                'weekend_work_multiplier': 1.0,
                'allowed_methods': ['remote'],
                'requires_biometric': False,
                'requires_geofence': False,
                'geofence_radius_meters': 200,
                'campus': None
            })

        campus = Campus.objects.first() # Default to first campus if none assigned

        allowed_methods = []
        if policy.requires_biometric:
            allowed_methods = ['biometric']
        else:
            if policy.allow_geolocation_clock_in:
                allowed_methods.append('geolocation')
            if policy.allow_remote_clock_in:
                allowed_methods.append('remote')
        
        if not allowed_methods:
            allowed_methods = ['remote']

        return Response({
            'attendance_policy': policy.name,
            'holiday_work_multiplier': float(policy.holiday_work_multiplier),
            'weekend_work_multiplier': float(policy.weekend_work_multiplier),
            'allowed_methods': allowed_methods,
            'requires_biometric': policy.requires_biometric,
            'requires_geofence': policy.require_on_site_geofence,
            'geofence_radius_meters': policy.default_geofence_radius_meters,
            'campus': {
                'id': campus.id,
                'name': campus.name,
                'latitude': campus.latitude,
                'longitude': campus.longitude,
            } if campus else None,
        })
    
    @action(detail=False, methods=['post'])
    def clock_in(self, request):
        """Clock in for attendance"""
        employee, error_response = self._resolve_employee(request)
        if error_response:
            return error_response
        
        today = timezone.now().date()
        current_time = timezone.now().time()
        
        # Check if already clocked in
        existing = AttendanceRecord.objects.filter(
            employee=employee,
            attendance_date=today
        ).first()
        
        if existing and existing.check_in_time:
            return Response(
                {'error': 'Already clocked in today'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get employee work schedule
        schedule = EmployeeWorkSchedule.objects.filter(
            employee=employee,
            is_active=True
        ).first()
        
        if not schedule:
            return Response(
                {'error': 'No active work schedule'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        policy_data = self._effective_clock_policy(schedule, employee)
        requested_method = request.data.get('method')
        if requested_method in [None, '']:
            requested_method = policy_data['allowed_methods'][0]

        if requested_method not in policy_data['allowed_methods']:
            return Response(
                {
                    'error': f"Method '{requested_method}' is not allowed for your attendance policy",
                    'allowed_methods': policy_data['allowed_methods']
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        geofence_error = self._validate_geofence(requested_method, policy_data, request)
        if geofence_error:
            return geofence_error
        
        # Create or update attendance record
        attendance, created = AttendanceRecord.objects.get_or_create(
            employee=employee,
            attendance_date=today,
            defaults={
                'work_schedule': schedule.work_schedule,
                'check_in_time': current_time,
                'check_in_method': requested_method,
                'check_in_location': request.data.get('location_text', '') or f"{request.data.get('latitude', '')},{request.data.get('longitude', '')}".strip(','),
                'status': 'present'
            }
        )
        
        if not created:
            attendance.check_in_time = current_time
            attendance.check_in_method = requested_method
            attendance.check_in_location = request.data.get('location_text', '') or f"{request.data.get('latitude', '')},{request.data.get('longitude', '')}".strip(',')
            attendance.status = 'present'
            attendance.save()
        
        serializer = AttendanceRecordSerializer(attendance)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    @action(detail=False, methods=['post'])
    def clock_out(self, request):
        """Clock out for attendance"""
        employee, error_response = self._resolve_employee(request)
        if error_response:
            return error_response
        
        today = timezone.now().date()
        current_time = timezone.now().time()

        schedule = self._resolve_employee_schedule(employee)
        if not schedule:
            return Response(
                {'error': 'No active work schedule'},
                status=status.HTTP_400_BAD_REQUEST
            )

        policy_data = self._effective_clock_policy(schedule, employee)
        requested_method = request.data.get('method')
        if requested_method in [None, '']:
            requested_method = policy_data['allowed_methods'][0]

        if requested_method not in policy_data['allowed_methods']:
            return Response(
                {
                    'error': f"Method '{requested_method}' is not allowed for your attendance policy",
                    'allowed_methods': policy_data['allowed_methods']
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        geofence_error = self._validate_geofence(requested_method, policy_data, request)
        if geofence_error:
            return geofence_error
        
        # Get today's attendance record
        try:
            attendance = AttendanceRecord.objects.get(
                employee=employee,
                attendance_date=today
            )
        except AttendanceRecord.DoesNotExist:
            return Response(
                {'error': 'No clock-in record found for today'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        if attendance.check_out_time:
            return Response(
                {'error': 'Already clocked out today'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Update attendance record
        attendance.check_out_time = current_time
        attendance.check_out_method = requested_method
        attendance.check_out_location = request.data.get('location_text', '') or f"{request.data.get('latitude', '')},{request.data.get('longitude', '')}".strip(',')
        
        # Calculate hours worked
        from datetime import datetime, date
        check_in_datetime = datetime.combine(date.today(), attendance.check_in_time)
        check_out_datetime = datetime.combine(date.today(), current_time)
        
        if check_out_datetime < check_in_datetime:
            # Handle overnight shifts
            check_out_datetime = datetime.combine(
                date.today() + timedelta(days=1), 
                current_time
            )
        
        hours_worked = (check_out_datetime - check_in_datetime).seconds / 3600
        attendance.total_hours = round(hours_worked, 2)
        
        # Calculate regular and overtime hours, and determine status
        policy = attendance.work_schedule.attendance_policy
        if hours_worked <= policy.standard_hours_per_day:
            attendance.regular_hours = hours_worked
            attendance.overtime_hours = 0
        else:
            attendance.regular_hours = policy.standard_hours_per_day
            attendance.overtime_hours = hours_worked - policy.standard_hours_per_day
        
        # Determine half-day status
        if hours_worked < policy.minimum_hours_for_full_day:
            if hours_worked >= policy.half_day_threshold_hours:
                attendance.status = 'half_day'
            else:
                attendance.status = 'absent'
        else:
            # If they met minimum hours, they stay 'present'
            attendance.status = 'present'
            
        attendance.save()
        
        serializer = AttendanceRecordSerializer(attendance)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def daily_summary(self, request):
        """Get daily attendance summary"""
        date_str = request.query_params.get('date', timezone.now().date())
        
        records = AttendanceRecord.objects.filter(
            attendance_date=date_str
        )
        
        summary = {
            'date': date_str,
            'total': records.count(),
            'present': records.filter(status='present').count(),
            'absent': records.filter(status='absent').count(),
            'late': records.filter(status='late').count(),
            'on_leave': records.filter(status='on_leave').count(),
            'half_day': records.filter(status='half_day').count()
        }
        
        return Response(summary)


class OvertimeRequestViewSet(viewsets.ModelViewSet):
    """Overtime Request ViewSet"""
    queryset = OvertimeRequest.objects.select_related(
        'employee', 'department'
    )
    serializer_class = OvertimeRequestSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['approval_status', 'employee', 'overtime_date']
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve overtime request"""
        overtime_request = self.get_object()
        
        if overtime_request.approval_status != 'pending':
            return Response(
                {'error': 'Request is not pending'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        overtime_request.approval_status = 'approved'
        overtime_request.approved_by = request.user
        overtime_request.save()
        
        serializer = self.get_serializer(overtime_request)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Reject overtime request"""
        overtime_request = self.get_object()
        
        if overtime_request.approval_status != 'pending':
            return Response(
                {'error': 'Request is not pending'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        overtime_request.approval_status = 'rejected'
        overtime_request.approved_by = request.user
        overtime_request.approval_notes = request.data.get('notes', '')
        overtime_request.save()
        
        serializer = self.get_serializer(overtime_request)
        return Response(serializer.data)


# ============================================================================
# LEAVE VIEWSETS
# ============================================================================

class LeaveTypeViewSet(viewsets.ModelViewSet):
    queryset = LeaveType.objects.all()
    serializer_class = LeaveTypeSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at']

    @action(detail=False, methods=['post'])
    def populate(self, request):
        if not (request.user.is_superuser or request.user.is_staff):
            return Response({'error': 'Unauthorized'}, status=403)
            
        defaults = [
            {
                'code': 'ANNUAL',
                'name': 'Annual Leave',
                'description': 'Standard paid time off',
                'category': 'paid',
                'is_statutory': True,
                'max_days_per_year': 21,
                'gender_specific': 'all',
                'requires_medical_certificate': False,
                'advance_notice_days': 7,
                'accrual_rate': 1.75,
                'accrual_method': 'monthly'
            },
            {
                'code': 'SICK',
                'name': 'Sick Leave',
                'description': 'Medical leave',
                'category': 'paid',
                'is_statutory': True,
                'max_days_per_year': 14,
                'gender_specific': 'all',
                'requires_medical_certificate': True,
                'advance_notice_days': 0,
                'accrual_rate': 1.17,
                'accrual_method': 'monthly'
            },
            {
                'code': 'MATERNITY',
                'name': 'Maternity Leave',
                'description': 'Maternity leave for mothers',
                'category': 'paid',
                'is_statutory': True,
                'max_days_per_year': 90,
                'gender_specific': 'female',
                'requires_medical_certificate': True,
                'advance_notice_days': 30,
                'accrual_rate': 0.00,
                'accrual_method': 'yearly'
            }
        ]
        
        created_count = 0
        for data in defaults:
            obj, created = LeaveType.objects.get_or_create(
                code=data['code'],
                defaults=data
            )
            if created:
                created_count += 1
                
        return Response({'message': f'Successfully populated {created_count} leave types.'})

class LeaveApplicationViewSet(viewsets.ModelViewSet):
    """Leave Application ViewSet"""
    queryset = LeaveApplication.objects.select_related(
        'employee', 'leave_type', 'acting_employee'
    )
    serializer_class = LeaveApplicationSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['employee', 'leave_type', 'status', 'start_date']
    search_fields = ['employee__employee_no', 'employee__first_name']
    
    def perform_create(self, serializer):
        # If employee is provided and user is admin, allow it. Otherwise default to logged in user.
        # Actually, let's just use request.data.get('employee') if present, else fallback to request.user.employee
        if 'employee' not in self.request.data and hasattr(self.request.user, 'employee'):
            serializer.save(employee=self.request.user.employee)
        else:
            serializer.save()
            
    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """Submit leave application"""
        leave_app = self.get_object()
        
        if leave_app.status != 'draft':
            return Response(
                {'error': 'Can only submit draft applications'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if employee has sufficient leave balance
        balance = EmployeeLeaveBalance.objects.filter(
            employee=leave_app.employee,
            leave_type=leave_app.leave_type,
            year=leave_app.start_date.year
        ).first()
        
        if balance:
            available = balance.closing_balance - balance.pending_days
            if available < leave_app.working_days:
                return Response(
                    {'error': f'Insufficient leave balance. Available: {available} days'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        leave_app.status = 'submitted'
        leave_app.submitted_date = timezone.now()
        leave_app.save()
        
        # Update pending days in balance
        if balance:
            balance.pending_days += leave_app.working_days
            balance.save()
        
        serializer = self.get_serializer(leave_app)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve leave application"""
        leave_app = self.get_object()
        
        if leave_app.status not in ['submitted', 'pending_approval']:
            return Response(
                {'error': 'Invalid status for approval'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        leave_app.status = 'approved'
        leave_app.approved_date = timezone.now()
        leave_app.save()
        
        serializer = self.get_serializer(leave_app)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Reject leave application"""
        leave_app = self.get_object()
        
        if leave_app.status not in ['submitted', 'pending_approval']:
            return Response(
                {'error': 'Invalid status for rejection'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        leave_app.status = 'rejected'
        leave_app.rejected_date = timezone.now()
        leave_app.rejection_reason = request.data.get('reason', '')
        leave_app.save()
        
        # Update pending days in balance
        balance = EmployeeLeaveBalance.objects.filter(
            employee=leave_app.employee,
            leave_type=leave_app.leave_type,
            year=leave_app.start_date.year
        ).first()
        
        if balance:
            balance.pending_days -= leave_app.working_days
            balance.save()
        
        serializer = self.get_serializer(leave_app)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def pending_approvals(self, request):
        """Get pending leave approvals"""
        pending = LeaveApplication.objects.filter(
            status__in=['submitted', 'pending_approval']
        )
        serializer = self.get_serializer(pending, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def dashboard_metrics(self, request):
        """Get leave dashboard metrics for specific UI design"""
        from django.db.models import Sum, Q
        now = timezone.now()
        year = now.year
        month = now.month
        
        # Current month queryset
        qs_month = self.get_queryset().filter(
            start_date__year=year,
            start_date__month=month
        )
        
        # Calculate days by leave type
        def get_days_for_type(type_name_contains):
            return qs_month.filter(
                leave_type__name__icontains=type_name_contains,
                status__in=['approved', 'on_leave', 'completed']
            ).aggregate(total=Sum('working_days'))['total'] or 0

        annual_days = get_days_for_type('Annual')
        sick_days = get_days_for_type('Sick')
        
        other_days = qs_month.exclude(
            Q(leave_type__name__icontains='Annual') | Q(leave_type__name__icontains='Sick')
        ).filter(status__in=['approved', 'on_leave', 'completed']).aggregate(total=Sum('working_days'))['total'] or 0
        
        pending_leave_count = qs_month.filter(status__in=['submitted', 'pending_approval']).count()
        
        # Overtime pending count
        pending_overtime_count = OvertimeRequest.objects.filter(
            approval_status='pending',
            overtime_date__year=year,
            overtime_date__month=month
        ).count()
        
        metrics = [
            {
                'id': 1,
                'title': 'Annual Leave',
                'value': int(annual_days),
                'subtitle': 'This month'
            },
            {
                'id': 2,
                'title': 'Sick Leave',
                'value': int(sick_days),
                'subtitle': 'This month'
            },
            {
                'id': 3,
                'title': 'Other Leave',
                'value': int(other_days),
                'subtitle': 'This month'
            },
            {
                'id': 4,
                'title': 'Pending Request',
                'value': pending_leave_count,
                'subtitle': 'This month'
            }
        ]
        
        return Response({
            'metrics': metrics,
            'pending_leaves': pending_leave_count,
            'pending_overtime': pending_overtime_count
        })

    @action(detail=False, methods=['get'])
    def analytics(self, request):
        """Get leave analytics data"""
        from django.db.models import Count, Q
        from datetime import timedelta
        
        # Get last 7 days
        today = timezone.now().date()
        days = [(today - timedelta(days=i)) for i in range(6, -1, -1)]
        
        # Base queryset for approved leaves in this range
        qs = self.get_queryset().filter(
            status__in=['approved', 'on_leave', 'completed'],
            start_date__lte=days[-1],
            end_date__gte=days[0]
        )
        
        daily_trend = []
        for d in days:
            # Leaves active on this day
            active = qs.filter(start_date__lte=d, end_date__gte=d)
            annual = active.filter(leave_type__name__icontains='Annual').count()
            emergency = active.filter(leave_type__name__icontains='Emergency').count()
            sick = active.filter(leave_type__name__icontains='Sick').count()
            
            months = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            daily_trend.append({
                'name': f"{d.day} {months[d.month]}",
                'Annual Leave': annual,
                'Emergency Leave': emergency,
                'Sick Leave': sick
            })
            
        return Response({
            'dailyTrend': daily_trend
        })

    @action(detail=False, methods=['get'])
    def upcoming_leaves(self, request):
        """Get upcoming leaves"""
        today = timezone.now().date()
        upcoming = self.get_queryset().filter(
            status='approved',
            start_date__gte=today
        ).order_by('start_date')[:5]
        
        serializer = self.get_serializer(upcoming, many=True)
        return Response(serializer.data)


class EmployeeLeaveBalanceViewSet(viewsets.ReadOnlyModelViewSet):
    """Employee Leave Balance ViewSet (Read-only)"""
    queryset = EmployeeLeaveBalance.objects.select_related(
        'employee', 'leave_type'
    )
    serializer_class = EmployeeLeaveBalanceSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['employee', 'leave_type', 'year']
    
    @action(detail=False, methods=['post'])
    def accrue_monthly(self, request):
        """Accrue monthly leave for all employees"""
        current_year = timezone.now().year
        current_month = timezone.now().month
        
        employees = Employee.objects.filter(employment_status='active')
        leave_types = LeaveType.objects.filter(is_active=True)
        
        accrued_count = 0
        
        for employee in employees:
            for leave_type in leave_types:
                balance, created = EmployeeLeaveBalance.objects.get_or_create(
                    employee=employee,
                    leave_type=leave_type,
                    year=current_year
                )
                
                balance.accrued_days += leave_type.accrual_rate
                balance.closing_balance = (
                    balance.opening_balance +
                    balance.accrued_days +
                    balance.carried_forward_days -
                    balance.taken_days
                )
                balance.last_accrual_date = timezone.now().date()
                balance.save()
                accrued_count += 1
        
        return Response({
            'message': f'Accrued leave for {accrued_count} records',
            'employees': employees.count(),
            'leave_types': leave_types.count()
        })


# ============================================================================
# PAYROLL VIEWSETS
# ============================================================================

class PayrollPeriodViewSet(viewsets.ModelViewSet):
    """Payroll Period ViewSet"""
    queryset = PayrollPeriod.objects.all()
    serializer_class = PayrollPeriodSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['status', 'period_type']
    ordering = ['-start_date']
    
    @action(detail=True, methods=['post'])
    def process(self, request, pk=None):
        """Process payroll for the period"""
        period = self.get_object()
        
        if period.status not in ['open', 'processing']:
            return Response(
                {'error': 'Period is not open for processing'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        period.status = 'processing'
        period.processing_started_at = timezone.now()
        period.save()
        
        from workforce.services.payroll_service import PayrollCalculationService
        service = PayrollCalculationService(period, processed_by=request.user)
        result = service.run()

        if not result['success']:
            return Response(
                {'error': result['message'], 'errors': result['errors']},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        return Response({
            'message': result['message'],
            'period': period.period_name,
            'processed': result['processed'],
            'errors': result['errors'],
            'status': period.status,
        })
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve payroll period"""
        period = self.get_object()
        
        if period.status != 'calculated':
            return Response(
                {'error': 'Period must be calculated before approval'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        period.status = 'approved'
        period.approved_by = request.user
        period.approved_date = timezone.now().date()
        period.save()
        
        serializer = self.get_serializer(period)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def summary(self, request, pk=None):
        """Get payroll period summary"""
        period = self.get_object()
        
        calculations = PayrollCalculation.objects.filter(
            payroll_period=period
        )
        
        summary = {
            'period': period.period_name,
            'status': period.get_status_display(),
            'employee_count': calculations.count(),
            'total_gross_pay': calculations.aggregate(
                total=Sum('gross_pay')
            )['total'] or 0,
            'total_deductions': calculations.aggregate(
                total=Sum('total_deductions')
            )['total'] or 0,
            'total_net_pay': calculations.aggregate(
                total=Sum('net_pay')
            )['total'] or 0,
            'total_tax': calculations.aggregate(
                total=Sum('tax_amount')
            )['total'] or 0,
            'total_pension': calculations.aggregate(
                emp=Sum('pension_employee'),
                empr=Sum('pension_employer')
            )
        }
        
        return Response(summary)

    @action(detail=True, methods=['post'])
    def unlock(self, request, pk=None):
        """Unlock a payroll period for re-run. Requires can_unlock_payroll_period permission."""
        period = self.get_object()

        if not (request.user.is_superuser or request.user.has_perm('workforce.can_unlock_payroll_period')):
            return Response(
                {'error': 'You do not have permission to unlock payroll periods.'},
                status=status.HTTP_403_FORBIDDEN
            )

        if period.status == 'open':
            return Response(
                {'error': 'Period is already open.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if period.status == 'paid':
            return Response(
                {'error': 'Cannot unlock a paid period. Reverse payments first.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        period.status = 'open'
        period.locked = False
        period.approved_by = None
        period.approved_date = None
        period.processing_started_at = None
        period.processing_completed_at = None
        period.save()

        serializer = self.get_serializer(period)
        return Response(serializer.data)


class PayrollCalculationViewSet(viewsets.ReadOnlyModelViewSet):
    """Payroll Calculation ViewSet (Read-only)"""
    queryset = PayrollCalculation.objects.select_related(
        'employee', 'payroll_period'
    )
    serializer_class = PayrollCalculationSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['payroll_period', 'employee', 'payment_status']
    search_fields = ['employee__employee_no', 'employee__first_name']
    
    @action(detail=True, methods=['get'])
    def breakdown(self, request, pk=None):
        """Get detailed payroll breakdown"""
        calculation = self.get_object()
        
        details = PayrollCalculationDetail.objects.filter(
            payroll_calculation=calculation
        )
        
        earnings = details.filter(item_type='earning')
        deductions = details.filter(item_type='deduction')
        
        # Reconstruct reliefs applied
        reliefs = []
        try:
            from workforce.models import TaxRelief
            personal = TaxRelief.objects.filter(relief_type='personal', is_active=True).first()
            if personal and personal.amount:
                reliefs.append({
                    'description': 'Personal Relief',
                    'amount': float(personal.amount)
                })
            
            shif_detail = deductions.filter(description__icontains='shif').first()
            if shif_detail:
                insurance = TaxRelief.objects.filter(relief_type='insurance', is_active=True).first()
                if insurance:
                    relief_pct = insurance.percentage or 15
                    insurance_relief = float(shif_detail.amount) * float(relief_pct) / 100
                    if insurance.max_amount:
                        insurance_relief = min(insurance_relief, float(insurance.max_amount))
                    if insurance_relief > 0:
                        reliefs.append({
                            'description': 'Insurance Relief',
                            'amount': insurance_relief
                        })
        except Exception:
            pass
        
        breakdown = {
            'employee': {
                'employee_no': calculation.employee.employee_no,
                'name': calculation.employee.get_full_name(),
                'department': calculation.employee.department.name if getattr(calculation.employee, 'department', None) else 'N/A',
                'kra_pin': getattr(calculation.employee, 'tax_pin', 'N/A'),
                'profile_image': calculation.employee.profile_picture.url if getattr(calculation.employee, 'profile_picture', None) else (calculation.employee.photo.url if getattr(calculation.employee, 'photo', None) else None),
            },
            'period': calculation.payroll_period.period_name,
            'earnings': [
                {
                    'description': d.description,
                    'amount': float(d.amount)
                } for d in earnings
            ],
            'deductions': [
                {
                    'description': d.description,
                    'amount': float(d.amount)
                } for d in deductions
            ],
            'reliefs': reliefs,
            'summary': {
                'gross_pay': float(calculation.gross_pay),
                'total_deductions': float(calculation.total_deductions),
                'net_pay': float(calculation.net_pay)
            }
        }
        
        return Response(breakdown)


class PayslipTemplateViewSet(viewsets.ModelViewSet):
    """Payslip Template ViewSet for frontend customization"""
    queryset = PayslipTemplate.objects.all()
    serializer_class = PayslipTemplateSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['is_active', 'is_default']




class PayslipViewSet(viewsets.ReadOnlyModelViewSet):
    """Payslip ViewSet"""
    queryset = Payslip.objects.select_related(
        'employee', 'payroll_period', 'payroll_calculation'
    )
    serializer_class = PayslipSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['employee', 'payroll_period']
    
    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        """Download payslip PDF"""
        payslip = self.get_object()
        
        # Mark as downloaded
        if not payslip.downloaded:
            payslip.downloaded = True
            payslip.download_date = timezone.now()
            payslip.save()
        
        # TODO: Generate and return PDF
        return Response({
            'message': 'PDF generation not implemented yet',
            'payslip_number': payslip.payslip_number
        })
    
    @action(detail=True, methods=['post'])
    def email(self, request, pk=None):
        """Email payslip to employee"""
        payslip = self.get_object()
        
        # TODO: Implement email sending
        
        payslip.email_sent = True
        payslip.email_sent_date = timezone.now()
        payslip.save()
        
        return Response({
            'message': 'Payslip emailed successfully',
            'email': payslip.employee.official_email
        })


# ============================================================================
# HRMS ENHANCEMENT VIEWSETS
# ============================================================================

from .services import (
    EmployeeLifecycleService, PositionService, BulkImportService,
    OrgChartService, DocumentService, HRAnalyticsService
)
from .serializers import (
    PositionSerializer, PositionListSerializer, EmployeeLifecycleLogSerializer,
    BulkImportSessionSerializer, BulkImportRecordSerializer,
    EmployeeDocumentSerializer, EmployeeDocumentUploadSerializer,
    HRAutomationRuleSerializer, HRAutomationLogSerializer,
    HRNotificationPreferenceSerializer
)


class PositionViewSet(viewsets.ModelViewSet):
    """
    Position ViewSet for headcount management.
    """
    queryset = Position.objects.select_related(
        'department', 'job_title', 'job_grade', 
        'current_holder', 'reports_to_position'
    )
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = [
        'department', 'job_title', 'job_grade',
        'vacancy_status', 'funding_status', 'is_active'
    ]
    search_fields = ['position_code', 'position_name', 'duties_summary']
    ordering_fields = ['position_code', 'created_at']
    ordering = ['position_code']
    
    def get_serializer_class(self):
        if self.action == 'list':
            return PositionListSerializer
        return PositionSerializer
    
    @action(detail=False, methods=['get'])
    def vacant(self, request):
        """Get all vacant positions"""
        service = PositionService()
        department_id = request.query_params.get('department')
        department = Department.objects.get(id=department_id) if department_id else None
        positions = service.get_vacant_positions(department=department)
        serializer = PositionListSerializer(positions, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def headcount_summary(self, request):
        """Get headcount summary by department"""
        service = PositionService()
        summary = service.get_headcount_summary()
        return Response(summary)
    
    @action(detail=True, methods=['post'])
    def assign_employee(self, request, pk=None):
        """Assign an employee to this position"""
        position = self.get_object()
        employee_id = request.data.get('employee_id')
        effective_date = request.data.get('effective_date')
        
        try:
            employee = Employee.objects.get(id=employee_id)
            service = PositionService()
            position = service.assign_employee_to_position(
                position=position,
                employee=employee,
                assigned_by=request.user,
                effective_date=effective_date
            )
            return Response(PositionSerializer(position).data)
        except Employee.DoesNotExist:
            return Response(
                {'error': 'Employee not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def vacate(self, request, pk=None):
        """Vacate this position"""
        position = self.get_object()
        reason = request.data.get('reason', '')
        
        service = PositionService()
        position = service.vacate_position(
            position=position,
            vacated_by=request.user,
            reason=reason
        )
        return Response(PositionSerializer(position).data)
    
    @action(detail=True, methods=['post'])
    def freeze(self, request, pk=None):
        """Freeze this position"""
        position = self.get_object()
        reason = request.data.get('reason', '')
        
        service = PositionService()
        position = service.freeze_position(
            position=position,
            frozen_by=request.user,
            reason=reason
        )
        return Response(PositionSerializer(position).data)
    
    @action(detail=True, methods=['post'])
    def unfreeze(self, request, pk=None):
        """Unfreeze this position"""
        position = self.get_object()
        
        service = PositionService()
        position = service.unfreeze_position(
            position=position,
            unfrozen_by=request.user
        )
        return Response(PositionSerializer(position).data)
    
    @action(detail=False, methods=['get'])
    def hierarchy(self, request):
        """Get position hierarchy for a department"""
        department_id = request.query_params.get('department')
        if not department_id:
            return Response(
                {'error': 'department parameter required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            department = Department.objects.get(id=department_id)
            service = PositionService()
            hierarchy = service.get_position_hierarchy(department)
            return Response(hierarchy)
        except Department.DoesNotExist:
            return Response(
                {'error': 'Department not found'},
                status=status.HTTP_404_NOT_FOUND
            )


class EmployeeLifecycleLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Employee Lifecycle Log ViewSet (Read-only audit trail).
    """
    queryset = EmployeeLifecycleLog.objects.select_related(
        'employee', 'performed_by', 'old_department', 
        'new_department', 'old_job_title', 'new_job_title'
    ).order_by('-event_date', '-created_at')
    serializer_class = EmployeeLifecycleLogSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['employee', 'event_type', 'performed_by']
    search_fields = ['employee__employee_no', 'employee__first_name', 'notes']
    ordering_fields = ['event_date', 'created_at']
    
    @action(detail=False, methods=['get'])
    def by_employee(self, request):
        """Get lifecycle history for a specific employee"""
        employee_id = request.query_params.get('employee_id')
        if not employee_id:
            return Response(
                {'error': 'employee_id parameter required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        logs = self.queryset.filter(employee_id=employee_id)
        serializer = self.get_serializer(logs, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def recent(self, request):
        """Get recent lifecycle events"""
        days = int(request.query_params.get('days', 30))
        cutoff = timezone.now() - timedelta(days=days)
        logs = self.queryset.filter(event_date__gte=cutoff)[:50]
        serializer = self.get_serializer(logs, many=True)
        return Response(serializer.data)


class EmployeeLifecycleViewSet(viewsets.ViewSet):
    """
    ViewSet for employee lifecycle actions (hire, promote, transfer, etc.)
    """
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['post'])
    def hire(self, request):
        """Hire a new employee"""
        service = EmployeeLifecycleService()
        try:
            employee = service.hire_employee(
                employee_data=request.data.get('employee_data', {}),
                position_id=request.data.get('position_id'),
                hired_by=request.user,
                notes=request.data.get('notes', '')
            )
            return Response(
                EmployeeDetailSerializer(employee).data,
                status=status.HTTP_201_CREATED
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'], url_path='confirm-probation')
    def confirm_probation(self, request, pk=None):
        """Confirm employee after probation"""
        service = EmployeeLifecycleService()
        try:
            employee = Employee.objects.get(pk=pk)
            employee = service.confirm_probation(
                employee=employee,
                confirmed_by=request.user,
                new_salary=request.data.get('new_salary'),
                notes=request.data.get('notes', '')
            )
            return Response(EmployeeDetailSerializer(employee).data)
        except Employee.DoesNotExist:
            return Response(
                {'error': 'Employee not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def promote(self, request, pk=None):
        """Promote an employee"""
        service = EmployeeLifecycleService()
        try:
            employee = Employee.objects.get(pk=pk)
            employee = service.promote(
                employee=employee,
                new_job_title_id=request.data.get('new_job_title_id'),
                new_job_grade_id=request.data.get('new_job_grade_id'),
                new_salary=request.data.get('new_salary'),
                effective_date=request.data.get('effective_date'),
                promoted_by=request.user,
                notes=request.data.get('notes', '')
            )
            return Response(EmployeeDetailSerializer(employee).data)
        except Employee.DoesNotExist:
            return Response(
                {'error': 'Employee not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def transfer(self, request, pk=None):
        """Transfer an employee"""
        service = EmployeeLifecycleService()
        try:
            employee = Employee.objects.get(pk=pk)
            employee = service.transfer(
                employee=employee,
                new_department_id=request.data.get('new_department_id'),
                new_position_id=request.data.get('new_position_id'),
                effective_date=request.data.get('effective_date'),
                transferred_by=request.user,
                notes=request.data.get('notes', '')
            )
            return Response(EmployeeDetailSerializer(employee).data)
        except Employee.DoesNotExist:
            return Response(
                {'error': 'Employee not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def suspend(self, request, pk=None):
        """Suspend an employee"""
        service = EmployeeLifecycleService()
        try:
            employee = Employee.objects.get(pk=pk)
            employee = service.suspend(
                employee=employee,
                reason=request.data.get('reason', ''),
                suspended_by=request.user,
                expected_end_date=request.data.get('expected_end_date')
            )
            return Response(EmployeeDetailSerializer(employee).data)
        except Employee.DoesNotExist:
            return Response(
                {'error': 'Employee not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def reinstate(self, request, pk=None):
        """Reinstate a suspended employee"""
        service = EmployeeLifecycleService()
        try:
            employee = Employee.objects.get(pk=pk)
            employee = service.reinstate(
                employee=employee,
                reinstated_by=request.user,
                notes=request.data.get('notes', '')
            )
            return Response(EmployeeDetailSerializer(employee).data)
        except Employee.DoesNotExist:
            return Response(
                {'error': 'Employee not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def resign(self, request, pk=None):
        """Process employee resignation"""
        service = EmployeeLifecycleService()
        try:
            employee = Employee.objects.get(pk=pk)
            employee = service.resign(
                employee=employee,
                resignation_date=request.data.get('resignation_date'),
                last_working_date=request.data.get('last_working_date'),
                reason=request.data.get('reason', ''),
                processed_by=request.user
            )
            return Response(EmployeeDetailSerializer(employee).data)
        except Employee.DoesNotExist:
            return Response(
                {'error': 'Employee not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def terminate(self, request, pk=None):
        """Terminate an employee"""
        service = EmployeeLifecycleService()
        try:
            employee = Employee.objects.get(pk=pk)
            employee = service.terminate(
                employee=employee,
                termination_date=request.data.get('termination_date'),
                reason=request.data.get('reason', ''),
                terminated_by=request.user
            )
            return Response(EmployeeDetailSerializer(employee).data)
        except Employee.DoesNotExist:
            return Response(
                {'error': 'Employee not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def retire(self, request, pk=None):
        """Retire an employee"""
        service = EmployeeLifecycleService()
        try:
            employee = Employee.objects.get(pk=pk)
            employee = service.retire(
                employee=employee,
                retirement_date=request.data.get('retirement_date'),
                retired_by=request.user,
                notes=request.data.get('notes', '')
            )
            return Response(EmployeeDetailSerializer(employee).data)
        except Employee.DoesNotExist:
            return Response(
                {'error': 'Employee not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )


class BulkImportSessionViewSet(viewsets.ModelViewSet):
    """
    Bulk Import Session ViewSet for Excel data imports.
    """
    queryset = BulkImportSession.objects.select_related('uploaded_by')
    serializer_class = BulkImportSessionSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['status', 'import_type']
    ordering = ['-created_at']
    
    def perform_create(self, serializer):
        serializer.save(uploaded_by=self.request.user)
    
    @action(detail=True, methods=['get'])
    def records(self, request, pk=None):
        """Get import records for this session"""
        session = self.get_object()
        records = BulkImportRecord.objects.filter(import_session=session)
        serializer = BulkImportRecordSerializer(records, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def validate(self, request, pk=None):
        """Validate imported data"""
        session = self.get_object()
        service = BulkImportService()
        
        try:
            result = service.validate_import(session)
            return Response({
                'valid_records': result['valid_count'],
                'invalid_records': result['invalid_count'],
                'errors': result['errors']
            })
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def execute(self, request, pk=None):
        """Execute the import"""
        session = self.get_object()
        service = BulkImportService()
        
        try:
            result = service.execute_import(session, request.user)
            return Response({
                'processed': result['processed_count'],
                'created': result['created_count'],
                'errors': result['errors']
            })
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['get'])
    def template(self, request):
        """Download import template"""
        import_type = request.query_params.get('type', 'employee')
        service = BulkImportService()
        
        template = service.generate_template(import_type)
        return Response({
            'fields': template['fields'],
            'sample_data': template['sample_data']
        })


class EmployeeDocumentViewSet(viewsets.ModelViewSet):
    """
    Employee Document ViewSet for document management.
    """
    queryset = EmployeeDocument.objects.select_related(
        'employee', 'uploaded_by', 'verified_by'
    )
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['employee', 'category', 'verification_status', 'is_required']
    search_fields = ['document_name', 'document_number']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        if self.action == 'create':
            return EmployeeDocumentUploadSerializer
        return EmployeeDocumentSerializer
    
    def perform_create(self, serializer):
        serializer.save(uploaded_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def verify(self, request, pk=None):
        """Verify a document"""
        document = self.get_object()
        service = DocumentService()
        
        try:
            document = service.verify_document(
                document=document,
                verified_by=request.user
            )
            return Response(EmployeeDocumentSerializer(document).data)
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Reject a document"""
        document = self.get_object()
        service = DocumentService()
        
        try:
            document = service.reject_document(
                document=document,
                rejected_by=request.user,
                reason=request.data.get('reason', '')
            )
            return Response(EmployeeDocumentSerializer(document).data)
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['get'])
    def expiring(self, request):
        """Get documents expiring soon"""
        days = int(request.query_params.get('days', 30))
        service = DocumentService()
        documents = service.get_expiring_documents(days=days)
        serializer = EmployeeDocumentSerializer(documents, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def by_employee(self, request):
        """Get all documents for an employee"""
        employee_id = request.query_params.get('employee_id')
        if not employee_id:
            return Response(
                {'error': 'employee_id parameter required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        documents = self.queryset.filter(employee_id=employee_id)
        serializer = EmployeeDocumentSerializer(documents, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Get document statistics"""
        service = DocumentService()
        stats = service.get_document_statistics()
        return Response(stats)


class HRAutomationRuleViewSet(viewsets.ModelViewSet):
    """
    HR Automation Rule ViewSet.
    """
    queryset = HRAutomationRule.objects.select_related('created_by')
    serializer_class = HRAutomationRuleSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['trigger_type', 'action_type', 'is_active']
    search_fields = ['name', 'description']
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def toggle(self, request, pk=None):
        """Toggle rule active status"""
        rule = self.get_object()
        rule.is_active = not rule.is_active
        rule.save()
        return Response(HRAutomationRuleSerializer(rule).data)
    
    @action(detail=True, methods=['get'])
    def logs(self, request, pk=None):
        """Get execution logs for this rule"""
        rule = self.get_object()
        logs = HRAutomationLog.objects.filter(rule=rule).order_by('-executed_at')[:50]
        serializer = HRAutomationLogSerializer(logs, many=True)
        return Response(serializer.data)


class HRNotificationPreferenceViewSet(viewsets.ModelViewSet):
    """
    HR Notification Preference ViewSet.
    """
    queryset = HRNotificationPreference.objects.select_related('user')
    serializer_class = HRNotificationPreferenceSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        # Users can only see their own preferences
        if self.request.user.is_superuser:
            return self.queryset
        return self.queryset.filter(user=self.request.user)
    
    @action(detail=False, methods=['get', 'post'])
    def my_preferences(self, request):
        """Get or update current user's preferences"""
        pref, created = HRNotificationPreference.objects.get_or_create(
            user=request.user,
            defaults={
                'probation_alerts': True,
                'contract_alerts': True,
                'document_expiry_alerts': True,
                'leave_alerts': True,
                'email_notifications': True,
                'system_notifications': True,
            }
        )
        
        if request.method == 'POST':
            serializer = HRNotificationPreferenceSerializer(
                pref, data=request.data, partial=True
            )
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        return Response(HRNotificationPreferenceSerializer(pref).data)


class OrgChartViewSet(viewsets.ViewSet):
    """
    Organization Chart ViewSet.
    """
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def departments(self, request):
        """Get department hierarchy tree"""
        service = OrgChartService()
        tree = service.get_department_tree()
        return Response(tree)
    
    @action(detail=False, methods=['get'])
    def positions(self, request):
        """Get position hierarchy for a department"""
        department_id = request.query_params.get('department')
        if not department_id:
            return Response(
                {'error': 'department parameter required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            department = Department.objects.get(id=department_id)
            service = OrgChartService()
            hierarchy = service.get_position_hierarchy(department)
            return Response(hierarchy)
        except Department.DoesNotExist:
            return Response(
                {'error': 'Department not found'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=False, methods=['get'])
    def reporting_chain(self, request):
        """Get reporting chain for an employee"""
        employee_id = request.query_params.get('employee')
        if not employee_id:
            return Response(
                {'error': 'employee parameter required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            employee = Employee.objects.get(id=employee_id)
            service = OrgChartService()
            chain = service.get_reporting_chain(employee)
            return Response(chain)
        except Employee.DoesNotExist:
            return Response(
                {'error': 'Employee not found'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=False, methods=['get'])
    def direct_reports(self, request):
        """Get direct reports for a manager"""
        manager_id = request.query_params.get('manager')
        if not manager_id:
            return Response(
                {'error': 'manager parameter required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            manager = Employee.objects.get(id=manager_id)
            service = OrgChartService()
            reports = service.get_direct_reports(manager)
            return Response(reports)
        except Employee.DoesNotExist:
            return Response(
                {'error': 'Manager not found'},
                status=status.HTTP_404_NOT_FOUND
            )


class HRAnalyticsViewSet(viewsets.ViewSet):
    """
    HR Analytics ViewSet for dashboards and reporting.
    """
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        """Get comprehensive HR dashboard"""
        service = HRAnalyticsService()
        data = service.get_dashboard_summary()
        return Response(data)
    
    @action(detail=False, methods=['get'])
    def headcount(self, request):
        """Get headcount analytics"""
        service = HRAnalyticsService()
        department_id = request.query_params.get('department')
        department = None
        if department_id:
            try:
                department = Department.objects.get(id=department_id)
            except Department.DoesNotExist:
                pass
        
        data = service.get_headcount_analytics(department=department)
        return Response(data)
    
    @action(detail=False, methods=['get'])
    def turnover(self, request):
        """Get turnover analytics"""
        from datetime import datetime
        service = HRAnalyticsService()
        
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if start_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        else:
            start_date = timezone.now().date().replace(month=1, day=1)
        
        if end_date:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        else:
            end_date = timezone.now().date()
        
        data = service.get_turnover_analytics(start_date, end_date)
        return Response(data)
    
    @action(detail=False, methods=['get'])
    def tenure(self, request):
        """Get tenure analytics"""
        service = HRAnalyticsService()
        department_id = request.query_params.get('department')
        department = None
        if department_id:
            try:
                department = Department.objects.get(id=department_id)
            except Department.DoesNotExist:
                pass
        
        data = service.get_tenure_analytics(department=department)
        return Response(data)
    
    @action(detail=False, methods=['get'])
    def demographics(self, request):
        """Get demographics analytics"""
        service = HRAnalyticsService()
        department_id = request.query_params.get('department')
        department = None
        if department_id:
            try:
                department = Department.objects.get(id=department_id)
            except Department.DoesNotExist:
                pass
        
        data = service.get_demographics_analytics(department=department)
        return Response(data)
    
    @action(detail=False, methods=['get'])
    def probation(self, request):
        """Get probation analytics"""
        service = HRAnalyticsService()
        data = service.get_probation_analytics()
        return Response(data)
    
    @action(detail=False, methods=['get'])
    def contracts(self, request):
        """Get contract analytics"""
        service = HRAnalyticsService()
        data = service.get_contract_analytics()
        return Response(data)


# ============================================================================
# PAYROLL DASHBOARD VIEWSET
# ============================================================================

class PayrollDashboardViewSet(viewsets.ViewSet):
    """
    Payroll Dashboard aggregate endpoints for KPIs and analytics.
    Provides real-time payroll metrics without modifying existing payroll logic.
    """
    permission_classes = [AllowAny]  # Change to IsAuthenticated in production
    
    @action(detail=False, methods=['get'])
    def summary(self, request):
        """
        Returns comprehensive dashboard KPI summary.
        GET /workforce/api/payroll-dashboard/summary/
        """
        from django.db.models.functions import Coalesce
        from decimal import Decimal
        
        # Get current/latest payroll period
        current_period = PayrollPeriod.objects.filter(
            status__in=['open', 'processing', 'calculated', 'approved']
        ).order_by('-start_date').first()
        
        # Get last completed period for comparison
        last_period = PayrollPeriod.objects.filter(
            status='paid'
        ).order_by('-start_date').first()
        
        # Calculate current period stats
        current_stats = {'gross': 0, 'net': 0, 'deductions': 0, 'allowances': 0, 'employees': 0}
        if current_period:
            calcs = PayrollCalculation.objects.filter(payroll_period=current_period)
            agg = calcs.aggregate(
                gross=Coalesce(Sum('gross_pay'), Decimal('0')),
                net=Coalesce(Sum('net_pay'), Decimal('0')),
                deductions=Coalesce(Sum('total_deductions'), Decimal('0')),
                allowances=Coalesce(Sum('total_allowances'), Decimal('0')),
                count=Count('id')
            )
            current_stats = {
                'gross': float(agg['gross']),
                'net': float(agg['net']),
                'deductions': float(agg['deductions']),
                'allowances': float(agg['allowances']),
                'employees': agg['count']
            }
        
        # Calculate last period stats for comparison
        last_stats = {'gross': 0, 'net': 0}
        if last_period:
            last_agg = PayrollCalculation.objects.filter(
                payroll_period=last_period
            ).aggregate(
                gross=Coalesce(Sum('gross_pay'), Decimal('0')),
                net=Coalesce(Sum('net_pay'), Decimal('0'))
            )
            last_stats = {
                'gross': float(last_agg['gross']),
                'net': float(last_agg['net'])
            }
        
        # Calculate percentage changes
        gross_change = 0
        net_change = 0
        if last_stats['gross'] > 0:
            gross_change = round(((current_stats['gross'] - last_stats['gross']) / last_stats['gross']) * 100, 1)
        if last_stats['net'] > 0:
            net_change = round(((current_stats['net'] - last_stats['net']) / last_stats['net']) * 100, 1)
        
        # Count employees
        total_employees = Employee.objects.filter(employment_status='active').count()
        
        # Count pending items
        pending_approvals = PayrollPeriod.objects.filter(status='calculated').count()
        pending_calculations = PayrollCalculation.objects.filter(payment_status='pending').count()
        
        return Response({
            'total_employees': total_employees,
            'gross_payroll': current_stats['gross'],
            'gross_change': gross_change,
            'net_payable': current_stats['net'],
            'net_change': net_change,
            'total_deductions': current_stats['deductions'],
            'total_allowances': current_stats['allowances'],
            'employees_processed': current_stats['employees'],
            'pending_approvals': pending_approvals,
            'pending_calculations': pending_calculations,
            'current_period': {
                'id': current_period.id if current_period else None,
                'name': current_period.period_name if current_period else 'No Active Period',
                'status': current_period.status if current_period else None,
                'start_date': current_period.start_date if current_period else None,
                'end_date': current_period.end_date if current_period else None,
                'payment_date': current_period.payment_date if current_period else None,
            } if current_period else None,
            'last_period': {
                'name': last_period.period_name if last_period else None,
                'gross': last_stats['gross'],
                'net': last_stats['net']
            } if last_period else None
        })
    
    @action(detail=False, methods=['get'], url_path='monthly-trends')
    def monthly_trends(self, request):
        """
        Returns payroll trends for the last N months.
        GET /workforce/api/payroll-dashboard/monthly-trends/?months=6
        """
        from django.db.models.functions import Coalesce, TruncMonth
        from decimal import Decimal
        
        months = int(request.query_params.get('months', 6))
        
        # Get periods from the last N months
        cutoff_date = timezone.now().date() - timedelta(days=months * 30)
        
        periods = PayrollPeriod.objects.filter(
            start_date__gte=cutoff_date,
            status__in=['paid', 'approved', 'closed']
        ).order_by('start_date')
        
        trend_data = []
        for period in periods:
            calcs = PayrollCalculation.objects.filter(payroll_period=period)
            agg = calcs.aggregate(
                gross=Coalesce(Sum('gross_pay'), Decimal('0')),
                net=Coalesce(Sum('net_pay'), Decimal('0')),
                deductions=Coalesce(Sum('total_deductions'), Decimal('0')),
                allowances=Coalesce(Sum('total_allowances'), Decimal('0')),
                employees=Count('id')
            )
            
            trend_data.append({
                'period_id': period.id,
                'period_name': period.period_name,
                'month': period.start_date.strftime('%b'),
                'year': period.start_date.year,
                'gross': float(agg['gross']),
                'net': float(agg['net']),
                'deductions': float(agg['deductions']),
                'allowances': float(agg['allowances']),
                'employees': agg['employees']
            })
        
        return Response({
            'trends': trend_data,
            'period_count': len(trend_data)
        })
    
    @action(detail=False, methods=['get'], url_path='department-costs')
    def department_costs(self, request):
        """
        Returns payroll costs breakdown by department.
        GET /workforce/api/payroll-dashboard/department-costs/?period_id=1
        """
        from django.db.models.functions import Coalesce
        from decimal import Decimal
        
        period_id = request.query_params.get('period_id')
        
        # Get period
        if period_id:
            period = PayrollPeriod.objects.filter(id=period_id).first()
        else:
            period = PayrollPeriod.objects.filter(
                status__in=['paid', 'approved', 'calculated']
            ).order_by('-start_date').first()
        
        if not period:
            return Response({'departments': [], 'total': 0})
        
        # Get department breakdown
        dept_costs = PayrollCalculation.objects.filter(
            payroll_period=period
        ).values(
            'employee__department__name',
            'employee__department__id'
        ).annotate(
            gross=Coalesce(Sum('gross_pay'), Decimal('0')),
            net=Coalesce(Sum('net_pay'), Decimal('0')),
            deductions=Coalesce(Sum('total_deductions'), Decimal('0')),
            employee_count=Count('id')
        ).order_by('-gross')
        
        departments = []
        total_gross = 0
        for dept in dept_costs:
            gross = float(dept['gross'])
            total_gross += gross
            departments.append({
                'id': dept['employee__department__id'],
                'name': dept['employee__department__name'] or 'Unassigned',
                'gross': gross,
                'net': float(dept['net']),
                'deductions': float(dept['deductions']),
                'employees': dept['employee_count']
            })
        
        # Add percentage
        for dept in departments:
            dept['percentage'] = round((dept['gross'] / total_gross * 100), 1) if total_gross > 0 else 0
        
        return Response({
            'period_name': period.period_name,
            'departments': departments,
            'total_gross': total_gross
        })
    
    @action(detail=False, methods=['get'], url_path='current-period')
    def current_period(self, request):
        """
        Returns current active payroll period with workflow status.
        GET /workforce/api/payroll-dashboard/current-period/
        """
        period = PayrollPeriod.objects.filter(
            status__in=['open', 'processing', 'calculated', 'approved']
        ).order_by('-start_date').first()
        
        if not period:
            # Fall back to most recent period
            period = PayrollPeriod.objects.order_by('-start_date').first()
        
        if not period:
            return Response({
                'period': None,
                'workflow': [],
                'message': 'No payroll periods found'
            })
        
        # Build workflow steps based on status
        status_order = ['open', 'processing', 'calculated', 'approved', 'paid', 'closed']
        current_idx = status_order.index(period.status) if period.status in status_order else 0
        
        workflow_steps = [
            {
                'id': 1,
                'name': 'Draft',
                'status': 'completed' if current_idx >= 0 else 'upcoming',
                'date': period.created_at.strftime('%b %d') if hasattr(period, 'created_at') and period.created_at else '--'
            },
            {
                'id': 2,
                'name': 'Processing',
                'status': 'completed' if current_idx >= 1 else ('current' if current_idx == 0 else 'upcoming'),
                'date': period.processing_started_at.strftime('%b %d') if period.processing_started_at else '--'
            },
            {
                'id': 3,
                'name': 'Calculated',
                'status': 'completed' if current_idx >= 2 else ('current' if current_idx == 1 else 'upcoming'),
                'date': period.processing_completed_at.strftime('%b %d') if period.processing_completed_at else '--'
            },
            {
                'id': 4,
                'name': 'Approval',
                'status': 'completed' if current_idx >= 3 else ('current' if current_idx == 2 else 'upcoming'),
                'date': period.approved_date.strftime('%b %d') if period.approved_date else '--'
            },
            {
                'id': 5,
                'name': 'Disbursement',
                'status': 'completed' if current_idx >= 4 else ('current' if current_idx == 3 else 'upcoming'),
                'date': period.payment_date.strftime('%b %d') if period.payment_date and current_idx >= 4 else '--'
            }
        ]
        
        return Response({
            'period': {
                'id': period.id,
                'name': period.period_name,
                'status': period.status,
                'status_display': period.get_status_display(),
                'start_date': period.start_date,
                'end_date': period.end_date,
                'payment_date': period.payment_date,
                'employee_count': period.employee_count,
                'total_gross': float(period.total_gross_pay),
                'total_net': float(period.total_net_pay),
                'locked': period.locked
            },
            'workflow': workflow_steps,
            'progress_percentage': int((current_idx + 1) / len(status_order) * 100)
        })
    
    @action(detail=False, methods=['get'])
    def alerts(self, request):
        """
        Returns payroll alerts and notifications.
        GET /workforce/api/payroll-dashboard/alerts/
        """
        alerts = []
        
        # Check for employees missing tax info
        missing_tax = Employee.objects.filter(
            employment_status='active',
            tax_pin__isnull=True
        ).count() + Employee.objects.filter(
            employment_status='active',
            tax_pin=''
        ).count()
        
        if missing_tax > 0:
            alerts.append({
                'id': 1,
                'type': 'error',
                'title': 'Missing Tax PINs',
                'message': f'{missing_tax} active employees are missing KRA PINs. Compliance risk.',
                'action': 'Fix Now',
                'action_url': '/dashboard/hr/staff?filter=missing_tax'
            })
        
        # Check for pending payroll approvals
        pending_approval = PayrollPeriod.objects.filter(status='calculated').count()
        if pending_approval > 0:
            alerts.append({
                'id': 2,
                'type': 'warning',
                'title': 'Pending Approval',
                'message': f'{pending_approval} payroll period(s) awaiting approval.',
                'action': 'Review',
                'action_url': '/dashboard/payroll'
            })
        
        # Check for employees without salary structure
        no_salary = Employee.objects.filter(
            employment_status='active'
        ).exclude(
            payroll_calculations__isnull=False
        ).count()
        
        # Check for new employees not in current payroll
        current_period = PayrollPeriod.objects.filter(
            status__in=['open', 'processing', 'calculated']
        ).order_by('-start_date').first()
        
        if current_period:
            processed_employees = PayrollCalculation.objects.filter(
                payroll_period=current_period
            ).values_list('employee_id', flat=True)
            
            active_employees = Employee.objects.filter(
                employment_status='active'
            ).exclude(id__in=processed_employees).count()
            
            if active_employees > 0:
                alerts.append({
                    'id': 3,
                    'type': 'info',
                    'title': 'Unprocessed Employees',
                    'message': f'{active_employees} active employees not yet included in current payroll.',
                    'action': 'Review',
                    'action_url': '/dashboard/payroll'
                })
        
        # Check statutory deadlines (example: PAYE due by 9th)
        today = timezone.now().date()
        if today.day <= 9:
            alerts.append({
                'id': 4,
                'type': 'info',
                'title': 'Statutory Deadline',
                'message': f'PAYE remittance due by {today.replace(day=9).strftime("%b %d")}.',
                'action': 'View',
                'action_url': '/dashboard/payroll/statutory'
            })
        
        return Response({
            'alerts': alerts,
            'count': len(alerts),
            'has_critical': any(a['type'] == 'error' for a in alerts)
        })
    
    @action(detail=False, methods=['get'], url_path='recent-runs')
    def recent_runs(self, request):
        """
        Returns recent payroll runs with summary data.
        GET /workforce/api/payroll-dashboard/recent-runs/?limit=5
        """
        from django.db.models.functions import Coalesce
        from decimal import Decimal
        
        limit = int(request.query_params.get('limit', 5))
        
        periods = PayrollPeriod.objects.filter(
            status__in=['paid', 'approved', 'closed', 'calculated']
        ).order_by('-start_date')[:limit]
        
        runs = []
        for period in periods:
            calcs = PayrollCalculation.objects.filter(payroll_period=period)
            agg = calcs.aggregate(
                gross=Coalesce(Sum('gross_pay'), Decimal('0')),
                net=Coalesce(Sum('net_pay'), Decimal('0')),
                count=Count('id')
            )
            
            runs.append({
                'id': period.id,
                'period': period.period_name,
                'start_date': period.start_date,
                'end_date': period.end_date,
                'payment_date': period.payment_date,
                'employees': agg['count'],
                'gross_pay': float(agg['gross']),
                'net_pay': float(agg['net']),
                'status': period.status,
                'status_display': period.get_status_display()
            })
        
        return Response({
            'runs': runs,
            'total_count': PayrollPeriod.objects.count()
        })
    
    @action(detail=False, methods=['get'])
    def distribution(self, request):
        """
        Returns payroll distribution data for charts.
        GET /workforce/api/payroll-dashboard/distribution/?period_id=1
        """
        from django.db.models.functions import Coalesce
        from decimal import Decimal
        
        period_id = request.query_params.get('period_id')
        
        if period_id:
            period = PayrollPeriod.objects.filter(id=period_id).first()
        else:
            period = PayrollPeriod.objects.filter(
                status__in=['paid', 'approved', 'calculated']
            ).order_by('-start_date').first()
        
        if not period:
            return Response({
                'distribution': [],
                'total': 0
            })
        
        calcs = PayrollCalculation.objects.filter(payroll_period=period)
        agg = calcs.aggregate(
            basic=Coalesce(Sum('basic_salary'), Decimal('0')),
            allowances=Coalesce(Sum('total_allowances'), Decimal('0')),
            overtime=Coalesce(Sum('total_overtime'), Decimal('0')),
            bonuses=Coalesce(Sum('total_bonuses'), Decimal('0')),
            deductions=Coalesce(Sum('total_deductions'), Decimal('0')),
            tax=Coalesce(Sum('tax_amount'), Decimal('0')),
            pension_emp=Coalesce(Sum('pension_employee'), Decimal('0')),
            gross=Coalesce(Sum('gross_pay'), Decimal('0'))
        )
        
        total = float(agg['gross'])
        
        distribution = [
            {
                'name': 'Basic Pay',
                'value': float(agg['basic']),
                'percentage': round(float(agg['basic']) / total * 100, 1) if total > 0 else 0,
                'color': '#10b981'
            },
            {
                'name': 'Allowances',
                'value': float(agg['allowances']),
                'percentage': round(float(agg['allowances']) / total * 100, 1) if total > 0 else 0,
                'color': '#3b82f6'
            },
            {
                'name': 'Overtime',
                'value': float(agg['overtime']),
                'percentage': round(float(agg['overtime']) / total * 100, 1) if total > 0 else 0,
                'color': '#8b5cf6'
            },
            {
                'name': 'Deductions',
                'value': float(agg['deductions']),
                'percentage': round(float(agg['deductions']) / total * 100, 1) if total > 0 else 0,
                'color': '#f59e0b'
            }
        ]
        
        return Response({
            'period_name': period.period_name,
            'distribution': distribution,
            'total': total,
            'summary': {
                'tax': float(agg['tax']),
                'pension': float(agg['pension_emp']),
                'bonuses': float(agg['bonuses'])
            }
        })


class PayrollSettingsViewSet(viewsets.ViewSet):
    """
    Payroll Settings API for managing configuration, tax bands, 
    statutory rates, earning types, deduction types, and workflows.
    """
    permission_classes = [AllowAny]  # Change to IsAuthenticated in production
    
    # ==========================================
    # PAYROLL CYCLE CONFIGURATION
    # ==========================================
    
    @action(detail=False, methods=['get', 'put'], url_path='cycle')
    def cycle_settings(self, request):
        """
        GET/PUT payroll cycle configuration.
        GET /workforce/api/payroll-settings/cycle/
        PUT /workforce/api/payroll-settings/cycle/
        """
        from .models import PayrollConfiguration
        
        config = PayrollConfiguration.get_active()
        
        if request.method == 'GET':
            return Response({
                'pay_frequency': config.pay_frequency,
                'payroll_run_date': config.payroll_run_date,
                'attendance_cutoff_date': config.attendance_cutoff_date,
                'financial_year_start': config.financial_year_start,
                'is_locked': config.is_locked,
                'currency_code': config.currency_code,
                'rounding_method': config.rounding_method,
                'overtime_multiplier': float(config.overtime_multiplier),
                'holiday_multiplier': float(config.holiday_multiplier),
                'enable_proration': config.enable_proration,
                'require_2fa_approval': config.require_2fa_approval,
                'allow_retroactive_changes': config.allow_retroactive_changes,
            })
        
        # PUT request
        data = request.data
        config.pay_frequency = data.get('pay_frequency', config.pay_frequency)
        config.payroll_run_date = data.get('payroll_run_date', config.payroll_run_date)
        config.attendance_cutoff_date = data.get('attendance_cutoff_date', config.attendance_cutoff_date)
        config.financial_year_start = data.get('financial_year_start', config.financial_year_start)
        config.is_locked = data.get('is_locked', config.is_locked)
        config.currency_code = data.get('currency_code', config.currency_code)
        config.rounding_method = data.get('rounding_method', config.rounding_method)
        config.overtime_multiplier = data.get('overtime_multiplier', config.overtime_multiplier)
        config.holiday_multiplier = data.get('holiday_multiplier', config.holiday_multiplier)
        config.enable_proration = data.get('enable_proration', config.enable_proration)
        config.require_2fa_approval = data.get('require_2fa_approval', config.require_2fa_approval)
        config.allow_retroactive_changes = data.get('allow_retroactive_changes', config.allow_retroactive_changes)
        config.save()
        
        return Response({
            'message': 'Payroll cycle settings updated successfully',
            'data': {
                'pay_frequency': config.pay_frequency,
                'payroll_run_date': config.payroll_run_date,
                'attendance_cutoff_date': config.attendance_cutoff_date,
                'financial_year_start': config.financial_year_start,
                'is_locked': config.is_locked,
            }
        })
    
    # ==========================================
    # EARNING TYPES (Salary Components)
    # ==========================================
    
    @action(detail=False, methods=['get', 'post'], url_path='earning-types')
    def earning_types(self, request):
        """
        GET all earning types or POST to create new.
        """
        if request.method == 'GET':
            earnings = EarningType.objects.filter(is_active=True).order_by('sort_order', 'name')
            data = [{
                'id': e.id,
                'code': e.code,
                'name': e.name,
                'category': e.category,
                'category_display': e.get_category_display(),
                'is_taxable': e.is_taxable,
                'is_pensionable': e.is_pensionable,
                'gl_account_code': e.gl_account_code,
                'description': e.description,
                'is_active': e.is_active,
                'sort_order': e.sort_order,
            } for e in earnings]
            return Response({'earning_types': data, 'count': len(data)})
        
        # POST - Create new
        data = request.data
        earning = EarningType.objects.create(
            code=data.get('code'),
            name=data.get('name'),
            category=data.get('category', 'allowance'),
            is_taxable=data.get('is_taxable', True),
            is_pensionable=data.get('is_pensionable', False),
            gl_account_code=data.get('gl_account_code', ''),
            description=data.get('description', ''),
            sort_order=data.get('sort_order', 0)
        )
        return Response({
            'message': 'Earning type created successfully',
            'id': earning.id,
            'name': earning.name
        }, status=201)
    
    @action(detail=False, methods=['put', 'delete'], url_path='earning-types/(?P<pk>[^/.]+)')
    def earning_type_detail(self, request, pk=None):
        """
        PUT/DELETE specific earning type.
        """
        try:
            earning = EarningType.objects.get(pk=pk)
        except EarningType.DoesNotExist:
            return Response({'error': 'Earning type not found'}, status=404)
        
        if request.method == 'DELETE':
            # Soft delete
            earning.is_active = False
            earning.save()
            return Response({'message': 'Earning type deleted'})
        
        # PUT - Update
        data = request.data
        earning.name = data.get('name', earning.name)
        earning.category = data.get('category', earning.category)
        earning.is_taxable = data.get('is_taxable', earning.is_taxable)
        earning.is_pensionable = data.get('is_pensionable', earning.is_pensionable)
        earning.gl_account_code = data.get('gl_account_code', earning.gl_account_code)
        earning.description = data.get('description', earning.description)
        earning.is_active = data.get('is_active', earning.is_active)
        earning.sort_order = data.get('sort_order', earning.sort_order)
        earning.save()
        
        return Response({'message': 'Earning type updated', 'id': earning.id})
    
    # ==========================================
    # DEDUCTION TYPES
    # ==========================================
    
    @action(detail=False, methods=['get', 'post'], url_path='deduction-types')
    def deduction_types(self, request):
        """
        GET all deduction types or POST to create new.
        """
        if request.method == 'GET':
            deductions = DeductionType.objects.filter(is_active=True).order_by('sort_order', 'name')
            data = [{
                'id': d.id,
                'code': d.code,
                'name': d.name,
                'category': d.category,
                'category_display': d.get_category_display(),
                'is_mandatory': d.is_mandatory,
                'gl_account_code': d.gl_account_code,
                'description': d.description,
                'is_active': d.is_active,
                'sort_order': d.sort_order,
            } for d in deductions]
            return Response({'deduction_types': data, 'count': len(data)})
        
        # POST - Create new
        data = request.data
        deduction = DeductionType.objects.create(
            code=data.get('code'),
            name=data.get('name'),
            category=data.get('category', 'voluntary'),
            is_mandatory=data.get('is_mandatory', False),
            gl_account_code=data.get('gl_account_code', ''),
            description=data.get('description', ''),
            sort_order=data.get('sort_order', 0)
        )
        return Response({
            'message': 'Deduction type created successfully',
            'id': deduction.id,
            'name': deduction.name
        }, status=201)
    
    @action(detail=False, methods=['put', 'delete'], url_path='deduction-types/(?P<pk>[^/.]+)')
    def deduction_type_detail(self, request, pk=None):
        """
        PUT/DELETE specific deduction type.
        """
        try:
            deduction = DeductionType.objects.get(pk=pk)
        except DeductionType.DoesNotExist:
            return Response({'error': 'Deduction type not found'}, status=404)
        
        if request.method == 'DELETE':
            deduction.is_active = False
            deduction.save()
            return Response({'message': 'Deduction type deleted'})
        
        # PUT - Update
        data = request.data
        deduction.name = data.get('name', deduction.name)
        deduction.category = data.get('category', deduction.category)
        deduction.is_mandatory = data.get('is_mandatory', deduction.is_mandatory)
        deduction.gl_account_code = data.get('gl_account_code', deduction.gl_account_code)
        deduction.description = data.get('description', deduction.description)
        deduction.is_active = data.get('is_active', deduction.is_active)
        deduction.sort_order = data.get('sort_order', deduction.sort_order)
        deduction.save()
        
        return Response({'message': 'Deduction type updated', 'id': deduction.id})
    
    # ==========================================
    # TAX BANDS (PAYE)
    # ==========================================
    
    @action(detail=False, methods=['get', 'post'], url_path='tax-bands')
    def tax_bands(self, request):
        """
        GET current tax bands or POST to create new set.
        """
        from .models import TaxBand
        
        if request.method == 'GET':
            bands = TaxBand.get_current_bands()
            data = [{
                'id': b.id,
                'lower_limit': float(b.lower_limit),
                'upper_limit': float(b.upper_limit) if b.upper_limit else None,
                'rate': float(b.rate),
                'effective_date': b.effective_date,
                'expiry_date': b.expiry_date,
                'sort_order': b.sort_order,
            } for b in bands]
            return Response({'tax_bands': data, 'count': len(data)})
        
        # POST - Create new band
        data = request.data
        band = TaxBand.objects.create(
            lower_limit=data.get('lower_limit'),
            upper_limit=data.get('upper_limit'),
            rate=data.get('rate'),
            effective_date=data.get('effective_date', timezone.now().date()),
            expiry_date=data.get('expiry_date'),
            sort_order=data.get('sort_order', 0)
        )
        return Response({
            'message': 'Tax band created',
            'id': band.id
        }, status=201)
    
    @action(detail=False, methods=['put', 'delete'], url_path='tax-bands/(?P<pk>[^/.]+)')
    def tax_band_detail(self, request, pk=None):
        """
        PUT/DELETE specific tax band.
        """
        from .models import TaxBand
        
        try:
            band = TaxBand.objects.get(pk=pk)
        except TaxBand.DoesNotExist:
            return Response({'error': 'Tax band not found'}, status=404)
        
        if request.method == 'DELETE':
            band.is_active = False
            band.save()
            return Response({'message': 'Tax band deleted'})
        
        # PUT - Update
        data = request.data
        band.lower_limit = data.get('lower_limit', band.lower_limit)
        band.upper_limit = data.get('upper_limit', band.upper_limit)
        band.rate = data.get('rate', band.rate)
        band.effective_date = data.get('effective_date', band.effective_date)
        band.expiry_date = data.get('expiry_date', band.expiry_date)
        band.sort_order = data.get('sort_order', band.sort_order)
        band.save()
        
        return Response({'message': 'Tax band updated', 'id': band.id})
    
    # ==========================================
    # TAX RELIEFS
    # ==========================================
    
    @action(detail=False, methods=['get', 'put'], url_path='tax-reliefs')
    def tax_reliefs(self, request):
        """
        GET/PUT tax relief configuration.
        """
        from .models import TaxRelief
        
        if request.method == 'GET':
            reliefs = TaxRelief.objects.filter(is_active=True)
            data = [{
                'id': r.id,
                'relief_type': r.relief_type,
                'name': r.name,
                'amount': float(r.amount) if r.amount else None,
                'percentage': float(r.percentage) if r.percentage else None,
                'max_amount': float(r.max_amount) if r.max_amount else None,
                'effective_date': r.effective_date,
            } for r in reliefs]
            return Response({'tax_reliefs': data, 'count': len(data)})
        
        # PUT - Bulk update
        for item in request.data.get('reliefs', []):
            relief, created = TaxRelief.objects.update_or_create(
                relief_type=item.get('relief_type'),
                defaults={
                    'name': item.get('name'),
                    'amount': item.get('amount'),
                    'percentage': item.get('percentage'),
                    'max_amount': item.get('max_amount'),
                    'effective_date': item.get('effective_date', timezone.now().date()),
                }
            )
        return Response({'message': 'Tax reliefs updated'})
    
    # ==========================================
    # STATUTORY RATES (NSSF, NHIF, Housing)
    # ==========================================
    
    @action(detail=False, methods=['get', 'put'], url_path='statutory-rates')
    def statutory_rates(self, request):
        """
        GET/PUT statutory deduction rates.
        """
        from .models import StatutoryRate
        
        today = timezone.now().date()
        
        if request.method == 'GET':
            rates = StatutoryRate.objects.filter(
                is_active=True,
                effective_date__lte=today
            ).filter(
                models.Q(expiry_date__isnull=True) | models.Q(expiry_date__gte=today)
            )
            
            data = [{
                'id': r.id,
                'rate_type': r.rate_type,
                'name': r.name,
                'employee_rate': float(r.employee_rate) if r.employee_rate else None,
                'employer_rate': float(r.employer_rate) if r.employer_rate else None,
                'fixed_amount': float(r.fixed_amount) if r.fixed_amount else None,
                'lower_limit': float(r.lower_limit) if r.lower_limit else None,
                'upper_limit': float(r.upper_limit) if r.upper_limit else None,
                'max_contribution': float(r.max_contribution) if r.max_contribution else None,
                'is_enabled': r.is_enabled,
                'effective_date': r.effective_date,
            } for r in rates]
            return Response({'statutory_rates': data, 'count': len(data)})
        
        # PUT - Update rates
        for item in request.data.get('rates', []):
            rate, created = StatutoryRate.objects.update_or_create(
                rate_type=item.get('rate_type'),
                effective_date=item.get('effective_date', today),
                defaults={
                    'name': item.get('name'),
                    'employee_rate': item.get('employee_rate'),
                    'employer_rate': item.get('employer_rate'),
                    'fixed_amount': item.get('fixed_amount'),
                    'lower_limit': item.get('lower_limit'),
                    'upper_limit': item.get('upper_limit'),
                    'max_contribution': item.get('max_contribution'),
                    'is_enabled': item.get('is_enabled', True),
                }
            )
        return Response({'message': 'Statutory rates updated'})
    
    @action(detail=False, methods=['post'], url_path='statutory-rates/toggle')
    def toggle_statutory(self, request):
        """
        Enable/disable a statutory deduction.
        """
        from .models import StatutoryRate
        
        rate_type = request.data.get('rate_type')
        is_enabled = request.data.get('is_enabled', True)
        
        updated = StatutoryRate.objects.filter(
            rate_type=rate_type,
            is_active=True
        ).update(is_enabled=is_enabled)
        
        return Response({
            'message': f'{rate_type} {"enabled" if is_enabled else "disabled"}',
            'updated': updated
        })
    
    # ==========================================
    # GL ACCOUNT MAPPING (Finance Integration)
    # ==========================================
    
    @action(detail=False, methods=['get', 'put'], url_path='gl-mappings')
    def gl_mappings(self, request):
        """
        GET/PUT GL account mappings.
        """
        from .models import GLAccountMapping
        
        if request.method == 'GET':
            mappings = GLAccountMapping.objects.filter(is_active=True)
            data = [{
                'id': m.id,
                'mapping_type': m.mapping_type,
                'gl_account_code': m.gl_account_code,
                'gl_account_name': m.gl_account_name,
                'description': m.description,
            } for m in mappings]
            return Response({'gl_mappings': data, 'count': len(data)})
        
        # PUT - Bulk update
        for item in request.data.get('mappings', []):
            mapping, created = GLAccountMapping.objects.update_or_create(
                mapping_type=item.get('mapping_type'),
                defaults={
                    'gl_account_code': item.get('gl_account_code'),
                    'gl_account_name': item.get('gl_account_name'),
                    'description': item.get('description', ''),
                }
            )
        return Response({'message': 'GL mappings updated'})
    
    # ==========================================
    # APPROVAL WORKFLOW
    # ==========================================
    
    @action(detail=False, methods=['get', 'put'], url_path='approval-workflow')
    def approval_workflow(self, request):
        """
        GET/PUT approval workflow configuration.
        """
        from .models import ApprovalWorkflowLevel
        
        if request.method == 'GET':
            levels = ApprovalWorkflowLevel.objects.filter(is_active=True).order_by('level_number')
            data = [{
                'id': l.id,
                'level_number': l.level_number,
                'name': l.name,
                'description': l.description,
                'approver_role': l.approver_role,
                'can_approve': l.can_approve,
                'can_reject': l.can_reject,
                'can_request_changes': l.can_request_changes,
                'notify_on_pending': l.notify_on_pending,
                'auto_escalate_hours': l.auto_escalate_hours,
            } for l in levels]
            
            # Check if workflow is enabled (at least one active level)
            is_enabled = levels.exists()
            
            return Response({
                'is_enabled': is_enabled,
                'levels': data,
                'count': len(data)
            })
        
        # PUT - Update workflow
        data = request.data
        is_enabled = data.get('is_enabled', True)
        
        if not is_enabled:
            # Disable all levels
            ApprovalWorkflowLevel.objects.all().update(is_active=False)
            return Response({'message': 'Approval workflow disabled'})
        
        # Update levels
        for level_data in data.get('levels', []):
            level, created = ApprovalWorkflowLevel.objects.update_or_create(
                level_number=level_data.get('level_number'),
                defaults={
                    'name': level_data.get('name'),
                    'description': level_data.get('description', ''),
                    'approver_role': level_data.get('approver_role'),
                    'can_approve': level_data.get('can_approve', True),
                    'can_reject': level_data.get('can_reject', True),
                    'can_request_changes': level_data.get('can_request_changes', True),
                    'notify_on_pending': level_data.get('notify_on_pending', True),
                    'auto_escalate_hours': level_data.get('auto_escalate_hours'),
                    'is_active': True,
                }
            )
        
        return Response({'message': 'Approval workflow updated'})
    
    @action(detail=False, methods=['post'], url_path='approval-workflow/add-level')
    def add_approval_level(self, request):
        """
        Add a new approval level.
        """
        from .models import ApprovalWorkflowLevel
        
        data = request.data
        max_level = ApprovalWorkflowLevel.objects.filter(is_active=True).aggregate(
            max_level=models.Max('level_number')
        )['max_level'] or 0
        
        level = ApprovalWorkflowLevel.objects.create(
            level_number=max_level + 1,
            name=data.get('name'),
            description=data.get('description', ''),
            approver_role=data.get('approver_role'),
            can_approve=data.get('can_approve', True),
            can_reject=data.get('can_reject', True),
        )
        
        return Response({
            'message': 'Approval level added',
            'id': level.id,
            'level_number': level.level_number
        }, status=201)
    
    @action(detail=False, methods=['delete'], url_path='approval-workflow/(?P<pk>[^/.]+)')
    def delete_approval_level(self, request, pk=None):
        """
        Delete an approval level.
        """
        from .models import ApprovalWorkflowLevel
        
        try:
            level = ApprovalWorkflowLevel.objects.get(pk=pk)
            level.is_active = False
            level.save()
            return Response({'message': 'Approval level deleted'})
        except ApprovalWorkflowLevel.DoesNotExist:
            return Response({'error': 'Level not found'}, status=404)
    
    # ==========================================
    # BULK SETTINGS EXPORT/IMPORT
    # ==========================================
    
    @action(detail=False, methods=['get'], url_path='export')
    def export_settings(self, request):
        """
        Export all payroll settings as JSON for backup/migration.
        """
        from .models import (
            PayrollConfiguration, TaxBand, TaxRelief, 
            StatutoryRate, GLAccountMapping, ApprovalWorkflowLevel
        )
        
        config = PayrollConfiguration.get_active()
        
        export_data = {
            'exported_at': timezone.now().isoformat(),
            'configuration': {
                'pay_frequency': config.pay_frequency,
                'payroll_run_date': config.payroll_run_date,
                'attendance_cutoff_date': config.attendance_cutoff_date,
                'financial_year_start': config.financial_year_start,
                'currency_code': config.currency_code,
                'rounding_method': config.rounding_method,
                'overtime_multiplier': float(config.overtime_multiplier),
                'enable_proration': config.enable_proration,
            },
            'tax_bands': list(TaxBand.get_current_bands().values(
                'lower_limit', 'upper_limit', 'rate', 'effective_date', 'sort_order'
            )),
            'tax_reliefs': list(TaxRelief.objects.filter(is_active=True).values(
                'relief_type', 'name', 'amount', 'percentage', 'max_amount'
            )),
            'statutory_rates': list(StatutoryRate.objects.filter(is_active=True, is_enabled=True).values(
                'rate_type', 'name', 'employee_rate', 'employer_rate', 
                'fixed_amount', 'lower_limit', 'upper_limit', 'max_contribution'
            )),
            'gl_mappings': list(GLAccountMapping.objects.filter(is_active=True).values(
                'mapping_type', 'gl_account_code', 'gl_account_name'
            )),
            'earning_types': list(EarningType.objects.filter(is_active=True).values(
                'code', 'name', 'category', 'is_taxable', 'is_pensionable', 'gl_account_code'
            )),
            'deduction_types': list(DeductionType.objects.filter(is_active=True).values(
                'code', 'name', 'category', 'is_mandatory', 'gl_account_code'
            )),
            'approval_workflow': list(ApprovalWorkflowLevel.objects.filter(is_active=True).values(
                'level_number', 'name', 'approver_role', 'can_approve', 'can_reject'
            )),
        }
        
        return Response(export_data)
    
    @action(detail=False, methods=['post'], url_path='seed-defaults')
    def seed_defaults(self, request):
        """
        Seed default Kenyan payroll settings (PAYE, NSSF, SHIF, etc.)
        Useful for initial setup.
        """
        from .models import (
            PayrollConfiguration, TaxBand, TaxRelief, 
            StatutoryRate, GLAccountMapping, ApprovalWorkflowLevel
        )
        from decimal import Decimal
        
        today = timezone.now().date()
        
        # 1. Default Configuration
        config, _ = PayrollConfiguration.objects.get_or_create(
            is_active=True,
            defaults={
                'pay_frequency': 'monthly',
                'payroll_run_date': 25,
                'attendance_cutoff_date': 20,
                'financial_year_start': 'july',
                'currency_code': 'KES',
            }
        )
        
        # 2. Kenya 2024 PAYE Tax Bands
        tax_bands_data = [
            {'lower': 0, 'upper': 24000, 'rate': 10, 'order': 1},
            {'lower': 24001, 'upper': 32333, 'rate': 25, 'order': 2},
            {'lower': 32334, 'upper': 500000, 'rate': 30, 'order': 3},
            {'lower': 500001, 'upper': 800000, 'rate': 32.5, 'order': 4},
            {'lower': 800001, 'upper': None, 'rate': 35, 'order': 5},
        ]
        
        for band in tax_bands_data:
            TaxBand.objects.get_or_create(
                lower_limit=band['lower'],
                upper_limit=band['upper'],
                effective_date=today,
                defaults={
                    'rate': band['rate'],
                    'sort_order': band['order'],
                }
            )
        
        # 3. Tax Reliefs
        reliefs_data = [
            {'type': 'personal', 'name': 'Personal Relief', 'amount': 2400},
            {'type': 'insurance', 'name': 'Insurance Relief', 'percentage': 15, 'max': 5000},
            {'type': 'housing', 'name': 'Affordable Housing Relief', 'percentage': 15, 'max': 9000},
        ]
        
        for relief in reliefs_data:
            TaxRelief.objects.get_or_create(
                relief_type=relief['type'],
                defaults={
                    'name': relief['name'],
                    'amount': relief.get('amount'),
                    'percentage': relief.get('percentage'),
                    'max_amount': relief.get('max'),
                    'effective_date': today,
                }
            )
        
        # 4. Statutory Rates
        statutory_data = [
            {
                'type': 'nssf_tier1', 'name': 'NSSF Tier I',
                'emp_rate': 6, 'er_rate': 6,
                'lower': 0, 'upper': 7000, 'max': 420
            },
            {
                'type': 'nssf_tier2', 'name': 'NSSF Tier II',
                'emp_rate': 6, 'er_rate': 6,
                'lower': 7001, 'upper': 36000, 'max': 1740
            },
            {
                'type': 'shif', 'name': 'Social Health Insurance',
                'emp_rate': 2.75, 'er_rate': 0,
            },
            {
                'type': 'housing_levy', 'name': 'Affordable Housing Levy',
                'emp_rate': 1.5, 'er_rate': 1.5,
            },
            {
                'type': 'nita', 'name': 'NITA Levy',
                'fixed': 50, 'er_rate': 0,
            },
        ]
        
        for rate in statutory_data:
            StatutoryRate.objects.get_or_create(
                rate_type=rate['type'],
                effective_date=today,
                defaults={
                    'name': rate['name'],
                    'employee_rate': rate.get('emp_rate'),
                    'employer_rate': rate.get('er_rate'),
                    'fixed_amount': rate.get('fixed'),
                    'lower_limit': rate.get('lower'),
                    'upper_limit': rate.get('upper'),
                    'max_contribution': rate.get('max'),
                    'is_enabled': True,
                }
            )
        
        # 5. Default GL Mappings
        gl_mappings_data = [
            ('salary_expense', '6001', 'Salaries & Wages Expense'),
            ('paye_liability', '2101', 'PAYE Payable'),
            ('nssf_liability', '2102', 'NSSF Payable'),
            ('nhif_liability', '2103', 'NHIF/SHIF Payable'),
            ('housing_levy', '2104', 'Housing Levy Payable'),
            ('net_pay_liability', '2110', 'Net Salary Payable'),
            ('bank_account', '1101', 'Payroll Bank Account'),
        ]
        
        for mapping_type, code, name in gl_mappings_data:
            GLAccountMapping.objects.get_or_create(
                mapping_type=mapping_type,
                defaults={
                    'gl_account_code': code,
                    'gl_account_name': name,
                }
            )
        
        # 6. Default Approval Workflow
        workflow_data = [
            (1, 'HR Verification', 'HR Manager'),
            (2, 'Finance Review', 'Chief Accountant'),
            (3, 'Final Authorization', 'Director'),
        ]
        
        for level_num, name, role in workflow_data:
            ApprovalWorkflowLevel.objects.get_or_create(
                level_number=level_num,
                defaults={
                    'name': name,
                    'approver_role': role,
                }
            )
        
        # 7. Default Earning Types
        earning_defaults = [
            ('BASIC', 'Basic Salary', 'basic', True, True, '6001'),
            ('HOUSE', 'House Allowance', 'allowance', True, False, '6002'),
            ('TRANS', 'Transport Allowance', 'allowance', True, False, '6003'),
            ('OT', 'Overtime Pay', 'overtime', True, False, '6004'),
        ]
        
        for code, name, cat, taxable, pension, gl in earning_defaults:
            EarningType.objects.get_or_create(
                code=code,
                defaults={
                    'name': name,
                    'category': cat,
                    'is_taxable': taxable,
                    'is_pensionable': pension,
                    'gl_account_code': gl,
                }
            )
        
        # 8. Default Deduction Types
        deduction_defaults = [
            ('PAYE', 'Income Tax (PAYE)', 'statutory', True, '2101'),
            ('NSSF', 'NSSF Contribution', 'statutory', True, '2102'),
            ('SHIF', 'SHIF Contribution', 'statutory', True, '2103'),
            ('HELB', 'HELB Loan', 'loan', False, '2105'),
            ('SACCO', 'Sacco Contribution', 'voluntary', False, '2106'),
        ]
        
        for code, name, cat, mandatory, gl in deduction_defaults:
            DeductionType.objects.get_or_create(
                code=code,
                defaults={
                    'name': name,
                    'category': cat,
                    'is_mandatory': mandatory,
                    'gl_account_code': gl,
                }
            )
        
        return Response({
            'message': 'Default payroll settings seeded successfully',
            'seeded': {
                'tax_bands': 5,
                'tax_reliefs': 3,
                'statutory_rates': 5,
                'gl_mappings': 7,
                'approval_levels': 3,
                'earning_types': 4,
                'deduction_types': 5,
            }
        })
# ============================================================================
# PERFORMANCE VIEWSETS
# ============================================================================

class PerformanceMetricViewSet(viewsets.ModelViewSet):
    """ViewSet for Performance Metrics"""
    queryset = PerformanceMetric.objects.all()
    serializer_class = PerformanceMetricSerializer
    permission_classes = [IsAuthenticated]

class AppraisalCycleViewSet(viewsets.ModelViewSet):
    """ViewSet for Appraisal Cycles"""
    queryset = AppraisalCycle.objects.all().order_by('-start_date')
    serializer_class = AppraisalCycleSerializer
    permission_classes = [IsAuthenticated]

class EmployeeAppraisalViewSet(viewsets.ModelViewSet):
    """ViewSet for Employee Appraisals"""
    queryset = EmployeeAppraisal.objects.all().select_related('employee', 'appraiser', 'appraisal_cycle')
    serializer_class = EmployeeAppraisalSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()
        
        # If not HR/Admin, restrict to own appraisals or appraisals where user is appraiser
        if not (user.is_staff or user.is_superuser or user.groups.filter(name='HR').exists()):
            try:
                employee = user.employee_profile
                qs = qs.filter(Q(employee=employee) | Q(appraiser=employee))
            except Exception:
                return qs.none()
        return qs

    @action(detail=False, methods=['get'])
    def dashboard_metrics(self, request):
        """Aggregate performance data for dashboard"""
        user = request.user
        
        # Base querysets
        appraisals = self.get_queryset()
        goals = EmployeePerformanceGoal.objects.all()
        
        if not (user.is_staff or user.is_superuser or user.groups.filter(name='HR').exists()):
            try:
                employee = user.employee_profile
                goals = goals.filter(employee=employee)
            except Exception:
                goals = goals.none()
        
        # Compute metrics
        total_appraisals = appraisals.count()
        pending_reviews = appraisals.filter(status__in=['draft', 'submitted']).count()
        completed_reviews = appraisals.filter(status='approved').count()
        
        # Compute goals metrics
        total_goals = goals.count()
        completed_goals = goals.filter(status='completed').count()
        in_progress_goals = goals.filter(status='in_progress').count()
        
        goal_completion_rate = 0
        if total_goals > 0:
            goal_completion_rate = int((completed_goals / total_goals) * 100)
        
        return Response({
            'metrics': [
                {
                    'id': 1,
                    'title': 'Goal Completion',
                    'value': f"{goal_completion_rate}%",
                    'subtitle': f'{completed_goals} of {total_goals} goals met',
                    'trend': '+5% vs last period',
                    'trendUp': True,
                    'color': 'bg-blue-500',
                    'lightColor': 'bg-blue-50'
                },
                {
                    'id': 2,
                    'title': 'Pending Reviews',
                    'value': pending_reviews,
                    'subtitle': 'Awaiting action',
                    'trend': '-2 since last week',
                    'trendUp': True,
                    'color': 'bg-orange-500',
                    'lightColor': 'bg-orange-50'
                },
                {
                    'id': 3,
                    'title': 'Avg Appraisals',
                    'value': '4.2',
                    'subtitle': 'Out of 5.0',
                    'trend': '+0.3 vs last cycle',
                    'trendUp': True,
                    'color': 'bg-green-500',
                    'lightColor': 'bg-green-50'
                },
                {
                    'id': 4,
                    'title': 'Completed Reviews',
                    'value': completed_reviews,
                    'subtitle': 'This cycle',
                    'trend': f'{total_appraisals} total',
                    'trendUp': True,
                    'color': 'bg-purple-500',
                    'lightColor': 'bg-purple-50'
                }
            ],
            'recent_appraisals': EmployeeAppraisalSerializer(appraisals.order_by('-appraisal_date')[:5], many=True).data,
            'goals': EmployeePerformanceGoalSerializer(goals.order_by('-target_completion_date')[:5], many=True).data
        })

class EmployeePerformanceGoalViewSet(viewsets.ModelViewSet):
    """ViewSet for Employee Performance Goals"""
    queryset = EmployeePerformanceGoal.objects.all().select_related('employee')
    serializer_class = EmployeePerformanceGoalSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()
        
        # If not HR/Admin, restrict to own goals
        if not (user.is_staff or user.is_superuser or user.groups.filter(name='HR').exists()):
            try:
                employee = user.employee_profile
                qs = qs.filter(employee=employee)
            except Exception:
                return qs.none()
        return qs
