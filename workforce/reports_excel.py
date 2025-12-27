"""
HR & Payroll Excel Report Generators
Part 2: Excel Reports using openpyxl
"""

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side
)
from openpyxl.utils import get_column_letter
from io import BytesIO
from decimal import Decimal


class ExcelReportGenerator:
    """Base class for Excel report generation"""
    
    def __init__(self):
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        
        # Define styles
        self.header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(
            start_color='1a237e',
            end_color='1a237e',
            fill_type='solid'
        )
        
        self.subheader_font = Font(name='Arial', size=11, bold=True)
        self.subheader_fill = PatternFill(
            start_color='E3F2FD',
            end_color='E3F2FD',
            fill_type='solid'
        )
        
        self.title_font = Font(name='Arial', size=16, bold=True)
        self.normal_font = Font(name='Arial', size=10)
        
        self.center_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        self.left_alignment = Alignment(horizontal='left', vertical='center')
        self.right_alignment = Alignment(horizontal='right', vertical='center')
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _set_column_widths(self, widths):
        """Set column widths"""
        for idx, width in enumerate(widths, 1):
            self.worksheet.column_dimensions[get_column_letter(idx)].width = width
    
    def _merge_and_style_header(self, row, start_col, end_col, text, style='title'):
        """Merge cells and apply header styling"""
        cell = self.worksheet.cell(row=row, column=start_col)
        cell.value = text
        
        if end_col > start_col:
            self.worksheet.merge_cells(
                start_row=row,
                start_column=start_col,
                end_row=row,
                end_column=end_col
            )
        
        if style == 'title':
            cell.font = self.title_font
            cell.alignment = self.center_alignment
        elif style == 'header':
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
        elif style == 'subheader':
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.alignment = self.center_alignment
    
    def _apply_table_styling(self, start_row, end_row, start_col, end_col):
        """Apply table styling to a range"""
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = self.worksheet.cell(row=row, column=col)
                cell.border = self.border
                cell.alignment = self.left_alignment
                
                if row == start_row:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.center_alignment
    
    def generate_response(self, filename):
        """Generate HTTP response with Excel file"""
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(buffer.getvalue())
        buffer.close()
        
        return response


class PayrollRegisterExcelGenerator(ExcelReportGenerator):
    """Generate payroll register Excel report"""
    
    def generate(self, payroll_period):
        """Generate payroll register"""
        self.worksheet.title = 'Payroll Register'
        
        # Title
        self._merge_and_style_header(1, 1, 10, 'PAYROLL REGISTER', 'title')
        self._merge_and_style_header(
            2, 1, 10,
            f"Period: {payroll_period.period_name} | Payment Date: {payroll_period.payment_date}",
            'subheader'
        )
        
        # Headers
        headers = [
            'Emp No', 'Name', 'Department', 'Basic Salary',
            'Allowances', 'Gross Pay', 'Deductions', 'Tax',
            'Net Pay', 'Bank Account'
        ]
        
        row = 4
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Data
        calculations = payroll_period.calculations.select_related(
            'employee', 'employee__department'
        ).order_by('employee__employee_no')
        
        row = 5
        total_gross = Decimal('0')
        total_deductions = Decimal('0')
        total_net = Decimal('0')
        
        for calc in calculations:
            data = [
                calc.employee.employee_no,
                calc.employee.get_full_name(),
                calc.employee.department.name if calc.employee.department else 'N/A',
                float(calc.basic_salary),
                float(calc.total_allowances),
                float(calc.gross_pay),
                float(calc.total_deductions),
                float(calc.tax_amount),
                float(calc.net_pay),
                calc.bank_account_number
            ]
            
            for col, value in enumerate(data, 1):
                cell = self.worksheet.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                
                # Format currency columns
                if col in [4, 5, 6, 7, 8, 9]:
                    cell.number_format = '#,##0.00'
                    cell.alignment = self.right_alignment
                else:
                    cell.alignment = self.left_alignment
            
            total_gross += calc.gross_pay
            total_deductions += calc.total_deductions
            total_net += calc.net_pay
            row += 1
        
        # Totals row
        totals_row = row
        self.worksheet.cell(row=totals_row, column=1).value = 'TOTAL'
        self.worksheet.cell(row=totals_row, column=2).value = f"{calculations.count()} employees"
        self.worksheet.cell(row=totals_row, column=6).value = float(total_gross)
        self.worksheet.cell(row=totals_row, column=7).value = float(total_deductions)
        self.worksheet.cell(row=totals_row, column=9).value = float(total_net)
        
        # Style totals row
        for col in range(1, 11):
            cell = self.worksheet.cell(row=totals_row, column=col)
            cell.font = Font(name='Arial', size=10, bold=True)
            cell.fill = PatternFill(
                start_color='E3F2FD',
                end_color='E3F2FD',
                fill_type='solid'
            )
            cell.border = self.border
            
            if col in [6, 7, 9]:
                cell.number_format = '#,##0.00'
                cell.alignment = self.right_alignment
        
        # Summary section
        summary_row = totals_row + 2
        self._merge_and_style_header(
            summary_row, 1, 4,
            'PAYMENT SUMMARY',
            'subheader'
        )
        
        summary_data = [
            ['Total Gross Pay', float(total_gross)],
            ['Total Deductions', float(total_deductions)],
            ['Total Net Pay', float(total_net)],
            ['Number of Employees', calculations.count()],
            ['Average Net Pay', float(total_net / calculations.count() if calculations.count() > 0 else 0)]
        ]
        
        summary_start = summary_row + 1
        for idx, (label, value) in enumerate(summary_data):
            label_cell = self.worksheet.cell(row=summary_start + idx, column=1)
            value_cell = self.worksheet.cell(row=summary_start + idx, column=2)
            
            label_cell.value = label
            value_cell.value = value
            
            label_cell.font = self.subheader_font
            value_cell.font = self.normal_font
            
            if isinstance(value, (int, float, Decimal)) and idx < 3:
                value_cell.number_format = '#,##0.00'
                value_cell.alignment = self.right_alignment
        
        # Set column widths
        self._set_column_widths([12, 25, 20, 12, 12, 12, 12, 12, 12, 18])
        
        return self.generate_response(
            f"payroll_register_{payroll_period.period_name}.xlsx"
        )


class EmployeeListExcelGenerator(ExcelReportGenerator):
    """Generate employee list Excel report"""
    
    def generate(self, employees):
        """Generate employee list"""
        self.worksheet.title = 'Employee List'
        
        # Title
        self._merge_and_style_header(
            1, 1, 12,
            'EMPLOYEE MASTER LIST',
            'title'
        )
        self._merge_and_style_header(
            2, 1, 12,
            f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')} | Total: {employees.count()}",
            'subheader'
        )
        
        # Headers
        headers = [
            'Emp No', 'First Name', 'Last Name', 'Email',
            'Phone', 'Department', 'Job Grade', 'Category',
            'Status', 'Hire Date', 'Confirmation Date', 'Termination Date'
        ]
        
        row = 4
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Data
        row = 5
        for emp in employees:
            data = [
                emp.employee_no,
                emp.first_name,
                emp.last_name,
                emp.official_email,
                emp.phone_primary,
                emp.department.name if emp.department else 'N/A',
                emp.job_grade.name if emp.job_grade else 'N/A',
                emp.get_employee_category_display(),
                emp.get_employment_status_display(),
                emp.hire_date,
                emp.confirmation_date,
                emp.termination_date
            ]
            
            for col, value in enumerate(data, 1):
                cell = self.worksheet.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                
                # Format date columns
                if col in [10, 11, 12] and value:
                    cell.number_format = 'YYYY-MM-DD'
                
                cell.alignment = self.left_alignment
            
            row += 1
        
        # Set column widths
        self._set_column_widths([12, 15, 15, 25, 15, 20, 15, 15, 15, 12, 12, 12])
        
        # Add summary sheet
        summary_ws = self.workbook.create_sheet('Summary')
        
        from django.db.models import Count
        
        # By Status
        summary_ws['A1'] = 'By Employment Status'
        summary_ws['A1'].font = self.subheader_font
        summary_ws['A2'] = 'Status'
        summary_ws['B2'] = 'Count'
        
        status_summary = employees.values('employment_status').annotate(
            count=Count('id')
        )
        
        row = 3
        for item in status_summary:
            summary_ws[f'A{row}'] = item['employment_status'].replace('_', ' ').title()
            summary_ws[f'B{row}'] = item['count']
            row += 1
        
        # By Category
        summary_ws['D1'] = 'By Employee Category'
        summary_ws['D1'].font = self.subheader_font
        summary_ws['D2'] = 'Category'
        summary_ws['E2'] = 'Count'
        
        category_summary = employees.values('employee_category').annotate(
            count=Count('id')
        )
        
        row = 3
        for item in category_summary:
            summary_ws[f'D{row}'] = item['employee_category'].replace('_', ' ').title()
            summary_ws[f'E{row}'] = item['count']
            row += 1
        
        return self.generate_response(
            f"employee_list_{timezone.now().strftime('%Y%m%d')}.xlsx"
        )


class AttendanceReportExcelGenerator(ExcelReportGenerator):
    """Generate attendance report Excel"""
    
    def generate(self, records, start_date, end_date):
        """Generate attendance report"""
        self.worksheet.title = 'Attendance Report'
        
        # Title
        self._merge_and_style_header(
            1, 1, 8,
            'ATTENDANCE REPORT',
            'title'
        )
        self._merge_and_style_header(
            2, 1, 8,
            f"Period: {start_date} to {end_date}",
            'subheader'
        )
        
        # Headers
        headers = [
            'Employee No', 'Name', 'Department', 'Date',
            'Check In', 'Check Out', 'Hours', 'Status'
        ]
        
        row = 4
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Data
        row = 5
        for record in records:
            data = [
                record.employee.employee_no,
                record.employee.get_full_name(),
                record.employee.department.name if record.employee.department else 'N/A',
                record.attendance_date,
                record.check_in_time,
                record.check_out_time,
                float(record.total_hours) if record.total_hours else 0,
                record.get_status_display()
            ]
            
            for col, value in enumerate(data, 1):
                cell = self.worksheet.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                
                if col == 4:  # Date
                    cell.number_format = 'YYYY-MM-DD'
                elif col in [5, 6]:  # Time
                    cell.number_format = 'HH:MM'
                elif col == 7:  # Hours
                    cell.number_format = '0.00'
                    cell.alignment = self.right_alignment
                else:
                    cell.alignment = self.left_alignment
            
            row += 1
        
        # Set column widths
        self._set_column_widths([15, 25, 20, 12, 10, 10, 10, 15])
        
        # Summary sheet
        summary_ws = self.workbook.create_sheet('Summary')
        
        from django.db.models import Count, Sum
        from collections import defaultdict
        
        # By employee
        emp_summary = defaultdict(lambda: {
            'present': 0, 'absent': 0, 'late': 0,
            'leave': 0, 'total_hours': Decimal('0')
        })
        
        for record in records:
            emp_key = (record.employee.employee_no, record.employee.get_full_name())
            
            if record.status == 'present':
                emp_summary[emp_key]['present'] += 1
            elif record.status == 'absent':
                emp_summary[emp_key]['absent'] += 1
            elif record.status == 'late':
                emp_summary[emp_key]['late'] += 1
            elif record.status == 'on_leave':
                emp_summary[emp_key]['leave'] += 1
            
            emp_summary[emp_key]['total_hours'] += record.total_hours or Decimal('0')
        
        # Write summary
        summary_ws['A1'] = 'Employee Summary'
        summary_ws['A1'].font = self.subheader_font
        
        headers = ['Emp No', 'Name', 'Present', 'Absent', 'Late', 'Leave', 'Total Hours']
        for col, header in enumerate(headers, 1):
            cell = summary_ws.cell(row=2, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        row = 3
        for (emp_no, name), data in sorted(emp_summary.items()):
            summary_ws[f'A{row}'] = emp_no
            summary_ws[f'B{row}'] = name
            summary_ws[f'C{row}'] = data['present']
            summary_ws[f'D{row}'] = data['absent']
            summary_ws[f'E{row}'] = data['late']
            summary_ws[f'F{row}'] = data['leave']
            summary_ws[f'G{row}'] = float(data['total_hours'])
            summary_ws[f'G{row}'].number_format = '0.00'
            row += 1
        
        return self.generate_response(
            f"attendance_report_{start_date}_{end_date}.xlsx"
        )


class LeaveBalanceExcelGenerator(ExcelReportGenerator):
    """Generate leave balance Excel report"""
    
    def generate(self, balances, year=None):
        """Generate leave balance report"""
        year = year or timezone.now().year
        self.worksheet.title = 'Leave Balances'
        
        # Title
        self._merge_and_style_header(
            1, 1, 10,
            'EMPLOYEE LEAVE BALANCES',
            'title'
        )
        self._merge_and_style_header(
            2, 1, 10,
            f"Year: {year}",
            'subheader'
        )
        
        # Headers
        headers = [
            'Emp No', 'Name', 'Leave Type', 'Opening',
            'Accrued', 'Carried Fwd', 'Taken', 'Pending',
            'Encashed', 'Closing'
        ]
        
        row = 4
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Data
        row = 5
        for balance in balances:
            data = [
                balance.employee.employee_no,
                balance.employee.get_full_name(),
                balance.leave_type.name,
                float(balance.opening_balance),
                float(balance.accrued_days),
                float(balance.carried_forward_days),
                float(balance.taken_days),
                float(balance.pending_days),
                float(balance.encashed_days),
                float(balance.closing_balance)
            ]
            
            for col, value in enumerate(data, 1):
                cell = self.worksheet.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                
                # Format number columns
                if col > 3:
                    cell.number_format = '0.00'
                    cell.alignment = self.right_alignment
                else:
                    cell.alignment = self.left_alignment
            
            row += 1
        
        # Set column widths
        self._set_column_widths([12, 25, 15, 10, 10, 10, 10, 10, 10, 10])
        
        return self.generate_response(
            f"leave_balances_{year}.xlsx"
        )


class PayrollComparisonExcelGenerator(ExcelReportGenerator):
    """Generate payroll comparison Excel report"""
    
    def generate(self, periods):
        """Generate payroll comparison across multiple periods"""
        self.worksheet.title = 'Payroll Comparison'
        
        # Title
        self._merge_and_style_header(
            1, 1, len(periods) + 2,
            'PAYROLL COMPARISON REPORT',
            'title'
        )
        
        # Headers
        headers = ['Metric'] + [p.period_name for p in periods]
        
        row = 3
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Metrics
        metrics = [
            'Employee Count',
            'Total Gross Pay',
            'Total Deductions',
            'Total Net Pay',
            'Average Gross Pay',
            'Average Net Pay'
        ]
        
        row = 4
        for metric in metrics:
            cell = self.worksheet.cell(row=row, column=1)
            cell.value = metric
            cell.font = self.subheader_font
            cell.border = self.border
            
            for col, period in enumerate(periods, 2):
                calcs = period.calculations.all()
                
                if metric == 'Employee Count':
                    value = calcs.count()
                elif metric == 'Total Gross Pay':
                    value = float(calcs.aggregate(Sum('gross_pay'))['gross_pay__sum'] or 0)
                elif metric == 'Total Deductions':
                    value = float(calcs.aggregate(Sum('total_deductions'))['total_deductions__sum'] or 0)
                elif metric == 'Total Net Pay':
                    value = float(calcs.aggregate(Sum('net_pay'))['net_pay__sum'] or 0)
                elif metric == 'Average Gross Pay':
                    avg = calcs.aggregate(Avg('gross_pay'))['gross_pay__avg']
                    value = float(avg) if avg else 0
                else:  # Average Net Pay
                    avg = calcs.aggregate(Avg('net_pay'))['net_pay__avg']
                    value = float(avg) if avg else 0
                
                cell = self.worksheet.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                cell.alignment = self.right_alignment
                
                if metric != 'Employee Count':
                    cell.number_format = '#,##0.00'
            
            row += 1
        
        # Set column widths
        widths = [25] + [15] * len(periods)
        self._set_column_widths(widths)
        
        return self.generate_response(
            f"payroll_comparison_{timezone.now().strftime('%Y%m%d')}.xlsx"
        )