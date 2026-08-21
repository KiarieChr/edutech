from decimal import Decimal
from statistics import median

from django.db.models import Avg, Max, Min, Count, Sum, Q, F
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from .models import (
    GradingScale, GradingLevel, AssessmentType, Examination,
    StudentMark, TermResult, TermSubjectResult, ReportCardTemplate
)
from .serializers import (
    GradingScaleListSerializer, GradingScaleDetailSerializer,
    GradingLevelSerializer, AssessmentTypeSerializer,
    ExaminationListSerializer, ExaminationDetailSerializer,
    StudentMarkSerializer, BulkMarksSerializer,
    TermResultListSerializer, TermResultDetailSerializer,
    ReportCardTemplateSerializer,
)


# =============================================================================
# GRADING SCALE
# =============================================================================

class GradingScaleViewSet(viewsets.ModelViewSet):
    queryset = GradingScale.objects.select_related(
        'curriculum', 'curriculum_level'
    ).prefetch_related('levels').filter(is_deleted=False)
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'list':
            return GradingScaleListSerializer
        return GradingScaleDetailSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        curriculum = self.request.query_params.get('curriculum')
        level = self.request.query_params.get('curriculum_level')
        active = self.request.query_params.get('is_active')

        if curriculum:
            qs = qs.filter(curriculum_id=curriculum)
        if level:
            qs = qs.filter(curriculum_level_id=level)
        if active is not None:
            qs = qs.filter(is_active=active.lower() == 'true')
        return qs

    @action(detail=True, methods=['post'])
    def duplicate(self, request, pk=None):
        """Duplicate a grading scale with all its levels."""
        original = self.get_object()
        new_code = f"{original.code}_copy"
        new_name = f"{original.name} (Copy)"

        scale = GradingScale.objects.create(
            name=new_name,
            code=new_code,
            curriculum=original.curriculum,
            curriculum_level=original.curriculum_level,
            scale_type=original.scale_type,
            max_mark=original.max_mark,
            pass_mark=original.pass_mark,
            description=original.description,
            created_by=request.user,
        )
        for level in original.levels.all():
            GradingLevel.objects.create(
                scale=scale,
                grade=level.grade,
                label=level.label,
                min_mark=level.min_mark,
                max_mark=level.max_mark,
                points=level.points,
                order=level.order,
                color_hex=level.color_hex,
            )
        return Response(
            GradingScaleDetailSerializer(scale).data,
            status=status.HTTP_201_CREATED
        )


# =============================================================================
# ASSESSMENT TYPES
# =============================================================================

class AssessmentTypeViewSet(viewsets.ModelViewSet):
    queryset = AssessmentType.objects.select_related(
        'curriculum', 'curriculum_level'
    ).filter(is_deleted=False)
    serializer_class = AssessmentTypeSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None

    def get_queryset(self):
        qs = super().get_queryset()
        curriculum = self.request.query_params.get('curriculum')
        level = self.request.query_params.get('curriculum_level')

        if curriculum:
            qs = qs.filter(curriculum_id=curriculum)
        if level:
            qs = qs.filter(
                Q(curriculum_level_id=level) | Q(curriculum_level__isnull=True)
            )
        return qs


# =============================================================================
# EXAMINATIONS
# =============================================================================

class ExaminationViewSet(viewsets.ModelViewSet):
    queryset = Examination.objects.select_related(
        'class_session__grade', 'class_session__term', 'class_session__academic_year',
        'class_session__curriculum', 'subject', 'assessment_type',
        'grading_scale', 'teacher'
    ).filter(is_deleted=False)
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'list':
            return ExaminationListSerializer
        return ExaminationDetailSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        session = self.request.query_params.get('class_session')
        subject = self.request.query_params.get('subject')
        assessment_type = self.request.query_params.get('assessment_type')
        exam_status = self.request.query_params.get('status')
        academic_year = self.request.query_params.get('academic_year')
        term = self.request.query_params.get('term')
        grade = self.request.query_params.get('grade')

        if session:
            qs = qs.filter(class_session_id=session)
        if subject:
            qs = qs.filter(subject_id=subject)
        if assessment_type:
            qs = qs.filter(assessment_type_id=assessment_type)
        if exam_status:
            qs = qs.filter(status=exam_status)
        if academic_year:
            qs = qs.filter(class_session__academic_year_id=academic_year)
        if term:
            qs = qs.filter(class_session__term_id=term)
        if grade:
            qs = qs.filter(class_session__grade_id=grade)
        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    # --- Bulk marks entry ---
    @action(detail=True, methods=['post'])
    def bulk_marks(self, request, pk=None):
        """
        Enter/update marks for multiple students at once.
        Body: { "marks": [{ "student": 1, "raw_mark": 85.0, "is_absent": false }, ...] }
        """
        exam = self.get_object()
        serializer = BulkMarksSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        results = {'created': 0, 'updated': 0, 'errors': []}

        for entry in serializer.validated_data['marks']:
            student_id = entry.get('student')
            try:
                mark, created = StudentMark.objects.update_or_create(
                    examination=exam,
                    student_id=student_id,
                    defaults={
                        'raw_mark': entry.get('raw_mark'),
                        'is_absent': entry.get('is_absent', False),
                        'is_exempted': entry.get('is_exempted', False),
                        'teacher_remark': entry.get('teacher_remark', ''),
                        'stream_id': entry.get('stream'),
                        'entered_by': request.user,
                    }
                )
                if created:
                    results['created'] += 1
                else:
                    results['updated'] += 1
            except Exception as e:
                results['errors'].append({'student': student_id, 'error': str(e)})

        return Response(results)

    # --- Exam analysis ---
    @action(detail=True, methods=['get'])
    def analysis(self, request, pk=None):
        """Analysis for a single examination."""
        exam = self.get_object()
        marks = exam.marks.filter(is_exempted=False)

        scored = marks.filter(is_absent=False, raw_mark__isnull=False)
        raw_values = list(scored.values_list('normalized_mark', flat=True))

        total_students = marks.count()
        absent = marks.filter(is_absent=True).count()

        pass_mark = exam.grading_scale.pass_mark
        passed = scored.filter(normalized_mark__gte=pass_mark).count()

        # Grade distribution
        grade_dist = {}
        for level in exam.grading_scale.levels.order_by('order'):
            count = scored.filter(grade=level.grade).count()
            grade_dist[level.grade] = {
                'count': count,
                'label': level.label,
                'color': level.color_hex,
            }

        data = {
            'exam_id': exam.id,
            'exam_name': exam.name,
            'subject_name': exam.subject.name,
            'total_students': total_students,
            'marks_entered': scored.count(),
            'absent': absent,
            'mean_mark': float(sum(raw_values) / len(raw_values)) if raw_values else 0,
            'median_mark': float(median(raw_values)) if raw_values else None,
            'highest_mark': float(max(raw_values)) if raw_values else None,
            'lowest_mark': float(min(raw_values)) if raw_values else None,
            'pass_rate': round((passed / len(raw_values)) * 100, 1) if raw_values else 0,
            'grade_distribution': grade_dist,
        }
        return Response(data)

    # --- Publish marks ---
    @action(detail=True, methods=['post'])
    def publish(self, request, pk=None):
        """Move exam to published status."""
        exam = self.get_object()
        exam.status = 'published'
        exam.published_at = timezone.now()
        exam.published_by = request.user
        exam.save()
        return Response({'status': 'published'})

    # --- Get enrolled students for an exam (or a class session) ---
    @action(detail=True, methods=['get'])
    def students(self, request, pk=None):
        """
        Return all enrolled students for this exam's class session,
        with any existing marks pre-filled.
        """
        exam = self.get_object()
        from academics.models import StudentSessionEnrollment

        enrollments = StudentSessionEnrollment.objects.filter(
            session=exam.class_session,
            status='active',
        ).select_related('student__student', 'stream')

        # Get existing marks for this exam
        existing_marks = {
            m.student_id: m
            for m in exam.marks.select_related('student__student', 'stream').all()
        }

        students = []
        for enr in enrollments:
            mark = existing_marks.get(enr.student_id)
            students.append({
                'student_id': enr.student_id,
                'admission_number': enr.student.admission_number,
                'student_name': (
                    f"{enr.student.student.first_name} {enr.student.student.last_name}".strip()
                    or enr.student.student.username
                ),
                'stream_id': enr.stream_id,
                'stream_name': enr.stream.name if enr.stream else None,
                'raw_mark': float(mark.raw_mark) if mark and mark.raw_mark is not None else None,
                'normalized_mark': float(mark.normalized_mark) if mark and mark.normalized_mark is not None else None,
                'grade': mark.grade if mark else None,
                'points': float(mark.points) if mark and mark.points is not None else None,
                'grade_label': mark.grade_label if mark else None,
                'is_absent': mark.is_absent if mark else False,
                'is_exempted': mark.is_exempted if mark else False,
                'teacher_remark': mark.teacher_remark if mark else '',
            })

        return Response({
            'exam_id': exam.id,
            'exam_name': exam.name,
            'max_mark': float(exam.max_mark),
            'grading_scale': exam.grading_scale_id,
            'total_students': len(students),
            'students': students,
        })

    # --- Upload marks via CSV/Excel ---
    @action(detail=True, methods=['post'])
    def upload_marks(self, request, pk=None):
        """
        Upload marks from a CSV/Excel file.
        Expected columns: admission_number, raw_mark
        Optional columns: is_absent, teacher_remark
        """
        import csv
        import io

        exam = self.get_object()
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        filename = file.name.lower()
        rows = []

        try:
            if filename.endswith('.csv'):
                decoded = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(decoded))
                rows = list(reader)
            elif filename.endswith(('.xlsx', '.xls')):
                try:
                    import openpyxl
                except ImportError:
                    return Response(
                        {'error': 'openpyxl is required for Excel uploads. Install with: pip install openpyxl'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                wb = openpyxl.load_workbook(file, read_only=True)
                ws = wb.active
                headers = [str(cell.value or '').strip().lower().replace(' ', '_') for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            else:
                return Response({'error': 'Unsupported file format. Use .csv or .xlsx'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': f'Failed to parse file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        if not rows:
            return Response({'error': 'File is empty'}, status=status.HTTP_400_BAD_REQUEST)

        # Build admission_number → student_id map
        from student_management.models import Student
        from academics.models import StudentSessionEnrollment

        enrolled_students = StudentSessionEnrollment.objects.filter(
            session=exam.class_session, status='active',
        ).select_related('student')

        adm_to_student = {}
        for enr in enrolled_students:
            adm = enr.student.admission_number
            if adm:
                adm_to_student[str(adm).strip()] = enr.student_id

        results = {'created': 0, 'updated': 0, 'skipped': 0, 'errors': []}

        for i, row in enumerate(rows, start=2):
            adm = str(row.get('admission_number', '') or '').strip()
            if not adm:
                results['skipped'] += 1
                continue

            student_id = adm_to_student.get(adm)
            if not student_id:
                results['errors'].append({'row': i, 'admission_number': adm, 'error': 'Student not found or not enrolled'})
                continue

            raw_mark_str = str(row.get('raw_mark', '') or row.get('marks', '') or row.get('score', '') or '').strip()
            is_absent_str = str(row.get('is_absent', '') or '').strip().lower()
            remark = str(row.get('teacher_remark', '') or row.get('remarks', '') or row.get('remark', '') or '').strip()

            is_absent = is_absent_str in ('true', '1', 'yes', 'absent')

            if not raw_mark_str and not is_absent:
                results['skipped'] += 1
                continue

            try:
                raw_mark = float(raw_mark_str) if raw_mark_str else None
                mark, created = StudentMark.objects.update_or_create(
                    examination=exam,
                    student_id=student_id,
                    defaults={
                        'raw_mark': raw_mark,
                        'is_absent': is_absent,
                        'teacher_remark': remark,
                        'entered_by': request.user,
                    }
                )
                if created:
                    results['created'] += 1
                else:
                    results['updated'] += 1
            except Exception as e:
                results['errors'].append({'row': i, 'admission_number': adm, 'error': str(e)})

        return Response(results)

    # --- Generate/compute term results for all students in a class session ---
    @action(detail=False, methods=['post'])
    def compute_term_results(self, request):
        """
        Compute weighted term results for all students in a class session.
        Body: { "class_session": 1, "grading_scale": 1 }
        """
        session_id = request.data.get('class_session')
        scale_id = request.data.get('grading_scale')

        if not session_id or not scale_id:
            return Response(
                {'error': 'class_session and grading_scale are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        from academics.models import ClassSession, StudentSessionEnrollment
        session = ClassSession.objects.get(id=session_id)
        scale = GradingScale.objects.prefetch_related('levels').get(id=scale_id)
        levels_desc = scale.levels.order_by('-min_mark')

        # Get all exams for this session
        exams = Examination.objects.filter(
            class_session=session,
            status__in=['completed', 'published']
        ).select_related('assessment_type', 'subject')

        if not exams.exists():
            return Response({'error': 'No completed exams found'}, status=status.HTTP_400_BAD_REQUEST)

        # Get assessment type weights
        assessment_weights = {}
        for exam in exams:
            at = exam.assessment_type
            assessment_weights[at.id] = {
                'weight': at.weight,
                'code': at.code,
                'name': at.name,
            }

        # Get enrolled students
        enrollments = StudentSessionEnrollment.objects.filter(
            session=session, status='active'
        ).select_related('student')
        student_ids = [e.student_id for e in enrollments]

        # Build per-student, per-subject results
        results_created = 0
        for student_id in student_ids:
            # Get/create TermResult
            enrollment = enrollments.get(student_id=student_id)
            stream = getattr(enrollment, 'stream', None)
            term_result, _ = TermResult.objects.update_or_create(
                student_id=student_id,
                class_session=session,
                defaults={'stream': stream, 'created_by': request.user}
            )

            # Group marks by subject
            marks = StudentMark.objects.filter(
                examination__class_session=session,
                student_id=student_id,
                is_absent=False, is_exempted=False,
                raw_mark__isnull=False,
            ).select_related('examination__assessment_type', 'examination__subject')

            subject_marks = {}
            for mark in marks:
                subj_id = mark.examination.subject_id
                if subj_id not in subject_marks:
                    subject_marks[subj_id] = []
                subject_marks[subj_id].append(mark)

            total_weighted = Decimal('0')
            total_points = Decimal('0')
            subject_count = 0
            assessment_breakdown_all = {}

            for subj_id, subj_marks in subject_marks.items():
                weighted = Decimal('0')
                breakdown = {}

                for mark in subj_marks:
                    at = mark.examination.assessment_type
                    weight = at.weight / Decimal('100')
                    contribution = mark.normalized_mark * weight
                    weighted += contribution
                    breakdown[at.code] = float(mark.normalized_mark)

                # Look up grade
                subj_grade = ''
                subj_points = Decimal('0')
                subj_label = ''
                for level in levels_desc:
                    if weighted >= level.min_mark:
                        subj_grade = level.grade
                        subj_points = level.points
                        subj_label = level.label
                        break

                TermSubjectResult.objects.update_or_create(
                    term_result=term_result,
                    subject_id=subj_id,
                    defaults={
                        'weighted_mark': weighted,
                        'grade': subj_grade,
                        'points': subj_points,
                        'grade_label': subj_label,
                        'assessment_breakdown': breakdown,
                        'created_by': request.user,
                    }
                )

                total_weighted += weighted
                total_points += subj_points
                subject_count += 1

            # Update TermResult aggregates
            if subject_count > 0:
                term_result.total_marks = total_weighted
                term_result.total_points = total_points
                term_result.average_mark = total_weighted / subject_count
                term_result.average_points = total_points / subject_count
                term_result.subjects_taken = subject_count

                # Overall grade from average
                avg = term_result.average_mark
                for level in levels_desc:
                    if avg >= level.min_mark:
                        term_result.overall_grade = level.grade
                        term_result.overall_grade_label = level.label
                        break

                term_result.save()
            results_created += 1

        # Compute rankings
        self._compute_rankings(session)

        return Response({
            'message': f'Term results computed for {results_created} students',
            'class_session': session_id,
        })

    def _compute_rankings(self, session):
        """Compute class, stream, and grade-wide rankings."""
        results = TermResult.objects.filter(class_session=session).order_by('-average_mark')

        # Class rank
        for i, result in enumerate(results, 1):
            result.class_rank = i
            result.total_in_class = results.count()

        # Stream rank
        streams = results.values_list('stream', flat=True).distinct()
        for stream_id in streams:
            stream_results = results.filter(stream_id=stream_id).order_by('-average_mark')
            for i, result in enumerate(stream_results, 1):
                result.stream_rank = i
                result.total_in_stream = stream_results.count()

        # Grade rank (across all sessions for the same grade + term)
        grade = session.grade
        term = session.term
        grade_results = TermResult.objects.filter(
            class_session__grade=grade,
            class_session__term=term,
        ).order_by('-average_mark')
        total_grade = grade_results.count()
        for i, result in enumerate(grade_results, 1):
            result.grade_rank = i
            result.total_in_grade = total_grade

        # Bulk update
        TermResult.objects.bulk_update(
            list(results) + list(grade_results),
            ['class_rank', 'total_in_class', 'stream_rank', 'total_in_stream',
             'grade_rank', 'total_in_grade']
        )

        # Subject rankings
        subjects = TermSubjectResult.objects.filter(
            term_result__class_session=session
        ).values_list('subject_id', flat=True).distinct()

        for subj_id in subjects:
            subj_results = TermSubjectResult.objects.filter(
                term_result__class_session=session,
                subject_id=subj_id,
            ).order_by('-weighted_mark')
            total = subj_results.count()
            updates = []
            for i, sr in enumerate(subj_results, 1):
                sr.subject_rank = i
                sr.total_in_subject = total
                updates.append(sr)
            TermSubjectResult.objects.bulk_update(updates, ['subject_rank', 'total_in_subject'])


# =============================================================================
# TERM RESULTS
# =============================================================================

class TermResultViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = TermResult.objects.select_related(
        'student__student', 'class_session__grade', 'class_session__term',
        'class_session__academic_year', 'stream'
    ).prefetch_related('subject_results__subject').filter(is_deleted=False)
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'list':
            return TermResultListSerializer
        return TermResultDetailSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        session = self.request.query_params.get('class_session')
        student = self.request.query_params.get('student')
        published = self.request.query_params.get('is_published')

        if session:
            qs = qs.filter(class_session_id=session)
        if student:
            qs = qs.filter(student_id=student)
        if published is not None:
            qs = qs.filter(is_published=published.lower() == 'true')
        return qs

    @action(detail=False, methods=['post'])
    def publish(self, request):
        """Publish all term results for a class session."""
        session_id = request.data.get('class_session')
        if not session_id:
            return Response({'error': 'class_session required'}, status=status.HTTP_400_BAD_REQUEST)

        updated = TermResult.objects.filter(
            class_session_id=session_id, is_published=False
        ).update(is_published=True, published_at=timezone.now())

        return Response({'published': updated})

    @action(detail=False, methods=['get'])
    def class_analysis(self, request):
        """
        Class-wide analysis for a class session.
        Query params: class_session (required)
        """
        session_id = request.query_params.get('class_session')
        if not session_id:
            return Response({'error': 'class_session required'}, status=status.HTTP_400_BAD_REQUEST)

        from academics.models import ClassSession
        session = ClassSession.objects.select_related('grade', 'term', 'academic_year').get(id=session_id)

        results = TermResult.objects.filter(class_session=session)
        subject_results = TermSubjectResult.objects.filter(
            term_result__class_session=session
        ).select_related('subject')

        # Subject means
        subject_stats = subject_results.values(
            'subject__name', 'subject__code'
        ).annotate(
            mean=Avg('weighted_mark'),
            highest=Max('weighted_mark'),
            lowest=Min('weighted_mark'),
            count=Count('id'),
        ).order_by('subject__name')

        # Overall grade distribution
        grade_dist = {}
        for r in results:
            g = r.overall_grade or 'N/A'
            grade_dist[g] = grade_dist.get(g, 0) + 1

        # Top 10 students
        top = results.order_by('class_rank')[:10]
        top_list = TermResultListSerializer(top, many=True).data

        data = {
            'class_session_id': session.id,
            'class_name': session.grade.name,
            'term_name': session.term.name,
            'total_students': results.count(),
            'subjects': list(subject_stats.values_list('subject__name', flat=True)),
            'overall_mean': float(results.aggregate(m=Avg('average_mark'))['m'] or 0),
            'subject_means': [
                {
                    'subject': s['subject__name'],
                    'code': s['subject__code'],
                    'mean': round(float(s['mean'] or 0), 1),
                    'highest': round(float(s['highest'] or 0), 1),
                    'lowest': round(float(s['lowest'] or 0), 1),
                    'entries': s['count'],
                }
                for s in subject_stats
            ],
            'grade_distribution': grade_dist,
            'top_students': top_list,
        }
        return Response(data)


# =============================================================================
# REPORT CARD TEMPLATE
# =============================================================================

class ReportCardTemplateViewSet(viewsets.ModelViewSet):
    queryset = ReportCardTemplate.objects.select_related(
        'curriculum', 'curriculum_level'
    ).filter(is_deleted=False)
    serializer_class = ReportCardTemplateSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None

    def get_queryset(self):
        qs = super().get_queryset()
        curriculum = self.request.query_params.get('curriculum')
        if curriculum:
            qs = qs.filter(curriculum_id=curriculum)
        return qs
