import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db import transaction
from django.conf import settings
from django.utils import timezone
from datetime import datetime, date as _date

# Models
from accounts.models import User, Student, Parent
from student_settings.models import (
    Intake, Curriculum, CurriculumLevel, GradeStructure, Stream, 
    AcademicYear, Term
)
from student_management.models.admission import Admission
from student_settings.models import Enrollment
from student_management.models.application import Application

class ExcelImportUtils:
    
    HEADERS = [
        # Field, Required, Label
        ('first_name', True, 'First Name'),
        ('last_name', True, 'Last Name'),
        ('gender', True, 'Gender (M/F)'),
        ('date_of_birth', True, 'Date of Birth (YYYY-MM-DD)'),
        ('admission_number', True, 'Admission Number'), # Can be auto? For bulk import usually manual/existing
        ('admission_date', True, 'Admission Date (YYYY-MM-DD)'),
        ('curriculum_code', True, 'Curriculum Code'), # e.g. CBC
        ('level_name', True, 'Curriculum Level'), # e.g. Primary
        ('grade_name', True, 'Class/Grade'), # e.g. Grade 1
        ('stream_name', False, 'Stream'), # e.g. North
        ('intake_name', True, 'Intake'), # e.g. Jan 2026
        ('academic_year', True, 'Year'), # e.g. 2026
        ('term_name', True, 'Term'), # e.g. Term 1
        ('guardian_name', False, 'Guardian Name'),
        ('guardian_phone', False, 'Guardian Phone'),
        ('guardian_email', False, 'Guardian Email'),
    ]

    @classmethod
    def generate_template(cls):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"

        # 1. Setup Headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        
        for col_idx, (field, required, label) in enumerate(cls.HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=label + (" *" if required else ""))
            cell.font = header_font
            cell.fill = header_fill
            # Set column width
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

        # 2. Add Validations (Dropdowns)
        # We need to fetch data for validations. 
        # Excel dropdowns strictly from a list or a range.
        # If list is too long (>255 chars), we must use a separate sheet.
        
        data_sheet = wb.create_sheet("ReferenceData")
        data_sheet.protection.sheet = True # Hide/Protect
        
        # Helper to write list column and create named range
        def create_dropdown(options, col_letter, name):
            if not options:
                return None
            for i, opt in enumerate(options, 1):
                data_sheet[f"{col_letter}{i}"] = opt
            # Define named range
            count = len(options)
            wb.create_named_range(name, data_sheet, f"${col_letter}$1:${col_letter}${count}")
            return f"={name}"

        # Fetch Options
        genders = ['M', 'F']
        curriculums = list(Curriculum.objects.values_list('code', flat=True))
        levels = list(CurriculumLevel.objects.values_list('name', flat=True))
        grades = list(GradeStructure.objects.values_list('name', flat=True))
        streams = list(Stream.objects.values_list('name', flat=True))
        intakes = list(Intake.objects.values_list('name', flat=True))
        years = list(AcademicYear.objects.values_list('name', flat=True))
        terms = list(Term.objects.values_list('name', flat=True))

        validations = {
            'C': create_dropdown(genders, 'A', 'GenderRange'),
            'G': create_dropdown(curriculums, 'B', 'CurrRange'),
            'H': create_dropdown(levels, 'C', 'LevelRange'),
            'I': create_dropdown(grades, 'D', 'GradeRange'),
            'J': create_dropdown(streams, 'E', 'StreamRange'),
            'K': create_dropdown(intakes, 'F', 'IntakeRange'),
            'L': create_dropdown(years, 'G', 'YearRange'),
            'M': create_dropdown(terms, 'H', 'TermRange'),
        }

        # Apply Validations to rows 2-1000
        for col_letter, formula in validations.items():
            if formula:
                dv = DataValidation(type="list", formula1=formula, allow_blank=True)
                ws.add_data_validation(dv)
                dv.add(f"{col_letter}2:{col_letter}1000")

        # 3. Add Sample Row
        # Optional: Add a sample row in grey?
        # ws.append(["John", "Doe", "M", "2015-01-01", "ADM001", ...])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Student_Import_Template.xlsx'
        wb.save(response)
        return response

    @classmethod
    def parse_and_validate(cls, file):
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active # Assume first sheet
        
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        results = {
            "success": False,
            "total_rows": len(rows),
            "valid_rows": [],
            "errors": [],
            "preview": []
        }
        
        # Pre-fetch caches to avoid N+1 queries during validation
        # Store dicts of {name: id} or {code: id}
        cache = {
            'curriculums': {c.code: c for c in Curriculum.objects.all()},
            'levels': {l.name: l for l in CurriculumLevel.objects.all()},
            'grades': {g.name: g for g in GradeStructure.objects.all()},
            'streams': {s.name: s for s in Stream.objects.all()},
            'streams_by_grade': {
                (s.name, s.grade_id): s
                for s in Stream.objects.select_related('grade').all()
            },
            'intakes': {i.name: i for i in Intake.objects.all()},
            'years': {str(y.name): y for y in AcademicYear.objects.all()}, # Cast to string just in case Excel sends int
            'terms': {},  # keyed by (term_name, year_name) — populated below
        }
        # Build term cache keyed by (term_name, year_name) to avoid cross-year collisions
        for t in Term.objects.select_related('academic_year').all():
            cache['terms'][(t.name, str(t.academic_year.name))] = t
        # Also keep a flat name→term map as fallback for rows missing year
        cache['terms_by_name'] = {t.name: t for t in Term.objects.all()}
        
        existing_adm_nos = set(Student.objects.values_list('admission_number', flat=True))

        for idx, row in enumerate(rows, 2): # Start from row 2 (1-based index)
            if not any(row): continue # Skip empty rows

            row_data = {}
            row_errors = []
            
            # Helper to safely get value
            def get_val(col_idx):
                if col_idx < len(row):
                    val = row[col_idx]
                    return str(val).strip() if val is not None else ""
                return ""

            # Mapping based on headers order
            # 0: First Name, 1: Last Name ...
            
            # --- 1. Basic Fields ---
            row_data['first_name'] = get_val(0)
            row_data['last_name'] = get_val(1)
            row_data['gender'] = get_val(2).upper()
            row_data['dob'] = get_val(3)
            row_data['adm_no'] = get_val(4)
            row_data['adm_date'] = get_val(5)
            
            # --- 2. Academic Links ---
            curr_code = get_val(6)
            level_name = get_val(7)
            grade_name = get_val(8)
            stream_name = get_val(9)
            intake_name = get_val(10)
            year_name = get_val(11).split('.')[0] # Handle 2026.0 from excel
            term_name = get_val(12)
            
            # --- 3. Guardian ---
            row_data['guardian_name'] = get_val(13)
            row_data['guardian_phone'] = get_val(14)
            row_data['guardian_email'] = get_val(15)

            # --- VALIDATION ---
            
            # Required Fields
            required_fields = ['first_name', 'last_name', 'gender', 'adm_no', 'curriculum_code', 'grade_name', 'intake_name', 'academic_year']
            # Map keys? 
            # Let's validate manually
            if not row_data['first_name']: row_errors.append("First Name is required")
            if not row_data['last_name']: row_errors.append("Last Name is required")
            if not row_data['adm_no']: 
                row_errors.append("Admission Number is required")
            elif row_data['adm_no'] in existing_adm_nos:
                row_errors.append(f"Admission Number {row_data['adm_no']} already exists")
            
            # Relations
            if curr_code in cache['curriculums']:
                row_data['curriculum'] = cache['curriculums'][curr_code]
            else:
                row_errors.append(f"Curriculum '{curr_code}' not found")

            if level_name:
                # Level needs to match curriculum? Not enforcing yet, but good to check
                if level_name in cache['levels']:
                    row_data['level'] = cache['levels'][level_name]
                else:
                    row_errors.append(f"Level '{level_name}' not found")
            
            if grade_name in cache['grades']:
                row_data['grade'] = cache['grades'][grade_name]
            else:
                row_errors.append(f"Grade '{grade_name}' not found")

            # Stream logic: Try to find stream by name AND grade if possible
            # Simplified: Just check if stream name exists, we might need a better UI for stream selection or validation
            if stream_name:
                if 'grade' in row_data:
                    stream_key = (stream_name, row_data['grade'].pk)
                    if stream_key in cache['streams_by_grade']:
                        row_data['stream'] = cache['streams_by_grade'][stream_key]
                    else:
                        row_errors.append(f"Stream '{stream_name}' not found in {grade_name}")
            else:
                row_data['stream'] = None

            if intake_name in cache['intakes']:
                row_data['intake'] = cache['intakes'][intake_name]
            else:
                row_errors.append(f"Intake '{intake_name}' not found")

            if year_name in cache['years']:
                row_data['year'] = cache['years'][year_name]
            else:
                row_errors.append(f"Academic Year '{year_name}' not found")

            # Look up term scoped to the academic year to avoid cross-year mismatch
            term_key = (term_name, year_name)
            if term_key in cache['terms']:
                row_data['term'] = cache['terms'][term_key]
            elif term_name in cache['terms_by_name']:
                row_data['term'] = cache['terms_by_name'][term_name]
            else:
                row_errors.append(f"Term '{term_name}' not found")
            
            # --- Date Parsing ---
            try:
                # Handle Excel format or string
                # openpyxl with data_only=True might return datetime object or string
                if isinstance(row[3], datetime):
                    row_data['dob_obj'] = row[3].date()
                elif row_data['dob']:
                     row_data['dob_obj'] = datetime.strptime(row_data['dob'], "%Y-%m-%d").date()
                else:
                    row_data['dob_obj'] = None
            except:
                row_errors.append("Invalid DOB format (Use YYYY-MM-DD)")

            try:
                if isinstance(row[5], datetime):
                    row_data['adm_date_obj'] = row[5].date()
                elif row_data['adm_date']:
                     row_data['adm_date_obj'] = datetime.strptime(row_data['adm_date'], "%Y-%m-%d").date()
                else:
                     row_data['adm_date_obj'] = timezone.now().date()
            except:
                row_errors.append("Invalid Admission Date format")


            # Store Result
            if row_errors:
                results['errors'].append({
                    'row': idx,
                    'errors': row_errors,
                    'data': row_data # Send back data for preview/correction context
                })
            else:
                # Store valid object for processing
                results['valid_rows'].append(row_data)

            # Add to preview (first 10)
            if idx < 12:
                preview_item = {k: str(v) for k,v in row_data.items() if not k.endswith('_obj') and not k in ['curriculum', 'level', 'grade', 'stream', 'intake', 'year', 'term']}
                preview_item['status'] = 'Error' if row_errors else 'Valid'
                results['preview'].append(preview_item)

        results['success'] = len(results['errors']) == 0
        return results

    @classmethod
    @transaction.atomic
    def process_import(cls, valid_data_list, user=None):
        """
        Bulk-import students using bulk_create for performance.
        With 900 rows this reduces ~5400 DB queries to ~7 batch inserts.
        """
        if not valid_data_list:
            return 0

        importing_user = user
        now_date = timezone.now().date()

        # ── 1. Pre-generate unique usernames in memory ───────────
        existing_usernames = set(
            User.objects.values_list('username', flat=True)
        )
        usernames = []
        for data in valid_data_list:
            base = f"{data['first_name'].lower()}.{data['last_name'].lower()}"
            username = base
            counter = 1
            while username in existing_usernames:
                username = f"{base}{counter}"
                counter += 1
            existing_usernames.add(username)
            usernames.append(username)

        # ── 2. Build & bulk-create User objects ──────────────────
        #   Students get their admission_number as initial password
        #   and should change it on first login.
        from django.contrib.auth.hashers import make_password
        # Pre-hash passwords with a dedup cache to avoid redundant bcrypt calls
        pw_cache = {}
        for data in valid_data_list:
            pwd = data['adm_no']
            if pwd not in pw_cache:
                pw_cache[pwd] = make_password(pwd)
        user_objs = []
        for i, data in enumerate(valid_data_list):
            u = User(
                username=usernames[i],
                email=data.get('guardian_email', ''),
                first_name=data['first_name'],
                last_name=data['last_name'],
                is_student=True,
                gender=data['gender'],
                password=pw_cache[data['adm_no']],
                is_first_login=True,
            )
            user_objs.append(u)
        created_users = User.objects.bulk_create(user_objs)

        # ── 3. Bulk-create Students ──────────────────────────────
        student_objs = [
            Student(
                student=created_users[i],
                admission_number=data['adm_no'],
                date_of_birth=data.get('dob_obj'),
                intake=data['intake'],
                admission_date=data.get('adm_date_obj'),
            )
            for i, data in enumerate(valid_data_list)
        ]
        created_students = Student.objects.bulk_create(student_objs)

        # ── 4. Bulk-create Applications ──────────────────────────
        app_objs = [
            Application(
                first_name=data['first_name'],
                last_name=data['last_name'],
                date_of_birth=data.get('dob_obj'),
                gender=data['gender'],
                intake=data['intake'],
                applying_for_curriculum=data['curriculum'],
                applying_for_level=data.get('level'),
                applying_for_grade=data['grade'],
                guardian_name=data.get('guardian_name'),
                email=data.get('guardian_email'),
                phone_number=data.get('guardian_phone'),
                application_status='accepted',
            )
            for data in valid_data_list
        ]
        created_apps = Application.objects.bulk_create(app_objs)

        # ── 5. Bulk-create Admissions ────────────────────────────
        Admission.objects.bulk_create([
            Admission(
                application=created_apps[i],
                student=created_students[i],
                admission_number=data['adm_no'],
                admission_date=data.get('adm_date_obj') or now_date,
            )
            for i, data in enumerate(valid_data_list)
        ])

        # ── 6. Bulk-create Enrollments (skips full_clean — data already validated) ──
        Enrollment.objects.bulk_create([
            Enrollment(
                student=created_students[i],
                intake=data.get('intake'),
                academic_year=data['year'],
                term=data['term'],
                curriculum=data['curriculum'],
                curriculum_level=data.get('level'),
                grade=data['grade'],
                stream=data.get('stream'),
                enrollment_type='new_admission',
                status='active',
                is_active=True,
                enrollment_date=data.get('adm_date_obj') or now_date,
                created_by=importing_user,
                updated_by=importing_user,
            )
            for i, data in enumerate(valid_data_list)
        ])

        # ── 7. Bulk-create Parents / Guardians ───────────────────
        guardian_rows = [
            (i, data) for i, data in enumerate(valid_data_list)
            if data.get('guardian_name')
        ]
        if guardian_rows:
            p_user_objs = []
            for i, data in guardian_rows:
                p_username = f"P{data['adm_no']}"
                counter = 1
                base_p = p_username
                while p_username in existing_usernames:
                    p_username = f"{base_p}_{counter}"
                    counter += 1
                existing_usernames.add(p_username)

                parts = data['guardian_name'].split()
                pu = User(
                    username=p_username,
                    email=data.get('guardian_email', ''),
                    first_name=parts[0],
                    last_name=' '.join(parts[1:]) if len(parts) > 1 else '',
                    is_parent=True,
                    password=make_password(p_username),
                    is_first_login=True,
                )
                p_user_objs.append(pu)

            created_p_users = User.objects.bulk_create(p_user_objs)

            Parent.objects.bulk_create([
                Parent(
                    user=created_p_users[idx],
                    student=created_students[i],
                    first_name=created_p_users[idx].first_name,
                    last_name=created_p_users[idx].last_name,
                    phone=data.get('guardian_phone', ''),
                    email=data.get('guardian_email', ''),
                )
                for idx, (i, data) in enumerate(guardian_rows)
            ])

        return len(valid_data_list)

    # ── Chunked import helpers ───────────────────────────────────

    @staticmethod
    def _serialize_valid_rows(valid_rows):
        """Convert validated rows (with ORM objects) to JSON-safe dicts for caching."""
        serialized = []
        for data in valid_rows:
            row = {
                'first_name': data['first_name'],
                'last_name': data['last_name'],
                'gender': data['gender'],
                'adm_no': data['adm_no'],
                'dob_obj': data.get('dob_obj').isoformat() if data.get('dob_obj') else None,
                'adm_date_obj': data.get('adm_date_obj').isoformat() if data.get('adm_date_obj') else None,
                'guardian_name': data.get('guardian_name', ''),
                'guardian_phone': data.get('guardian_phone', ''),
                'guardian_email': data.get('guardian_email', ''),
                'curriculum_id': data['curriculum'].pk,
                'level_id': data.get('level').pk if data.get('level') else None,
                'grade_id': data['grade'].pk,
                'stream_id': data.get('stream').pk if data.get('stream') else None,
                'intake_id': data['intake'].pk,
                'year_id': data['year'].pk,
                'term_id': data['term'].pk,
            }
            serialized.append(row)
        return serialized

    @classmethod
    @transaction.atomic
    def process_chunk(cls, chunk_data, importing_user=None):
        """
        Process a single chunk of serialized (ID-based) rows.
        Each chunk is its own atomic transaction — previous chunks survive if this fails.
        """
        if not chunk_data:
            return 0

        now_date = timezone.now().date()
        _to_date = lambda s: _date.fromisoformat(s) if s else None

        # ── 1. Generate unique usernames ─────────────────────────
        existing_usernames = set(
            User.objects.values_list('username', flat=True)
        )
        usernames = []
        for data in chunk_data:
            base = f"{data['first_name'].lower()}.{data['last_name'].lower()}"
            username = base
            counter = 1
            while username in existing_usernames:
                username = f"{base}{counter}"
                counter += 1
            existing_usernames.add(username)
            usernames.append(username)

        # ── 2. Hash passwords (dedup cache) ──────────────────────
        from django.contrib.auth.hashers import make_password
        pw_cache = {}
        for data in chunk_data:
            pwd = data['adm_no']
            if pwd not in pw_cache:
                pw_cache[pwd] = make_password(pwd)

        # ── 3. Create User objects ───────────────────────────────
        user_objs = [
            User(
                username=usernames[i],
                email=data.get('guardian_email', ''),
                first_name=data['first_name'],
                last_name=data['last_name'],
                is_student=True,
                gender=data['gender'],
                password=pw_cache[data['adm_no']],
                is_first_login=True,
            )
            for i, data in enumerate(chunk_data)
        ]
        created_users = User.objects.bulk_create(user_objs)

        # ── 4. Create Students ───────────────────────────────────
        student_objs = [
            Student(
                student=created_users[i],
                admission_number=data['adm_no'],
                date_of_birth=_to_date(data.get('dob_obj')),
                intake_id=data['intake_id'],
                admission_date=_to_date(data.get('adm_date_obj')),
            )
            for i, data in enumerate(chunk_data)
        ]
        created_students = Student.objects.bulk_create(student_objs)

        # ── 5. Create Applications ───────────────────────────────
        app_objs = [
            Application(
                first_name=data['first_name'],
                last_name=data['last_name'],
                date_of_birth=_to_date(data.get('dob_obj')),
                gender=data['gender'],
                intake_id=data['intake_id'],
                applying_for_curriculum_id=data['curriculum_id'],
                applying_for_level_id=data.get('level_id'),
                applying_for_grade_id=data['grade_id'],
                guardian_name=data.get('guardian_name'),
                email=data.get('guardian_email'),
                phone_number=data.get('guardian_phone'),
                application_status='accepted',
            )
            for data in chunk_data
        ]
        created_apps = Application.objects.bulk_create(app_objs)

        # ── 6. Create Admissions ─────────────────────────────────
        Admission.objects.bulk_create([
            Admission(
                application=created_apps[i],
                student=created_students[i],
                admission_number=data['adm_no'],
                admission_date=_to_date(data.get('adm_date_obj')) or now_date,
            )
            for i, data in enumerate(chunk_data)
        ])

        # ── 7. Create Enrollments ────────────────────────────────
        Enrollment.objects.bulk_create([
            Enrollment(
                student=created_students[i],
                intake_id=data.get('intake_id'),
                academic_year_id=data['year_id'],
                term_id=data['term_id'],
                curriculum_id=data['curriculum_id'],
                curriculum_level_id=data.get('level_id'),
                grade_id=data['grade_id'],
                stream_id=data.get('stream_id'),
                enrollment_type='new_admission',
                status='active',
                is_active=True,
                enrollment_date=_to_date(data.get('adm_date_obj')) or now_date,
                created_by=importing_user,
                updated_by=importing_user,
            )
            for i, data in enumerate(chunk_data)
        ])

        # ── 8. Create Parents ────────────────────────────────────
        guardian_rows = [
            (i, data) for i, data in enumerate(chunk_data)
            if data.get('guardian_name')
        ]
        if guardian_rows:
            p_user_objs = []
            for i, data in guardian_rows:
                p_username = f"P{data['adm_no']}"
                counter = 1
                base_p = p_username
                while p_username in existing_usernames:
                    p_username = f"{base_p}_{counter}"
                    counter += 1
                existing_usernames.add(p_username)

                parts = data['guardian_name'].split()
                pu = User(
                    username=p_username,
                    email=data.get('guardian_email', ''),
                    first_name=parts[0],
                    last_name=' '.join(parts[1:]) if len(parts) > 1 else '',
                    is_parent=True,
                    password=make_password(p_username),
                    is_first_login=True,
                )
                p_user_objs.append(pu)

            created_p_users = User.objects.bulk_create(p_user_objs)

            Parent.objects.bulk_create([
                Parent(
                    user=created_p_users[idx],
                    student=created_students[i],
                    first_name=created_p_users[idx].first_name,
                    last_name=created_p_users[idx].last_name,
                    phone=data.get('guardian_phone', ''),
                    email=data.get('guardian_email', ''),
                )
                for idx, (i, data) in enumerate(guardian_rows)
            ])

        return len(chunk_data)
