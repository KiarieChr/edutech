"""
Inventory Services Layer

Handles business logic for:
- Stock adjustments with GL posting
- GRN confirmation (stock receipt) with GL posting
- Supply issue processing with GL posting
- Stock take reconciliation
- Excel bulk upload

All GL operations follow the same JournalService pattern as payables.
"""

from django.db import transaction
from django.core.exceptions import ValidationError
from django.utils import timezone
from decimal import Decimal

from .models import (
    InventoryItem, StockMovement, Category,
    GoodsReceivedNote, GRNLine,
    SupplyIssue, SupplyIssueItem,
    StockTake, StockTakeLine
)
from finance.models import Account, AccountSubType, FinanceSettings
from journals.services import JournalService
from journals.models import JournalEntry, JournalLine


class InventoryService:
    """Core inventory operations with optional GL integration."""

    # =========================================================================
    # STOCK ADJUSTMENT
    # =========================================================================

    @staticmethod
    @transaction.atomic
    def adjust_stock(item, adjustment_type, quantity, reason, user, post_to_gl=False):
        """
        Adjust stock quantity with audit trail.

        Args:
            item: InventoryItem
            adjustment_type: 'INCREASE' or 'DECREASE'
            quantity: positive Decimal
            reason: str
            user: User
            post_to_gl: bool - whether to create journal entry

        Returns:
            StockMovement record
        """
        if quantity <= 0:
            raise ValidationError("Adjustment quantity must be positive.")

        if adjustment_type == 'INCREASE':
            item.stock_quantity += quantity
            movement_type = 'ADJUSTMENT_IN'
        elif adjustment_type == 'DECREASE':
            if item.stock_quantity < quantity:
                raise ValidationError(
                    f"Insufficient stock. Available: {item.stock_quantity}, requested: {quantity}"
                )
            item.stock_quantity -= quantity
            movement_type = 'ADJUSTMENT_OUT'
        else:
            raise ValidationError("adjustment_type must be 'INCREASE' or 'DECREASE'.")

        item.save()

        movement = StockMovement.objects.create(
            item=item,
            movement_type=movement_type,
            quantity=quantity,
            unit_cost=item.unit_cost,
            balance_after=item.stock_quantity,
            reference_number=f"ADJ-{timezone.now().strftime('%Y%m%d%H%M%S')}",
            reason=reason,
            created_by=user
        )

        # Optional GL posting
        if post_to_gl:
            journal_entry = InventoryService._post_adjustment_to_gl(
                item, movement_type, quantity, reason, user
            )
            movement.journal_entry = journal_entry
            movement.save()

        return movement

    @staticmethod
    def _post_adjustment_to_gl(item, movement_type, quantity, reason, user):
        """
        Post stock adjustment to GL.
        Adjustment In:  DR Inventory Asset, CR Inventory Surplus
        Adjustment Out: DR Inventory Shortage/Expense, CR Inventory Asset
        """
        asset_account = item.gl_asset_account
        expense_account = item.gl_expense_account or (
            item.category.gl_expense_account if item.category else None
        )

        if not asset_account:
            asset_account = Account.objects.filter(
                sub_type=AccountSubType.INVENTORY, is_active=True, children__isnull=True
            ).first()
        if not expense_account:
            expense_account = Account.objects.filter(
                type='EXPENSE', is_active=True, children__isnull=True
            ).first()

        if not asset_account or not expense_account:
            raise ValidationError("No inventory asset or expense GL account configured.")

        total = quantity * item.unit_cost

        entry = JournalEntry.objects.create(
            date=timezone.now().date(),
            description=f"Stock Adjustment: {item.name} - {reason}",
            journal_type='ADJUSTMENT',
            reference=f"SADJ-{item.code}",
            created_by=user
        )

        if movement_type == 'ADJUSTMENT_IN':
            JournalLine.objects.create(entry=entry, account=asset_account, debit=total, credit=Decimal('0.00'))
            JournalLine.objects.create(entry=entry, account=expense_account, debit=Decimal('0.00'), credit=total)
        else:
            JournalLine.objects.create(entry=entry, account=expense_account, debit=total, credit=Decimal('0.00'))
            JournalLine.objects.create(entry=entry, account=asset_account, debit=Decimal('0.00'), credit=total)

        JournalService.post_journal_entry(entry)
        return entry

    # =========================================================================
    # GRN CONFIRMATION (Stock Receipt)
    # =========================================================================

    @staticmethod
    @transaction.atomic
    def confirm_grn(grn, user):
        """
        Confirm a GRN — adds stock to items and creates movements.
        """
        if grn.status != 'DRAFT':
            raise ValidationError("Only DRAFT GRNs can be confirmed.")

        lines = grn.lines.select_related('item').all()
        if not lines.exists():
            raise ValidationError("GRN has no line items.")

        for line in lines:
            item = line.item
            item.stock_quantity += line.quantity_received
            item.unit_cost = line.unit_cost  # Update to latest cost
            item.save()

            StockMovement.objects.create(
                item=item,
                movement_type='RECEIPT',
                quantity=line.quantity_received,
                unit_cost=line.unit_cost,
                balance_after=item.stock_quantity,
                reference_number=grn.grn_number,
                reason=f"GRN from {grn.supplier.name}",
                created_by=user
            )

        grn.status = 'CONFIRMED'
        grn.save()
        return grn

    @staticmethod
    @transaction.atomic
    def post_grn_to_gl(grn, user):
        """
        Post confirmed GRN to GL.
        DR: Inventory Asset (per line)
        CR: Accounts Payable / GRN Clearing (total)
        """
        if grn.status != 'CONFIRMED':
            raise ValidationError("Only CONFIRMED GRNs can be posted to GL.")

        settings = FinanceSettings.load()
        ap_account = settings.default_payable_account
        if not ap_account:
            ap_account = Account.objects.filter(
                sub_type=AccountSubType.ACCOUNTS_PAYABLE, is_active=True, children__isnull=True
            ).first()
        if not ap_account:
            raise ValidationError("No Accounts Payable GL account configured.")

        lines = grn.lines.select_related('item', 'item__gl_asset_account').all()
        total = sum(line.total_cost for line in lines)

        entry = JournalEntry.objects.create(
            date=grn.received_date,
            description=f"GRN: {grn.grn_number} - {grn.supplier.name}",
            journal_type='PURCHASE',
            reference=grn.grn_number,
            created_by=user
        )

        for line in lines:
            asset_account = line.item.gl_asset_account
            if not asset_account:
                asset_account = Account.objects.filter(
                    sub_type=AccountSubType.INVENTORY, is_active=True, children__isnull=True
                ).first()
            if not asset_account:
                raise ValidationError(f"No inventory asset account for item: {line.item.name}")

            JournalLine.objects.create(
                entry=entry, account=asset_account,
                debit=line.total_cost, credit=Decimal('0.00'),
                description=f"{line.item.name} x {line.quantity_received}"
            )

        JournalLine.objects.create(
            entry=entry, account=ap_account,
            debit=Decimal('0.00'), credit=total,
            description=f"AP for {grn.grn_number}"
        )

        JournalService.post_journal_entry(entry)

        grn.journal_entry = entry
        grn.status = 'POSTED'
        grn.save()
        return grn

    # =========================================================================
    # SUPPLY ISSUE PROCESSING
    # =========================================================================

    @staticmethod
    @transaction.atomic
    def approve_issue(issue, user):
        """Approve a pending supply issue."""
        if issue.status != 'PENDING':
            raise ValidationError("Only PENDING issues can be approved.")

        if not issue.items.exists():
            raise ValidationError("Issue has no line items.")

        # Validate stock availability
        for line in issue.items.select_related('item').all():
            if line.item.stock_quantity < line.quantity_requested:
                raise ValidationError(
                    f"Insufficient stock for {line.item.name}. "
                    f"Available: {line.item.stock_quantity}, Requested: {line.quantity_requested}"
                )

        issue.status = 'APPROVED'
        issue.approved_by = user
        issue.save()
        return issue

    @staticmethod
    @transaction.atomic
    def process_issue(issue, user, post_to_gl=False):
        """
        Process an approved supply issue — deducts stock and creates movements.
        """
        if issue.status != 'APPROVED':
            raise ValidationError("Only APPROVED issues can be processed.")

        lines = issue.items.select_related('item').all()
        total_value = Decimal('0.00')

        for line in lines:
            item = line.item
            qty = line.quantity_issued if line.quantity_issued > 0 else line.quantity_requested

            if item.stock_quantity < qty:
                raise ValidationError(
                    f"Insufficient stock for {item.name}. "
                    f"Available: {item.stock_quantity}, Requested: {qty}"
                )

            item.stock_quantity -= qty
            item.save()

            line.quantity_issued = qty
            line.unit_cost = item.unit_cost
            line.save()

            total_value += line.total

            StockMovement.objects.create(
                item=item,
                movement_type='ISSUE',
                quantity=qty,
                unit_cost=item.unit_cost,
                balance_after=item.stock_quantity,
                reference_number=issue.issue_number,
                reason=f"Issued to {issue.department} - {issue.requested_by}",
                created_by=user
            )

        issue.total_value = total_value
        issue.status = 'ISSUED'
        issue.save()

        # Optional GL posting
        if post_to_gl:
            InventoryService._post_issue_to_gl(issue, user)

        return issue

    @staticmethod
    def _post_issue_to_gl(issue, user):
        """
        Post supply issue to GL.
        DR: Expense accounts (per item)
        CR: Inventory Asset (total)
        """
        lines = issue.items.select_related('item', 'item__gl_expense_account', 'item__gl_asset_account').all()
        total = sum(line.total for line in lines)

        asset_account = Account.objects.filter(
            sub_type=AccountSubType.INVENTORY, is_active=True, children__isnull=True
        ).first()
        if not asset_account:
            raise ValidationError("No inventory asset GL account configured.")

        entry = JournalEntry.objects.create(
            date=issue.issue_date,
            description=f"Supply Issue: {issue.issue_number} to {issue.department}",
            journal_type='GENERAL',
            reference=issue.issue_number,
            created_by=user
        )

        for line in lines:
            expense_account = (
                line.item.gl_expense_account or
                (line.item.category.gl_expense_account if line.item.category else None)
            )
            if not expense_account:
                expense_account = Account.objects.filter(
                    type='EXPENSE', is_active=True, children__isnull=True
                ).first()
            if not expense_account:
                raise ValidationError(f"No expense GL account for item: {line.item.name}")

            JournalLine.objects.create(
                entry=entry, account=expense_account,
                debit=line.total, credit=Decimal('0.00'),
                description=f"{line.item.name} x {line.quantity_issued}"
            )

        JournalLine.objects.create(
            entry=entry, account=asset_account,
            debit=Decimal('0.00'), credit=total,
            description=f"Stock issued: {issue.issue_number}"
        )

        JournalService.post_journal_entry(entry)

        issue.journal_entry = entry
        issue.save()

    # =========================================================================
    # STOCK TAKE RECONCILIATION
    # =========================================================================

    @staticmethod
    @transaction.atomic
    def reconcile_stock_take(stock_take, user, post_to_gl=False):
        """
        Apply stock take variances — adjusts stock and creates movements.
        """
        if stock_take.status != 'IN_PROGRESS':
            raise ValidationError("Only IN_PROGRESS stock takes can be reconciled.")

        lines = stock_take.lines.select_related('item').all()
        if not lines.exists():
            raise ValidationError("Stock take has no lines.")

        for line in lines:
            if line.variance == 0:
                continue

            item = line.item
            item.stock_quantity = line.physical_quantity
            item.save()

            movement_type = 'ADJUSTMENT_IN' if line.variance > 0 else 'ADJUSTMENT_OUT'
            StockMovement.objects.create(
                item=item,
                movement_type=movement_type,
                quantity=abs(line.variance),
                unit_cost=item.unit_cost,
                balance_after=item.stock_quantity,
                reference_number=stock_take.reference,
                reason=line.variance_reason or f"Stock take {stock_take.reference}",
                created_by=user
            )

        stock_take.status = 'COMPLETED'
        stock_take.save()

        if post_to_gl:
            InventoryService._post_stock_take_to_gl(stock_take, user)

        return stock_take

    @staticmethod
    def _post_stock_take_to_gl(stock_take, user):
        """Post stock take variances to GL as a single adjustment journal."""
        lines = stock_take.lines.select_related('item', 'item__gl_asset_account', 'item__gl_expense_account').all()
        variance_lines = [l for l in lines if l.variance != 0]

        if not variance_lines:
            return

        asset_account = Account.objects.filter(
            sub_type=AccountSubType.INVENTORY, is_active=True, children__isnull=True
        ).first()
        expense_account = Account.objects.filter(
            type='EXPENSE', is_active=True, children__isnull=True
        ).first()

        if not asset_account or not expense_account:
            raise ValidationError("No inventory asset or expense GL account configured.")

        entry = JournalEntry.objects.create(
            date=stock_take.date,
            description=f"Stock Take Reconciliation: {stock_take.reference}",
            journal_type='ADJUSTMENT',
            reference=stock_take.reference,
            created_by=user
        )

        for line in variance_lines:
            item_asset = line.item.gl_asset_account or asset_account
            item_expense = line.item.gl_expense_account or expense_account
            amount = abs(line.variance) * line.item.unit_cost

            if line.variance > 0:
                # Surplus: DR Inventory, CR Income/Surplus
                JournalLine.objects.create(entry=entry, account=item_asset, debit=amount, credit=Decimal('0.00'),
                                           description=f"Surplus: {line.item.name}")
                JournalLine.objects.create(entry=entry, account=item_expense, debit=Decimal('0.00'), credit=amount,
                                           description=f"Surplus: {line.item.name}")
            else:
                # Shortage: DR Expense/Shortage, CR Inventory
                JournalLine.objects.create(entry=entry, account=item_expense, debit=amount, credit=Decimal('0.00'),
                                           description=f"Shortage: {line.item.name}")
                JournalLine.objects.create(entry=entry, account=item_asset, debit=Decimal('0.00'), credit=amount,
                                           description=f"Shortage: {line.item.name}")

        JournalService.post_journal_entry(entry)

    # =========================================================================
    # EXCEL BULK UPLOAD
    # =========================================================================

    @staticmethod
    @transaction.atomic
    def upload_from_excel(file, user):
        """
        Bulk import inventory items from Excel file.

        Expected columns:
        code | name | category | item_type | unit_of_measure | unit_cost |
        stock_quantity | min_level | location | supplier_code | batch_number | expiry_date

        Returns:
            dict with 'created', 'errors', 'skipped' counts and details
        """
        try:
            import openpyxl
        except ImportError:
            raise ValidationError("openpyxl is required for Excel upload. Install it: pip install openpyxl")

        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if len(rows) < 2:
            raise ValidationError("Excel file must have a header row and at least one data row.")

        headers = [str(h).strip().lower().replace(' ', '_') if h else '' for h in rows[0]]
        required = {'name', 'category'}
        missing = required - set(headers)
        if missing:
            raise ValidationError(f"Missing required columns: {', '.join(missing)}")

        # Pre-fetch lookups
        from payables.models import Supplier
        categories = {c.name.lower(): c for c in Category.objects.all()}
        categories.update({c.code.lower(): c for c in Category.objects.all()})
        suppliers = {s.code.lower(): s for s in Supplier.objects.filter(is_active=True)}
        suppliers.update({s.name.lower(): s for s in Supplier.objects.filter(is_active=True)})

        valid_types = dict(InventoryItem.ITEM_TYPE_CHOICES)
        valid_units = dict(InventoryItem.UNIT_CHOICES)

        results = {'created': 0, 'updated': 0, 'errors': [], 'skipped': 0}

        for row_num, row in enumerate(rows[1:], start=2):
            row_data = dict(zip(headers, row))

            # Skip empty rows
            if not row_data.get('name'):
                results['skipped'] += 1
                continue

            try:
                name = str(row_data['name']).strip()
                cat_key = str(row_data.get('category', '')).strip().lower()

                # Resolve category (create if not exists)
                category = categories.get(cat_key)
                if not category:
                    cat_name = str(row_data['category']).strip()
                    category = Category.objects.create(
                        name=cat_name,
                        code=cat_name[:20].upper().replace(' ', '_')
                    )
                    categories[cat_key] = category

                # Resolve supplier
                supplier = None
                sup_key = str(row_data.get('supplier_code') or row_data.get('supplier', '') or '').strip().lower()
                if sup_key:
                    supplier = suppliers.get(sup_key)

                # Item type
                item_type = str(row_data.get('item_type', 'CONSUMABLE')).strip().upper()
                if item_type not in valid_types:
                    item_type = 'CONSUMABLE'

                # Unit
                unit = str(row_data.get('unit_of_measure') or row_data.get('unit', 'Pcs')).strip()
                if unit not in valid_units:
                    unit = 'Pcs'

                # Code
                code = str(row_data.get('code', '')).strip()

                # Check if item exists (update vs create)
                existing = None
                if code:
                    existing = InventoryItem.objects.filter(code=code).first()

                item_fields = {
                    'name': name,
                    'category': category,
                    'item_type': item_type,
                    'unit_of_measure': unit,
                    'unit_cost': Decimal(str(row_data.get('unit_cost', 0) or 0)),
                    'stock_quantity': Decimal(str(row_data.get('stock_quantity') or row_data.get('stock', 0) or 0)),
                    'min_level': Decimal(str(row_data.get('min_level', 0) or 0)),
                    'location': str(row_data.get('location', '') or '').strip() or None,
                    'supplier': supplier,
                    'batch_number': str(row_data.get('batch_number') or row_data.get('batch', '') or '').strip() or None,
                }

                # Parse expiry date
                expiry = row_data.get('expiry_date') or row_data.get('expiry')
                if expiry:
                    from datetime import date, datetime
                    if isinstance(expiry, (date, datetime)):
                        item_fields['expiry_date'] = expiry if isinstance(expiry, date) else expiry.date()
                    else:
                        try:
                            item_fields['expiry_date'] = datetime.strptime(str(expiry).strip(), '%Y-%m-%d').date()
                        except ValueError:
                            pass  # Skip invalid dates

                if existing:
                    for k, v in item_fields.items():
                        setattr(existing, k, v)
                    existing.save()
                    results['updated'] += 1
                else:
                    item = InventoryItem(**item_fields)
                    if code:
                        item.code = code
                    item.created_by = user
                    item.save()

                    # Create opening stock movement if stock > 0
                    if item.stock_quantity > 0:
                        StockMovement.objects.create(
                            item=item,
                            movement_type='OPENING',
                            quantity=item.stock_quantity,
                            unit_cost=item.unit_cost,
                            balance_after=item.stock_quantity,
                            reference_number='EXCEL-UPLOAD',
                            reason=f"Opening balance from Excel upload",
                            created_by=user
                        )
                    results['created'] += 1

            except Exception as e:
                results['errors'].append({
                    'row': row_num,
                    'name': str(row_data.get('name', '')),
                    'error': str(e)
                })

        return results

    @staticmethod
    def generate_upload_template():
        """Generate an Excel template for bulk upload."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory Items"

        headers = [
            'code', 'name', 'category', 'item_type', 'unit_of_measure',
            'unit_cost', 'stock_quantity', 'min_level', 'location',
            'supplier_code', 'batch_number', 'expiry_date'
        ]

        # Style headers
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2B5797', end_color='2B5797', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Sample data rows
        sample_data = [
            ['', 'A4 Paper Reams', 'Stationery', 'CONSUMABLE', 'Ream', 500, 100, 20, 'Store A-12', '', 'B-2026-001', ''],
            ['', 'Whiteboard Markers', 'Stationery', 'CONSUMABLE', 'Box', 1200, 50, 10, 'Store B-05', '', '', '2027-12-31'],
            ['', 'Office Projector', 'Electronics', 'CAPITAL_ASSET', 'Unit', 45000, 2, 1, 'Secure Store', '', 'SN-001', ''],
        ]

        for row_num, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border

        # Column widths
        widths = [12, 25, 15, 15, 15, 12, 15, 10, 15, 15, 15, 12]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        # Instructions sheet
        ws2 = wb.create_sheet("Instructions")
        instructions = [
            ["Column", "Required", "Description", "Valid Values"],
            ["code", "No", "Item code (auto-generated if blank)", "e.g. INV-00001"],
            ["name", "Yes", "Item name", ""],
            ["category", "Yes", "Category name (created if not found)", "e.g. Stationery, Cleaning"],
            ["item_type", "No", "Type of item (default: CONSUMABLE)", "CONSUMABLE, CAPITAL_ASSET, RAW_MATERIAL"],
            ["unit_of_measure", "No", "Unit (default: Pcs)", "Pcs, Box, Ream, Bale, Unit, Kg, Liters, etc."],
            ["unit_cost", "No", "Cost per unit", "Number"],
            ["stock_quantity", "No", "Current stock on hand", "Number"],
            ["min_level", "No", "Minimum stock level for alerts", "Number"],
            ["location", "No", "Storage location", "e.g. Store A-12"],
            ["supplier_code", "No", "Supplier code or name", "Must match existing supplier"],
            ["batch_number", "No", "Batch/serial number", ""],
            ["expiry_date", "No", "Expiry date", "YYYY-MM-DD format"],
        ]
        for row_num, row in enumerate(instructions, 1):
            for col, val in enumerate(row, 1):
                cell = ws2.cell(row=row_num, column=col, value=val)
                if row_num == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                cell.border = thin_border

        for col in range(1, 5):
            ws2.column_dimensions[chr(64 + col)].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
